from __future__ import annotations

import datetime as dt
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

from etapa2_lancar_evolucao_salus import (  # noqa: E402
    has_effective_discharge_at_end,
    parse_census_discharge,
    read_clinical,
    resolve_discharge,
    resolve_evolution_discharge,
    technical_discharge_time,
)


class AltaColunaTests(unittest.TestCase):
    def test_parse_census_discharge(self):
        self.assertEqual(
            parse_census_discharge(dt.datetime(2026, 7, 20, 15, 34)),
            ("20/07/2026", "15:34"),
        )
        self.assertEqual(
            parse_census_discharge("20/07/2026 15:34"),
            ("20/07/2026", "15:34"),
        )
        self.assertIsNone(parse_census_discharge("Internado"))
        self.assertIsNone(parse_census_discharge(dt.date(2026, 7, 20)))

    def test_column_is_fallback_and_evolution_has_priority(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Preenchimento"
        sheet.append(
            [
                "Nome paciente",
                "Iniciais",
                "Senha",
                "Dias internado",
                "ID internação",
                "evolucao",
                "Alta (data e hora)",
            ]
        )
        sheet.append(
            [
                "Paciente Censo",
                "PC",
                "ABC123",
                5,
                101,
                "Paciente estável, sem registro de alta no texto.",
                dt.datetime(2026, 7, 20, 15, 34),
            ]
        )
        sheet.append(
            [
                "Paciente Evolução",
                "PE",
                "XYZ789",
                3,
                102,
                "Paciente estável. Alta 21/07/2026 10:15.",
                dt.datetime(2026, 7, 20, 15, 34),
            ]
        )

        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "base.xlsx"
            workbook.save(path)
            patients, _meta, field_headers = read_clinical(path)

        census_values = patients["ABC123"][0].values
        self.assertEqual(census_values["Parecer do Auditor - Paciente permanece internado? *"], "Não")
        self.assertEqual(census_values["Parecer do Auditor - Data do desfecho * (cond.)"], "20/07/2026")
        self.assertEqual(census_values["Parecer do Auditor - Hora do desfecho * (cond.)"], "15:34")

        evolution_values = patients["XYZ789"][0].values
        self.assertEqual(evolution_values["Parecer do Auditor - Data do desfecho * (cond.)"], "21/07/2026")
        self.assertEqual(evolution_values["Parecer do Auditor - Hora do desfecho * (cond.)"], "10:15")
        self.assertNotIn("Alta (data e hora)", field_headers)

    def test_uses_technical_time_when_discharge_has_date_without_valid_time(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Preenchimento"
        sheet.append(["Nome paciente", "Senha", "evolucao"])
        sheet.append(["Paciente Sem Hora", "SEM123", "Conduta: alta dia 21/07/2026."])
        sheet.append(["Paciente Hora Invalida", "INV123", "Alta 21/07/2026 06:60."])

        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "base.xlsx"
            workbook.save(path)
            patients, _meta, _field_headers = read_clinical(path)

        for senha in ("SEM123", "INV123"):
            values = patients[senha][0].values
            self.assertEqual(values["Parecer do Auditor - Paciente permanece internado? *"], "Não")
            self.assertEqual(values["Parecer do Auditor - Data do desfecho * (cond.)"], "21/07/2026")
            hour = values["Parecer do Auditor - Hora do desfecho * (cond.)"]
            self.assertGreaterEqual(hour, "08:00")
            self.assertLessEqual(hour, "12:30")
            self.assertEqual(hour, technical_discharge_time(f"{senha}|21/07/2026"))

    def test_effective_discharge_without_date_uses_evolution_date(self):
        discharge = resolve_discharge(
            "Paciente estável. Conduta: alta hospitalar hoje.",
            evolution_date=dt.date(2026, 7, 23),
            seed="HOJE01",
        )
        self.assertIsNotNone(discharge)
        discharge_date, hour, outcome = discharge
        self.assertEqual(discharge_date, "23/07/2026")
        self.assertGreaterEqual(hour, "08:00")
        self.assertLessEqual(hour, "12:30")
        self.assertEqual(outcome, "Alta melhorada")

    def test_census_is_used_even_without_discharge_in_evolution(self):
        discharge = resolve_discharge(
            "Paciente permanece internado e estável.",
            census_value=dt.datetime(2026, 7, 23, 11, 42),
            evolution_date=dt.date(2026, 7, 23),
            seed="CENSO01",
        )
        self.assertEqual(discharge, ("23/07/2026", "11:42", "Alta melhorada"))

    def test_planned_or_internal_discharge_is_not_effective(self):
        examples = (
            "Previsão de alta amanhã.",
            "Paciente sem contraindicações para alta.",
            "Antes da alta, orientar familiares.",
            "Alta para UTI por instabilidade.",
            "Plano de alta após avaliação da equipe.",
        )
        for text in examples:
            with self.subTest(text=text):
                self.assertFalse(has_effective_discharge_at_end(text))
                self.assertIsNone(
                    resolve_discharge(
                        text,
                        evolution_date=dt.date(2026, 7, 23),
                        seed="NAOALTA",
                    )
                )

    def test_full_evolution_datetime_has_priority_over_census(self):
        discharge = resolve_discharge(
            "Conduta: alta 23/07/2026 às 10:18.",
            census_value=dt.datetime(2026, 7, 22, 9, 5),
            evolution_date=dt.date(2026, 7, 23),
            seed="PRIORIDADE",
        )
        self.assertEqual(discharge, ("23/07/2026", "10:18", "Alta melhorada"))

    def test_evolution_only_resolver_does_not_use_census(self):
        discharge = resolve_evolution_discharge(
            "Paciente estável. Conduta: alta hospitalar hoje.",
            evolution_date=dt.date(2026, 7, 23),
            seed="EVOLUCAO01",
        )
        self.assertIsNotNone(discharge)
        self.assertEqual(discharge[0], "23/07/2026")
        self.assertGreaterEqual(discharge[1], "08:00")
        self.assertLessEqual(discharge[1], "12:30")

    def test_read_clinical_preserves_explicit_adjusted_cid(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Preenchimento"
        sheet.append(
            [
                "Nome paciente",
                "Senha",
                "Dados da Internação - CID de internação *",
                "Dados da Internação - CID ajustado *",
            ]
        )
        sheet.append(["Paciente Com Mudança", "CID001", "I60.8", "I63.9"])
        sheet.append(["Paciente Sem Mudança", "CID002", "J18.9", None])

        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "base.xlsx"
            workbook.save(path)
            patients, _meta, _headers = read_clinical(path)

        changed = patients["CID001"][0].values
        unchanged = patients["CID002"][0].values
        self.assertEqual(changed["Dados da Internação - CID ajustado *"], "I63.9")
        self.assertEqual(unchanged["Dados da Internação - CID ajustado *"], "J18.9")


if __name__ == "__main__":
    unittest.main()
