from __future__ import annotations

import sys
import tempfile
import unittest
from copy import copy
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

from atualizar_novo_dia import generate_clinical_base  # noqa: E402
from etapa2_lancar_evolucao_salus import EVOLUTION_DISCHARGE_FILL  # noqa: E402


class AtualizarNovoDiaTests(unittest.TestCase):
    def test_reuses_evolution_dates_it_and_fills_derived_columns(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Preenchimento"
        headers = [
            "Nome",
            "Iniciais",
            "Senha",
            "Dias internado",
            "ID internação",
            "Dados da Internação - CID de internação *",
            "Dados da Internação - CID ajustado *",
            "Dados da Internação - Tempo de existência da doença *",
            "Dados da Internação - Nomenclatura do tempo de existência da doença *",
            "Exame Físico - Estado geral *",
            "Exame Físico - PA Sistólica max (mmHg) *",
            "Exame Físico - PA Diastólica max (mmHg) *",
            "Exame Físico - FC máx. (bpm) *",
            "Alta (data e hora)",
            "Parecer do Auditor - Paciente permanece internado? *",
            "Parecer do Auditor - Selecione o desfecho assistencial * (cond.)",
            "Parecer do Auditor - Data do desfecho * (cond.)",
            "Parecer do Auditor - Hora do desfecho * (cond.)",
            "Data da evolução",
            "evolucao",
        ]
        sheet.append(headers)
        sheet.append(
            [
                "Paciente Teste",
                "PT",
                "ABC123",
                4,
                99,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "22/07/2026 10:11",
                None,
                None,
                None,
                None,
                "21/07/2026",
                "CID J18.9\nBEG\nPA 120x80\nFC 88\nConduta: alta hospitalar hoje.",
            ]
        )

        patients = [
            {
                "nomeCompleto": "Paciente Teste",
                "nomeIniciais": "PT",
                "senha": "ABC123",
                "diasInternados": 5,
                "idInternacao": 99,
            }
        ]

        with tempfile.TemporaryDirectory() as temporary_directory:
            output = Path(temporary_directory) / "base.xlsx"
            stats = generate_clinical_base(
                patients,
                Path("modelo_nao_usado.xlsx"),
                output,
                previous_workbook=workbook,
                evolution_date=date(2026, 7, 22),
            )
            result = load_workbook(output)

        result_sheet = result["Preenchimento"]
        columns = {str(cell.value): cell.column for cell in result_sheet[1] if cell.value}
        self.assertEqual(
            result_sheet.cell(2, columns["evolucao"]).value,
            "CID J18.9\nBEG\nPA 120x80\nFC 88\nConduta: alta hospitalar hoje.",
        )
        self.assertEqual(result_sheet.cell(2, columns["Data da evolução"]).value, "22/07/2026")
        self.assertEqual(result_sheet.cell(2, columns["Dados da Internação - CID de internação *"]).value, "J18.9")
        self.assertEqual(result_sheet.cell(2, columns["Dados da Internação - CID ajustado *"]).value, "J18.9")
        self.assertEqual(result_sheet.cell(2, columns["Exame Físico - Estado geral *"]).value, "BEG – Bom Estado Geral")
        self.assertEqual(result_sheet.cell(2, columns["Exame Físico - PA Sistólica max (mmHg) *"]).value, 120)
        self.assertEqual(result_sheet.cell(2, columns["Exame Físico - PA Diastólica max (mmHg) *"]).value, 80)
        self.assertEqual(result_sheet.cell(2, columns["Exame Físico - FC máx. (bpm) *"]).value, 88)
        self.assertEqual(result_sheet.cell(2, columns["Dados da Internação - Tempo de existência da doença *"]).value, 5)
        self.assertEqual(result_sheet.cell(2, columns["Alta (data e hora)"]).value, "22/07/2026 10:11")
        self.assertEqual(
            result_sheet.cell(2, columns["Parecer do Auditor - Paciente permanece internado? *"]).value,
            "Não",
        )
        self.assertEqual(
            result_sheet.cell(2, columns["Parecer do Auditor - Data do desfecho * (cond.)"]).value,
            "22/07/2026",
        )
        self.assertEqual(
            result_sheet.cell(2, columns["Parecer do Auditor - Hora do desfecho * (cond.)"]).value,
            "10:11",
        )
        self.assertTrue(
            result_sheet.cell(2, columns["Alta (data e hora)"]).fill.fgColor.rgb.endswith(
                "FFC7CE"
            )
        )
        self.assertEqual(stats["evolucoes_reaproveitadas"], 1)
        self.assertEqual(stats["linhas_derivadas"], 1)
        self.assertEqual(result_sheet.max_row, 2)

    def test_evolution_discharge_fills_high_column_in_orange(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Preenchimento"
        sheet.append(
            [
                "Nome",
                "Iniciais",
                "Senha",
                "Dias internado",
                "ID internação",
                "Alta (data e hora)",
                "Parecer do Auditor - Paciente permanece internado? *",
                "Parecer do Auditor - Selecione o desfecho assistencial * (cond.)",
                "Parecer do Auditor - Data do desfecho * (cond.)",
                "Parecer do Auditor - Hora do desfecho * (cond.)",
                "Data da evolução",
                "evolucao",
            ]
        )
        sheet.append(
            [
                "Paciente Evolução",
                "PE",
                "EVOL01",
                2,
                321,
                "Internado",
                None,
                None,
                None,
                None,
                "22/07/2026",
                "Paciente estável. Conduta: alta hospitalar hoje.",
            ]
        )
        sheet.append(
            [
                "Paciente Evolução Anterior",
                "PEA",
                "EVOL02",
                2,
                322,
                datetime(2026, 7, 22, 9, 17),
                None,
                None,
                None,
                None,
                "22/07/2026",
                "Paciente estável. Conduta: alta hospitalar hoje.",
            ]
        )
        sheet.cell(3, 6).fill = PatternFill(
            "solid", fgColor=EVOLUTION_DISCHARGE_FILL
        )
        patients = [
            {
                "nomeCompleto": "Paciente Evolução",
                "nomeIniciais": "PE",
                "senha": "EVOL01",
                "diasInternados": 3,
                "idInternacao": 321,
            },
            {
                "nomeCompleto": "Paciente Evolução Anterior",
                "nomeIniciais": "PEA",
                "senha": "EVOL02",
                "diasInternados": 3,
                "idInternacao": 322,
            },
        ]

        with tempfile.TemporaryDirectory() as temporary_directory:
            output = Path(temporary_directory) / "base.xlsx"
            generate_clinical_base(
                patients,
                Path("modelo_nao_usado.xlsx"),
                output,
                previous_workbook=workbook,
                evolution_date=date(2026, 7, 23),
            )
            result = load_workbook(output)

        result_sheet = result["Preenchimento"]
        columns = {str(cell.value): cell.column for cell in result_sheet[1] if cell.value}
        high_cell = result_sheet.cell(2, columns["Alta (data e hora)"])
        self.assertEqual(high_cell.value.strftime("%d/%m/%Y"), "23/07/2026")
        self.assertGreaterEqual(high_cell.value.strftime("%H:%M"), "08:00")
        self.assertLessEqual(high_cell.value.strftime("%H:%M"), "12:30")
        self.assertTrue(high_cell.fill.fgColor.rgb.endswith(EVOLUTION_DISCHARGE_FILL))

        previous_high_cell = result_sheet.cell(3, columns["Alta (data e hora)"])
        self.assertEqual(
            previous_high_cell.value.strftime("%d/%m/%Y %H:%M"),
            "22/07/2026 09:17",
        )
        self.assertTrue(
            previous_high_cell.fill.fgColor.rgb.endswith(EVOLUTION_DISCHARGE_FILL)
        )
        self.assertEqual(result_sheet.max_row, 3)

    def test_census_discharge_fills_auditor_fields_without_evolution(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Preenchimento"
        sheet.append(
            [
                "Nome",
                "Iniciais",
                "Senha",
                "Dias internado",
                "ID internação",
                "Alta (data e hora)",
                "Parecer do Auditor - Paciente permanece internado? *",
                "Parecer do Auditor - Selecione o desfecho assistencial * (cond.)",
                "Parecer do Auditor - Data do desfecho * (cond.)",
                "Parecer do Auditor - Hora do desfecho * (cond.)",
                "evolucao",
            ]
        )
        sheet.append(
            [
                "Paciente Alta Censo",
                "PAC",
                "ALTA01",
                4,
                401,
                datetime(2026, 7, 23, 10, 12),
                None,
                None,
                None,
                None,
                None,
            ]
        )
        patients = [
            {
                "nomeCompleto": "Paciente Alta Censo",
                "nomeIniciais": "PAC",
                "senha": "ALTA01",
                "diasInternados": 4,
                "idInternacao": 401,
            }
        ]

        with tempfile.TemporaryDirectory() as temporary_directory:
            output = Path(temporary_directory) / "base.xlsx"
            generate_clinical_base(
                patients,
                Path("modelo_nao_usado.xlsx"),
                output,
                previous_workbook=workbook,
                evolution_date=date(2026, 7, 29),
            )
            result = load_workbook(output)

        result_sheet = result["Preenchimento"]
        columns = {str(cell.value): cell.column for cell in result_sheet[1] if cell.value}
        self.assertIsNone(result_sheet.cell(2, columns["evolucao"]).value)
        self.assertEqual(
            result_sheet.cell(
                2, columns["Parecer do Auditor - Paciente permanece internado? *"]
            ).value,
            "Não",
        )
        self.assertEqual(
            result_sheet.cell(
                2,
                columns[
                    "Parecer do Auditor - Selecione o desfecho assistencial * (cond.)"
                ],
            ).value,
            "Alta melhorada",
        )
        self.assertEqual(
            result_sheet.cell(
                2, columns["Parecer do Auditor - Data do desfecho * (cond.)"]
            ).value,
            "23/07/2026",
        )
        self.assertEqual(
            result_sheet.cell(
                2, columns["Parecer do Auditor - Hora do desfecho * (cond.)"]
            ).value,
            "10:12",
        )

    def test_lilac_name_is_preserved_only_for_same_admission(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Preenchimento"
        sheet.append(
            [
                "Nome paciente",
                "Iniciais",
                "Senha",
                "Dias internado",
                "ID internação",
                "Alta (data e hora)",
                "evolucao",
            ]
        )
        sheet.append(["Paciente Lilás", "PL", "LILAS01", 3, 501, None, None])
        sheet.append(["Paciente Neutro", "PN", "NEUTRO1", 2, 502, None, None])
        sheet.cell(2, 1).fill = PatternFill("solid", fgColor="E4DFEC")
        lilac_font = copy(sheet.cell(2, 1).font)
        lilac_font.color = "7030A0"
        lilac_font.bold = True
        sheet.cell(2, 1).font = lilac_font
        patients = [
            {
                "nomeCompleto": "Paciente Lilás",
                "nomeIniciais": "PL",
                "senha": "LILAS01",
                "diasInternados": 4,
                "idInternacao": 501,
            },
            {
                "nomeCompleto": "Paciente Neutro",
                "nomeIniciais": "PN",
                "senha": "NEUTRO1",
                "diasInternados": 3,
                "idInternacao": 502,
            },
            {
                "nomeCompleto": "Paciente Novo",
                "nomeIniciais": "PNO",
                "senha": "NOVO001",
                "diasInternados": 0,
                "idInternacao": 503,
            },
        ]

        with tempfile.TemporaryDirectory() as temporary_directory:
            output = Path(temporary_directory) / "base.xlsx"
            generate_clinical_base(
                patients,
                Path("modelo_nao_usado.xlsx"),
                output,
                previous_workbook=workbook,
                evolution_date=date(2026, 7, 30),
            )
            result = load_workbook(output)

        result_sheet = result["Preenchimento"]
        columns = {str(cell.value): cell.column for cell in result_sheet[1] if cell.value}
        name_column = columns["Nome paciente"]
        high_column = columns["Alta (data e hora)"]
        self.assertTrue(
            result_sheet.cell(2, name_column).fill.fgColor.rgb.endswith("E4DFEC")
        )
        self.assertFalse(
            str(result_sheet.cell(3, name_column).fill.fgColor.rgb or "").endswith(
                "E4DFEC"
            )
        )
        self.assertFalse(
            str(result_sheet.cell(4, name_column).fill.fgColor.rgb or "").endswith(
                "E4DFEC"
            )
        )
        self.assertIsNone(result_sheet.cell(4, high_column).fill.fill_type)

    def test_icu_evolution_overrides_reused_apartment_accommodation(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Preenchimento"
        sheet.append(
            [
                "Nome",
                "Iniciais",
                "Senha",
                "Dias internado",
                "ID internação",
                "Dados da Internação - Acomodação *",
                "Data da evolução",
                "evolucao",
            ]
        )
        sheet.append(
            [
                "Paciente UTI",
                "PU",
                "UTI001",
                4,
                701,
                "Apartamento / Enfermaria",
                "05/08/2026",
                "Paciente segue em UTI, em monitorização contínua.",
            ]
        )
        patients = [
            {
                "nomeCompleto": "Paciente UTI",
                "nomeIniciais": "PU",
                "senha": "UTI001",
                "diasInternados": 5,
                "idInternacao": 701,
            }
        ]

        with tempfile.TemporaryDirectory() as temporary_directory:
            output = Path(temporary_directory) / "base.xlsx"
            generate_clinical_base(
                patients,
                Path("modelo_nao_usado.xlsx"),
                output,
                previous_workbook=workbook,
                evolution_date=date(2026, 8, 6),
            )
            result = load_workbook(output)

        result_sheet = result["Preenchimento"]
        columns = {str(cell.value): cell.column for cell in result_sheet[1] if cell.value}
        self.assertEqual(
            result_sheet.cell(
                2, columns["Dados da Internação - Acomodação *"]
            ).value,
            "UTI",
        )


if __name__ == "__main__":
    unittest.main()
