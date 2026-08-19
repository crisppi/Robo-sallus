from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

from materializar_uti_semi import materialize, missing_uti_fields  # noqa: E402


HEADERS = [
    "Nome paciente",
    "Iniciais",
    "Senha",
    "Dias internado",
    "ID internação",
    "evolucao",
    "Dados da Internação - Acomodação *",
    "UTI - 1. Abertura Ocular (E) - Selecione a melhor resposta observada. *",
    "UTI - 2. Resposta Verbal (V) - Avaliar conteúdo da comunicação verbal. *",
    "UTI - 3. Melhor Resposta Motora (M) - Registrar a melhor resposta obtida. *",
    "UTI - 4. Resposta Pupilar (P) - Avaliar reatividade pupilar ao estímulo luminoso. *",
    "UTI - Monitorização *",
    "UTI - Tipo de monitorização * (cond.)",
    "UTI - Uso de droga vasoativa? *",
    "UTI - Drogas vasoativas em uso * (cond.)",
    "UTI - Creatinina sérica (mg/dL) *",
    "UTI - Não mensurado * (cond.)",
    "UTI - pH arterial *",
    "UTI - Não mensurado * (cond.) [2]",
    "UTI - PaO2 (mmHg) *",
    "UTI - Não mensurado * (cond.) [3]",
    "UTI - FiO2 (%) *",
    "UTI - Não mensurado * (cond.) [4]",
    "UTI - Categoria do diagnóstico principal *",
]


def save_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Preenchimento"
    sheet.append(HEADERS)
    sheet.append(
        [
            "Paciente UTI",
            "PU",
            "UTI001",
            3,
            1001,
            "Paciente em UTI por pneumonia, sem gasometria disponível.",
            "UTI",
        ]
    )
    sheet.append(
        [
            "Paciente Semi",
            "PS",
            "SEMI001",
            4,
            1002,
            "Paciente em semi por quadro cardiovascular.",
            "SEMI",
            "4 – Espontânea",
        ]
    )
    sheet.append(
        [
            "Paciente Apartamento",
            "PA",
            "APT001",
            1,
            1003,
            "Paciente estável.",
            "Apartamento / Enfermaria",
        ]
    )
    workbook.save(path)


class MaterializarUtiSemiTests(unittest.TestCase):
    def test_materializes_only_uti_and_semi_blank_fields(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "base.xlsx"
            save_workbook(path)

            result = materialize(path)

            self.assertEqual(result["linhas_uti_semi"], 2)
            self.assertEqual(result["linhas_alteradas"], 2)
            self.assertFalse(result["faltas"])

            workbook = load_workbook(path)
            sheet = workbook["Preenchimento"]
            headers = {str(cell.value): cell.column for cell in sheet[1] if cell.value}

            self.assertEqual(
                sheet.cell(2, headers["UTI - Monitorização *"]).value,
                "Não",
            )
            self.assertEqual(
                sheet.cell(2, headers["UTI - Categoria do diagnóstico principal *"]).value,
                "Respiratório",
            )
            self.assertEqual(
                sheet.cell(2, headers["UTI - Não mensurado * (cond.)"]).value,
                "Sim",
            )
            self.assertEqual(
                sheet.cell(3, headers["UTI - 1. Abertura Ocular (E) - Selecione a melhor resposta observada. *"]).value,
                "4 – Espontânea",
            )
            self.assertIsNone(sheet.cell(4, headers["UTI - Monitorização *"]).value)
            workbook.close()

    def test_dry_run_does_not_save(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "base.xlsx"
            save_workbook(path)

            result = materialize(path, dry_run=True)

            self.assertGreater(result["celulas_preenchidas"], 0)
            workbook = load_workbook(path)
            sheet = workbook["Preenchimento"]
            headers = {str(cell.value): cell.column for cell in sheet[1] if cell.value}
            self.assertIsNone(sheet.cell(2, headers["UTI - Monitorização *"]).value)
            self.assertTrue(missing_uti_fields(sheet, headers, 2))
            workbook.close()


if __name__ == "__main__":
    unittest.main()
