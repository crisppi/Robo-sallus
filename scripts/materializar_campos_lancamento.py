#!/usr/bin/env python3
"""Materializa campos operacionais obrigatorios antes do lancamento Salus."""

from __future__ import annotations

import argparse
import datetime as dt
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from etapa2_lancar_evolucao_salus import read_clinical, value_to_text
from lancar_evolucao_html_salus import filled_values


ROOT = Path(__file__).resolve().parents[1]
EXPORTS = ROOT / "exports"
ARCHIVE = EXPORTS / "arquivo"

FINALIZED_STATUSES = {
    "FINALIZADO",
    "SUCESSO",
    "SUCESSO_COM_ALERTA",
    "SUCESSO_MANUAL",
    "JA_LANCADO",
}

DEFAULT_VITALS = {
    "Exame Físico - PA Sistólica max (mmHg) *": 120,
    "Exame Físico - PA Diastólica max (mmHg) *": 80,
    "Exame Físico - FC máx. (bpm) *": 80,
    "Exame Físico - FR máx. (irpm) *": 16,
    "Exame Físico - SpO2 mín. (%) *": 97,
    "Exame Físico - Temperatura máx. (°C) *": 36.5,
}

MINIMUM_ADMISSION_FIELDS = [
    "Dados da Internação - Caráter da internação *",
    "Dados da Internação - Tipo da internação *",
    "Dados da Internação - Data da internação *",
    "Dados da Internação - CID de internação *",
    "Dados da Internação - CID ajustado *",
    "Dados da Internação - Tempo de existência da doença *",
    "Dados da Internação - Nomenclatura do tempo de existência da doença *",
]

OVERWRITE_TO_MATCH_RUNTIME = {
    "Conduta Clínica - Realizado procedimento cirúrgico? *",
    "Condição Adquirida - Paciente adquiriu alguma condição? *",
}


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def newest_base() -> Path:
    matches = sorted(
        EXPORTS.glob("data_base_lancar_*.xlsx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not matches:
        raise FileNotFoundError("Nenhuma base data_base_lancar_*.xlsx encontrada em exports.")
    return matches[0]


def create_backup(path: Path) -> Path:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    timestamp = dt.datetime.now().strftime("%H%M")
    backup = ARCHIVE / f"{path.stem}_antes_materializar_campos_lancamento_{timestamp}{path.suffix}"
    if backup.exists():
        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = ARCHIVE / f"{path.stem}_antes_materializar_campos_lancamento_{timestamp}{path.suffix}"
    shutil.copy2(path, backup)
    return backup


def has_launchable_content(values: dict[str, Any]) -> bool:
    return not is_blank(values.get("evolucao")) or not is_blank(values.get("Alta (data e hora)"))


def materialize(path: Path, *, dry_run: bool = False) -> dict[str, Any]:
    clinical_by_password, _, _ = read_clinical(path)
    clinical_by_row = {
        patient.row_number: patient
        for rows in clinical_by_password.values()
        if len(rows) == 1
        for patient in rows
    }

    workbook = load_workbook(path)
    sheet = workbook["Preenchimento"] if "Preenchimento" in workbook.sheetnames else workbook.active
    headers = {str(cell.value): cell.column for cell in sheet[1] if cell.value}

    rows_seen = 0
    rows_changed = 0
    cells_written = 0
    surgical_normalized = 0

    for row_number in range(2, sheet.max_row + 1):
        patient = clinical_by_row.get(row_number)
        if not patient:
            continue
        if not has_launchable_content(patient.values):
            continue
        status = value_to_text(patient.values.get("Lançamento Salus - Status")).upper()
        if status in FINALIZED_STATUSES:
            continue

        rows_seen += 1
        values = filled_values(patient)
        values.update(DEFAULT_VITALS)
        if (
            value_to_text(values.get("Exame Físico - Suporte respiratório *"))
            .lower()
            .startswith("suporte")
            and is_blank(values.get("Exame Físico - Frequência do suporte respiratório * (cond.)"))
        ):
            values["Exame Físico - Frequência do suporte respiratório * (cond.)"] = "Contínuo"
        if (
            value_to_text(values.get("Conduta Clínica - Uso de antibiótico? *"))
            .lower()
            .startswith("s")
            and is_blank(values.get("Conduta Clínica - Via do antibiótico * (cond.)"))
        ):
            values["Conduta Clínica - Via do antibiótico * (cond.)"] = "Via intravenosa"
        row_writes = 0

        for header, value in values.items():
            column = headers.get(header)
            if not column or is_blank(value):
                continue
            cell = sheet.cell(row_number, column)
            should_write = is_blank(cell.value)
            if header in OVERWRITE_TO_MATCH_RUNTIME and value_to_text(cell.value) != value_to_text(value):
                should_write = True
                if (
                    header == "Conduta Clínica - Realizado procedimento cirúrgico? *"
                    and value_to_text(cell.value).lower().startswith("s")
                    and value_to_text(value).lower().startswith("n")
                ):
                    surgical_normalized += 1
            if should_write:
                row_writes += 1
                cells_written += 1
                if not dry_run:
                    cell.value = value

        if row_writes:
            rows_changed += 1

        if is_blank(patient.values.get("evolucao")) and any(
            is_blank(values.get(field)) for field in MINIMUM_ADMISSION_FIELDS
        ):
            status_column = headers.get("Lançamento Salus - Status")
            date_column = headers.get("Lançamento Salus - Data/hora")
            message_column = headers.get("Lançamento Salus - Mensagem")
            if status_column and message_column:
                if not dry_run:
                    sheet.cell(row_number, status_column).value = "AGUARDANDO"
                    if date_column:
                        sheet.cell(row_number, date_column).value = dt.datetime.now().strftime(
                            "%d/%m/%Y %H:%M:%S"
                        )
                    sheet.cell(row_number, message_column).value = (
                        "Alta do censo sem evolucao/dados minimos de internacao; revisar antes do lancamento."
                    )
                cells_written += 2 if date_column else 1
                if not row_writes:
                    rows_changed += 1

    if not dry_run:
        workbook.save(path)
    workbook.close()

    return {
        "arquivo": path,
        "linhas_lancaveis": rows_seen,
        "linhas_alteradas": rows_changed,
        "celulas_preenchidas": cells_written,
        "cirurgicos_normalizados": surgical_normalized,
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Preenche defaults operacionais obrigatorios antes do lancamento Salus."
    )
    parser.add_argument("base", nargs="?", type=Path, default=None)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--sem-backup", action="store_true")
    args = parser.parse_args()

    base = args.base or newest_base()
    backup = None
    if not args.dry_run and not args.sem_backup:
        backup = create_backup(base)
    result = materialize(base, dry_run=args.dry_run)
    if backup:
        print(f"Backup: {backup}")
    print(f"Arquivo: {result['arquivo']}")
    print(f"Linhas lancaveis avaliadas: {result['linhas_lancaveis']}")
    print(f"Linhas alteradas: {result['linhas_alteradas']}")
    print(f"Celulas preenchidas: {result['celulas_preenchidas']}")
    print(f"Cirurgicos incompletos normalizados: {result['cirurgicos_normalizados']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
