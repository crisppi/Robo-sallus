#!/usr/bin/env python3
"""Sinaliza na base clinica os pacientes presentes no censo de altas."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import io
import re
import shutil
import subprocess
import unicodedata
from copy import copy
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


HIGH_COLUMN = "Alta (data e hora)"


@dataclass(frozen=True)
class Discharge:
    name: str
    discharged_at: dt.datetime


def normalize(value: object) -> str:
    text = str(value or "").strip().upper()
    text = "".join(
        character
        for character in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(character)
    )
    return " ".join(text.split())


def extract_discharges(pdf: Path) -> list[Discharge]:
    completed = subprocess.run(
        ["pdftotext", "-tsv", str(pdf), "-"],
        check=True,
        capture_output=True,
        text=True,
    )
    words = []
    for item in csv.DictReader(io.StringIO(completed.stdout), delimiter="\t"):
        if item.get("level") != "5" or not item.get("text"):
            continue
        words.append(
            {
                "page": int(item["page_num"]),
                "left": float(item["left"]),
                "top": float(item["top"]),
                "text": item["text"].strip(),
            }
        )

    births = sorted(
        [
            word
            for word in words
            if re.fullmatch(r"\d{2}/\d{2}/\d{4}", word["text"])
            and 80 <= word["left"] < 123
        ],
        key=lambda word: (word["page"], word["top"]),
    )
    discharges: list[Discharge] = []
    for index, birth in enumerate(births):
        next_top = birth["top"] + 19
        if index + 1 < len(births) and births[index + 1]["page"] == birth["page"]:
            next_top = births[index + 1]["top"] - 0.5

        name_words = sorted(
            [
                word
                for word in words
                if word["page"] == birth["page"]
                and word["left"] < 83
                and birth["top"] - 0.6 <= word["top"] < next_top
            ],
            key=lambda word: (word["top"], word["left"]),
        )
        discharge_date = next(
            (
                word["text"]
                for word in words
                if word["page"] == birth["page"]
                and abs(word["top"] - birth["top"]) < 0.7
                and 610 <= word["left"] < 650
                and re.fullmatch(r"\d{2}/\d{2}/\d{4}", word["text"])
            ),
            None,
        )
        discharge_time = next(
            (
                word["text"]
                for word in words
                if word["page"] == birth["page"]
                and abs(word["top"] - birth["top"]) < 0.7
                and 648 <= word["left"] < 678
                and re.fullmatch(r"\d{2}:\d{2}", word["text"])
            ),
            None,
        )
        name = " ".join(word["text"] for word in name_words).strip()
        if not name or not discharge_date or not discharge_time:
            continue
        discharges.append(
            Discharge(
                name=name,
                discharged_at=dt.datetime.strptime(
                    f"{discharge_date} {discharge_time}", "%d/%m/%Y %H:%M"
                ),
            )
        )
    return discharges


def extract_inpatient_names(pdf: Path) -> list[str]:
    completed = subprocess.run(
        ["pdftotext", "-tsv", str(pdf), "-"],
        check=True,
        capture_output=True,
        text=True,
    )
    words = []
    for item in csv.DictReader(io.StringIO(completed.stdout), delimiter="\t"):
        if item.get("level") != "5" or not item.get("text"):
            continue
        words.append(
            {
                "page": int(item["page_num"]),
                "left": float(item["left"]),
                "top": float(item["top"]),
                "text": item["text"].strip(),
            }
        )
    births = sorted(
        [
            word
            for word in words
            if re.fullmatch(r"\d{2}/\d{2}/\d{4}", word["text"])
            and 130 <= word["left"] < 174
        ],
        key=lambda word: (word["page"], word["top"]),
    )
    names = []
    for index, birth in enumerate(births):
        next_top = birth["top"] + 12.5
        if index + 1 < len(births) and births[index + 1]["page"] == birth["page"]:
            next_top = births[index + 1]["top"] - 0.5
        name_words = sorted(
            [
                word
                for word in words
                if word["page"] == birth["page"]
                and word["left"] < 133
                and birth["top"] - 0.6 <= word["top"] < next_top
            ],
            key=lambda word: (word["top"], word["left"]),
        )
        name = " ".join(word["text"] for word in name_words).strip()
        if name:
            names.append(name)
    return names


def match_name(name: str, candidates: dict[str, int]) -> tuple[int | None, str, float]:
    normalized = normalize(name)
    if normalized in candidates:
        return candidates[normalized], "exato", 1.0

    prefix_matches = [
        (candidate, row)
        for candidate, row in candidates.items()
        if len(normalized.split()) >= 2
        and (
            candidate.startswith(f"{normalized} ")
            or normalized.startswith(f"{candidate} ")
        )
    ]
    if len(prefix_matches) == 1:
        return prefix_matches[0][1], "prefixo unico", 0.99

    scores = sorted(
        (
            (SequenceMatcher(None, normalized, candidate).ratio(), candidate, row)
            for candidate, row in candidates.items()
        ),
        reverse=True,
    )
    if not scores:
        return None, "sem candidatos", 0.0
    best_score, _best_name, best_row = scores[0]
    second_score = scores[1][0] if len(scores) > 1 else 0.0
    if best_score >= 0.92 and best_score - second_score >= 0.04:
        return best_row, "grafia aproximada", best_score
    return None, "sem correspondencia segura", best_score


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("censo_altas", type=Path)
    parser.add_argument("base", type=Path)
    parser.add_argument("--censo-internados", type=Path)
    parser.add_argument(
        "--sinalizar-internados",
        action="store_true",
        help="Escreve 'Internado' na mesma coluna para os nomes presentes no censo de internados.",
    )
    parser.add_argument("--aplicar", action="store_true")
    parser.add_argument("--backup", type=Path)
    args = parser.parse_args()

    discharges = extract_discharges(args.censo_altas)
    inpatient_names = extract_inpatient_names(args.censo_internados) if args.censo_internados else []
    workbook = load_workbook(args.base)
    sheet = workbook["Preenchimento"]
    headers = {normalize(cell.value): cell.column for cell in sheet[1] if cell.value}
    name_column = (
        headers.get(normalize("Nome paciente"))
        or headers.get(normalize("Nome"))
        or headers.get(normalize("Paciente"))
    )
    evolution_column = headers.get(normalize("evolucao"))
    if not name_column or not evolution_column:
        raise RuntimeError("As colunas Nome/Paciente e evolucao sao obrigatorias.")

    candidates = {
        normalize(sheet.cell(row, name_column).value): row
        for row in range(2, sheet.max_row + 1)
        if normalize(sheet.cell(row, name_column).value)
    }
    matches = []
    unmatched = []
    for discharge in discharges:
        row, method, score = match_name(discharge.name, candidates)
        if row:
            matches.append((discharge, row, method, score))
        else:
            unmatched.append((discharge, score))

    inpatient_matches = []
    inpatient_unmatched = []
    if args.sinalizar_internados:
        if not args.censo_internados:
            raise RuntimeError("Use --censo-internados junto com --sinalizar-internados.")
        for inpatient_name in inpatient_names:
            row, method, score = match_name(inpatient_name, candidates)
            if row:
                inpatient_matches.append((inpatient_name, row, method, score))
            else:
                inpatient_unmatched.append((inpatient_name, score))

    print(f"Altas extraidas do PDF: {len(discharges)}")
    if args.censo_internados:
        inpatient_candidates = {normalize(name): index for index, name in enumerate(inpatient_names)}
        discharge_names_still_in_census = sum(
            match_name(discharge.name, inpatient_candidates)[0] is not None
            for discharge in discharges
        )
        print(f"Internados extraidos do outro PDF: {len(inpatient_names)}")
        print(f"Altas que tambem aparecem no censo de internados: {discharge_names_still_in_census}")
    print(f"Correspondencias seguras na base: {len(matches)}")
    print(f"Altas ausentes da base: {len(unmatched)}")
    if args.sinalizar_internados:
        print(f"Internados com correspondencia segura na base: {len(inpatient_matches)}")
        print(f"Internados ausentes da base: {len(inpatient_unmatched)}")
    for discharge, row, method, score in matches:
        base_name = str(sheet.cell(row, name_column).value or "").strip()
        print(
            f"SINALIZAR linha={row} base={base_name} pdf={discharge.name} "
            f"alta={discharge.discharged_at:%d/%m/%Y %H:%M} metodo={method} score={score:.3f}"
        )
    for inpatient_name, row, method, score in inpatient_matches:
        if method == "exato":
            continue
        base_name = str(sheet.cell(row, name_column).value or "").strip()
        print(
            f"INTERNADO_REVISADO linha={row} base={base_name} pdf={inpatient_name} "
            f"metodo={method} score={score:.3f}"
        )

    if not args.aplicar:
        print("Modo de conferencia: nenhum arquivo foi alterado.")
        return 0

    if args.backup:
        args.backup.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(args.base, args.backup)
        print(f"Backup: {args.backup}")

    high_column = headers.get(normalize(HIGH_COLUMN))
    if not high_column:
        high_column = evolution_column + 1
        sheet.insert_cols(high_column)
        source_header = sheet.cell(1, evolution_column)
        target_header = sheet.cell(1, high_column)
        target_header.value = HIGH_COLUMN
        target_header.font = copy(source_header.font)
        target_header.fill = copy(source_header.fill)
        target_header.border = copy(source_header.border)
        target_header.alignment = copy(source_header.alignment)
        target_header.protection = copy(source_header.protection)
        target_header.number_format = source_header.number_format
        sheet.column_dimensions[target_header.column_letter].width = 20
        for row in range(2, sheet.max_row + 1):
            source_cell = sheet.cell(row, evolution_column)
            target_cell = sheet.cell(row, high_column)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)

    for discharge, row, _method, _score in matches:
        cell = sheet.cell(row, high_column)
        cell.value = discharge.discharged_at
        cell.number_format = "dd/mm/yyyy hh:mm"
        cell.fill = PatternFill("solid", fgColor="FFC7CE")
        cell.font = Font(color="9C0006", bold=True)

    for _inpatient_name, row, _method, _score in inpatient_matches:
        cell = sheet.cell(row, high_column)
        if cell.value not in (None, ""):
            continue
        cell.value = "Internado"
        cell.number_format = "General"
        cell.fill = PatternFill("solid", fgColor="C6EFCE")
        cell.font = Font(color="006100", bold=True)

    workbook.save(args.base)
    print(f"Arquivo atualizado: {args.base}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
