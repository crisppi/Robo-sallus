#!/usr/bin/env python3
"""Etapa 2 do Robo Sallus: lancar evolucao clinica senha por senha.

Por seguranca, o script roda em modo dry-run por padrao. Nesse modo ele:
- le a fila oficial do Salus
- le a planilha clinica preenchida
- confere senha, nome e iniciais
- percorre campo a campo da linha do Excel
- valida listas, Sim/Nao, datas, numeros e multipla escolha
- gera relatorio de lancamentos

O lancamento real no Salus fica isolado na classe SalusExecutor. A regra de
percurso ja esta pronta; os seletores/API finais serao ajustados no teste real.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


IDENTITY_HEADERS = {
    "nome",
    "nome paciente",
    "paciente",
    "iniciais",
    "senha",
    "dias internado",
    "dias internados",
    "id internacao",
    "id internação",
    "alta (data e hora)",
}

SUCCESS_STATUSES = {"SUCESSO", "SUCESSO_COM_ALERTA", "SUCESSO_MANUAL"}


class AwaitingCidError(RuntimeError):
    pass


class AwaitingDischargeError(RuntimeError):
    pass
MULTIPLE_TYPES = {"LISTA_MULTIPLA"}
MULTIPLE_CONTROLS = {"CHECKBOX_MULTI", "MULTISELECT"}
YES_VALUES = {"sim", "s", "yes", "y", "true", "1"}
NO_VALUES = {"nao", "não", "n", "no", "false", "0"}


@dataclass
class QueuePatient:
    row_number: int
    senha: str
    nome: str = ""
    iniciais: str = ""
    dias_internado: str = ""


@dataclass
class ClinicalPatient:
    row_number: int
    senha: str
    nome: str = ""
    iniciais: str = ""
    dias_internado: str = ""
    id_internacao: str = ""
    values: dict[str, Any] = field(default_factory=dict)


@dataclass
class FieldMeta:
    coluna: str
    secao: str = ""
    campo: str = ""
    obrigatorio: str = ""
    tipo: str = ""
    controle: str = ""
    condicao: str = ""
    campo_pai: str = ""
    pergunta_id: str = ""
    opcoes: list[str] = field(default_factory=list)


@dataclass
class PreparedField:
    coluna: str
    valor_original: Any
    valor: str
    meta: FieldMeta
    status: str = "PENDENTE"
    mensagem: str = ""


@dataclass
class PatientResult:
    senha: str
    nome: str
    iniciais: str
    status: str
    mensagem: str
    campos_preenchidos: list[str] = field(default_factory=list)
    campos_ignorados: list[str] = field(default_factory=list)
    campos_com_erro: list[str] = field(default_factory=list)


def normalize(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text.casefold()


def header_key(value: Any) -> str:
    text = normalize(value)
    text = text.replace("*", "")
    text = text.replace("(cond.)", "")
    return text.strip()


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _outcome_tail(text: str) -> str:
    """Normaliza apenas o trecho final, onde o hospital registra o desfecho."""
    compact = re.sub(r"\s+", " ", str(text or "")).strip()
    compact = unicodedata.normalize("NFKD", compact)
    compact = "".join(ch for ch in compact if not unicodedata.combining(ch))
    return compact[-700:]


def _parse_outcome_date(day: str, month: str, year: str | None) -> dt.date | None:
    if not year:
        year = str(dt.date.today().year)
    elif len(year) == 2:
        year = "20" + year
    try:
        return dt.date(int(year), int(month), int(day))
    except ValueError:
        return None


def _is_planned_discharge_context(tail: str, start: int) -> bool:
    context = tail[max(0, start - 55):start].lower()
    return bool(
        re.search(
            r"(?:programacao|previsao|possibilidade)\s+de\s*$",
            context,
        )
    )


DISCHARGE_DATE_PATTERN = re.compile(
    r"\balta(?:\s+hospitalar|\s+para\s+casa)?"
    r"(?:\s*[.:;-]\s*)?(?:\s+(?:dia|em))?\s*"
    r"(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?",
    flags=re.IGNORECASE,
)

# O Salus exige hora para registrar o desfecho. Quando a evolucao confirma a
# alta e informa a data, mas nao traz uma hora valida, usamos um horario
# tecnico fixo para que o preenchimento seja reproduzivel e auditavel.
TECHNICAL_DISCHARGE_TIME = "12:00"


def extract_completed_discharge(text: str) -> tuple[str, str] | None:
    """Extrai alta efetiva no final, exigindo data e horário válidos."""
    tail = _outcome_tail(text)
    pattern = re.compile(
        DISCHARGE_DATE_PATTERN.pattern
        + r"\s*(?:[,.;:/-]|\b(?:as)\b)?\s*(\d{1,2}:\d{2})",
        flags=re.IGNORECASE,
    )
    valid: list[tuple[str, str]] = []
    for match in pattern.finditer(tail):
        if _is_planned_discharge_context(tail, match.start()):
            continue
        day, month, year, hour = match.groups()
        parsed_date = _parse_outcome_date(day, month, year)
        if parsed_date is None:
            continue
        try:
            parsed_time = dt.datetime.strptime(hour, "%H:%M").time()
        except ValueError:
            continue
        valid.append((parsed_date.strftime("%d/%m/%Y"), parsed_time.strftime("%H:%M")))
    return valid[-1] if valid else None


def extract_incomplete_discharge(text: str) -> tuple[str, str] | None:
    """Identifica alta efetiva com data, mas sem horário utilizável."""
    if extract_completed_discharge(text):
        return None
    tail = _outcome_tail(text)
    candidates: list[tuple[str, str]] = []
    for match in DISCHARGE_DATE_PATTERN.finditer(tail):
        if _is_planned_discharge_context(tail, match.start()):
            continue
        day, month, year = match.groups()
        parsed_date = _parse_outcome_date(day, month, year)
        if parsed_date is None:
            continue
        remainder = tail[match.end():match.end() + 24]
        time_match = re.match(r"\s*[,.;:/-]?\s*(\d{1,2}:\d{2})", remainder)
        invalid_time = time_match.group(1) if time_match else ""
        if invalid_time:
            try:
                dt.datetime.strptime(invalid_time, "%H:%M")
            except ValueError:
                pass
            else:
                continue
        candidates.append((parsed_date.strftime("%d/%m/%Y"), invalid_time))
    return candidates[-1] if candidates else None


def extract_completed_death(text: str) -> tuple[str, str] | None:
    """Extrai óbito quando o final informa alta por óbito e término do atendimento."""
    tail = _outcome_tail(text)
    death = re.search(r"\balta\s+por\s+obito\b", tail, flags=re.IGNORECASE)
    if not death:
        return None
    pattern = re.compile(
        r"data/hora\s+do\s+termino\s+do\s+atendimento\s*:\s*"
        r"(\d{1,2})/(\d{1,2})/(\d{2,4})\s+(\d{1,2}:\d{2})(?::\d{2})?",
        flags=re.IGNORECASE,
    )
    valid: list[tuple[str, str]] = []
    for match in pattern.finditer(tail[:death.end()]):
        day, month, year, hour = match.groups()
        parsed_date = _parse_outcome_date(day, month, year)
        if parsed_date is None:
            continue
        try:
            parsed_time = dt.datetime.strptime(hour, "%H:%M").time()
        except ValueError:
            continue
        valid.append((parsed_date.strftime("%d/%m/%Y"), parsed_time.strftime("%H:%M")))
    return valid[-1] if valid else None


def parse_census_discharge(value: Any) -> tuple[str, str] | None:
    """Converte a data/hora de alta do censo; textos como 'Internado' sao ignorados."""
    if isinstance(value, dt.datetime):
        return value.strftime("%d/%m/%Y"), value.strftime("%H:%M")
    if isinstance(value, dt.date):
        return None
    text = value_to_text(value)
    if not text or normalize(text) == "internado":
        return None
    for pattern in (
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
    ):
        try:
            parsed = dt.datetime.strptime(text, pattern)
        except ValueError:
            continue
        return parsed.strftime("%d/%m/%Y"), parsed.strftime("%H:%M")
    return None


def value_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, dt.date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_date_text(value: str) -> str | None:
    text = value.strip()
    if re.fullmatch(r"\d{2}/\d{2}/\d{4}", text):
        return text
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return dt.datetime.strptime(text[:19], fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return None


def find_column(headers: dict[str, int], *names: str) -> int | None:
    normalized_headers = {header_key(name): col for name, col in headers.items()}
    for name in names:
        col = normalized_headers.get(header_key(name))
        if col:
            return col
    return None


def worksheet_headers(ws) -> dict[str, int]:
    return {
        str(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(1, col).value not in (None, "")
    }


def require_column(headers: dict[str, int], *names: str) -> int:
    col = find_column(headers, *names)
    if not col:
        raise RuntimeError(f"Coluna obrigatoria nao encontrada: {', '.join(names)}")
    return col


def read_queue(path: Path) -> list[QueuePatient]:
    wb = load_workbook(path, data_only=True)
    ws = wb["Pacientes"] if "Pacientes" in wb.sheetnames else wb.active
    headers = worksheet_headers(ws)
    senha_col = require_column(headers, "Senha")
    nome_col = find_column(headers, "Nome", "Nome paciente", "Paciente")
    iniciais_col = find_column(headers, "Iniciais")
    dias_col = find_column(headers, "Dias internado", "Dias internados")

    patients: list[QueuePatient] = []
    for row in range(2, ws.max_row + 1):
        senha = value_to_text(ws.cell(row, senha_col).value)
        if not senha:
            continue
        patients.append(
            QueuePatient(
                row_number=row,
                senha=senha,
                nome=value_to_text(ws.cell(row, nome_col).value) if nome_col else "",
                iniciais=value_to_text(ws.cell(row, iniciais_col).value) if iniciais_col else "",
                dias_internado=value_to_text(ws.cell(row, dias_col).value) if dias_col else "",
            )
        )
    return patients


def read_field_meta(wb) -> dict[str, FieldMeta]:
    if "Campos" not in wb.sheetnames:
        return {}
    ws = wb["Campos"]
    headers = worksheet_headers(ws)
    col_coluna = require_column(headers, "Coluna")
    col_secao = find_column(headers, "Seção", "Secao")
    col_campo = find_column(headers, "Campo")
    col_obrigatorio = find_column(headers, "Obrigatório", "Obrigatorio")
    col_tipo = find_column(headers, "Tipo")
    col_controle = find_column(headers, "Controle")
    col_condicao = find_column(headers, "Condição/Quando preencher", "Condicao/Quando preencher")
    col_pai = find_column(headers, "Campo pai")
    col_id = find_column(headers, "ID formulário pergunta", "ID formulario pergunta")

    result: dict[str, FieldMeta] = {}
    for row in range(2, ws.max_row + 1):
        coluna = value_to_text(ws.cell(row, col_coluna).value)
        if not coluna:
            continue
        result[coluna] = FieldMeta(
            coluna=coluna,
            secao=value_to_text(ws.cell(row, col_secao).value) if col_secao else "",
            campo=value_to_text(ws.cell(row, col_campo).value) if col_campo else "",
            obrigatorio=value_to_text(ws.cell(row, col_obrigatorio).value) if col_obrigatorio else "",
            tipo=value_to_text(ws.cell(row, col_tipo).value).upper() if col_tipo else "",
            controle=value_to_text(ws.cell(row, col_controle).value).upper() if col_controle else "",
            condicao=value_to_text(ws.cell(row, col_condicao).value) if col_condicao else "",
            campo_pai=value_to_text(ws.cell(row, col_pai).value) if col_pai else "",
            pergunta_id=value_to_text(ws.cell(row, col_id).value) if col_id else "",
        )
    return result


def read_options(wb) -> dict[str, list[str]]:
    if "Opcoes" not in wb.sheetnames:
        return {}
    ws = wb["Opcoes"]
    headers = worksheet_headers(ws)
    campo_col = find_column(headers, "Campo")
    opcao_col = find_column(headers, "Opção", "Opcao")
    codigo_col = find_column(headers, "Código", "Codigo")
    if not campo_col or not opcao_col:
        return {}

    options: dict[str, list[str]] = {}
    for row in range(2, ws.max_row + 1):
        campo = value_to_text(ws.cell(row, campo_col).value)
        opcao = value_to_text(ws.cell(row, opcao_col).value)
        codigo = value_to_text(ws.cell(row, codigo_col).value) if codigo_col else ""
        if not campo or not opcao:
            continue
        if codigo == "API" or opcao.startswith("Lista pesquisável pela API"):
            continue
        options.setdefault(campo, [])
        if opcao not in options[campo]:
            options[campo].append(opcao)
    return options


def read_clinical(path: Path) -> tuple[dict[str, list[ClinicalPatient]], dict[str, FieldMeta], list[str]]:
    wb = load_workbook(path, data_only=True)
    if "Preenchimento" not in wb.sheetnames:
        raise RuntimeError("A planilha clinica precisa ter a aba 'Preenchimento'.")
    ws = wb["Preenchimento"]
    headers = worksheet_headers(ws)
    senha_col = require_column(headers, "Senha")
    nome_col = find_column(headers, "Nome", "Nome paciente", "Paciente")
    iniciais_col = find_column(headers, "Iniciais")
    dias_col = find_column(headers, "Dias internado", "Dias internados")
    id_col = find_column(headers, "ID internação", "ID internacao")
    census_discharge_col = find_column(headers, "Alta (data e hora)")

    meta = read_field_meta(wb)
    options = read_options(wb)
    for column_name, field_meta in meta.items():
        field_meta.opcoes = options.get(column_name, [])

    field_headers = [
        header
        for header in headers
        if header_key(header) not in IDENTITY_HEADERS
    ]

    by_password: dict[str, list[ClinicalPatient]] = {}
    for row in range(2, ws.max_row + 1):
        senha = value_to_text(ws.cell(row, senha_col).value)
        if not senha:
            continue
        values = {
            header: ws.cell(row, col).value
            for header, col in headers.items()
            if header in field_headers
        }
        evolution_text = value_to_text(values.get("evolucao")).lower()
        if any(term in evolution_text for term in ("dialise pausada", "diálise pausada", "pausado dialise", "pausada dialise")):
            therapies = value_to_text(values.get("Conduta Clínica - Terapias em andamento * (cond.)"))
            active_therapies = [
                item.strip()
                for item in therapies.split(";")
                if item.strip() and "renal substitutiva" not in item.lower()
            ]
            values["Conduta Clínica - Terapias em andamento * (cond.)"] = "; ".join(active_therapies)
            values["Conduta Clínica - Tipo de terapia renal substitutiva (TRS) * (cond.)"] = ""
            if not active_therapies:
                values["Conduta Clínica - Terapias Ativas (ex .: fisioterapia, suporte clínico) * *"] = "Não"
        admission_cid = values.get("Dados da Internação - CID de internação *")
        adjusted_cid = values.get("Dados da Internação - CID ajustado *")
        if not is_blank(admission_cid) and is_blank(adjusted_cid):
            # Regra operacional: se não houver mudança diagnóstica
            # explicitamente marcada na geração da base, repetir o CID inicial.
            values["Dados da Internação - CID ajustado *"] = admission_cid
        if is_blank(values.get("Dados da Internação - Comorbidades *")):
            values["Dados da Internação - Comorbidades *"] = "0SC - Sem comorbidades"
        if is_blank(values.get("Exame Físico - Estado geral *")):
            values["Exame Físico - Estado geral *"] = "REG – Regular Estado Geral"
        if is_blank(values.get("Exame Físico - Nível de consciência *")):
            values["Exame Físico - Nível de consciência *"] = "Orientado"
        if is_blank(values.get("Exame Físico - Mobilidade e dependência *")):
            values["Exame Físico - Mobilidade e dependência *"] = "Deambulando"
        if is_blank(values.get("Exame Físico - Acesso venoso? *")):
            values["Exame Físico - Acesso venoso? *"] = "Sim"
        if is_blank(values.get("Exame Físico - Qual o acesso venoso? * (cond.)")):
            values["Exame Físico - Qual o acesso venoso? * (cond.)"] = "Periférico"
        if is_blank(values.get("Exame Físico - Via respiratória *")):
            values["Exame Físico - Via respiratória *"] = "Normal"
        if is_blank(values.get("Exame Físico - Suporte respiratório *")):
            values["Exame Físico - Suporte respiratório *"] = "Ar ambiente"
        if is_blank(values.get("Exame Físico - Alimentação *")):
            values["Exame Físico - Alimentação *"] = "Oral"
        if is_blank(values.get("Exame Físico - Lesões na pele? *")):
            values["Exame Físico - Lesões na pele? *"] = "Não"
        if is_blank(values.get("Exame Físico - Controle de eliminação *")):
            values["Exame Físico - Controle de eliminação *"] = "Normal"
        # Regra operacional fixa para o bloco UTI.
        values["UTI - 1. Abertura Ocular (E) - Selecione a melhor resposta observada. *"] = (
            "NT – Não testável (ex.: olhos fechados por fator local)"
        )
        values["UTI - 2. Resposta Verbal (V) - Avaliar conteúdo da comunicação verbal. *"] = (
            "NT – Não testável (ex.: intubação, afasia)"
        )
        values["UTI - 3. Melhor Resposta Motora (M) - Registrar a melhor resposta obtida. *"] = (
            "NT – Não testável (fator limitante)"
        )
        values["UTI - 4. Resposta Pupilar (P) - Avaliar reatividade pupilar ao estímulo luminoso. *"] = (
            "0 – Reação bilateral ao estímulo"
        )
        for field in [
            "Conduta Clínica - Uso de antibiótico? *",
            "Conduta Clínica - Uso de antifúngico? *",
            "Conduta Clínica - Uso de antiviral? *",
            "Conduta Clínica - Câmara Hiperbárica * (cond.)",
            "Conduta Clínica - Administração de Imunoglobulina *",
            "Conduta Clínica - Terapias Ativas (ex .: fisioterapia, suporte clínico) * *",
            "Conduta Clínica - Realizado procedimento cirúrgico? *",
            "Condição Adquirida - Paciente adquiriu alguma condição? *",
        ]:
            if is_blank(values.get(field)):
                values[field] = "Não"
        # Condição adquirida nunca é inferida automaticamente. O valor "Sim"
        # somente é respeitado quando estiver explicitamente preenchido na base.
        if (
            "quimioterapia" in value_to_text(values.get("Conduta Clínica - Terapias em andamento * (cond.)")).lower()
            and is_blank(values.get("Conduta Clínica - Tipo de Quimioterapia * (cond.)"))
        ):
            values["Conduta Clínica - Tipo de Quimioterapia * (cond.)"] = "Curativa"
        if (
            "terapia renal substitutiva" in value_to_text(values.get("Conduta Clínica - Terapias em andamento * (cond.)")).lower()
            and is_blank(values.get("Conduta Clínica - Tipo de terapia renal substitutiva (TRS) * (cond.)"))
        ):
            values["Conduta Clínica - Tipo de terapia renal substitutiva (TRS) * (cond.)"] = "Hemodiálise"
        if (
            "radioterapia" in value_to_text(values.get("Conduta Clínica - Terapias em andamento * (cond.)")).lower()
            and is_blank(values.get("Conduta Clínica - Tipo de radioterapia * (cond.)"))
        ):
            values["Conduta Clínica - Tipo de radioterapia * (cond.)"] = "Convencional"

        if value_to_text(values.get("Conduta Clínica - Uso de antifúngico? *")).lower().startswith("s"):
            if is_blank(values.get("Conduta Clínica - Via do antifúngico * (cond.)")):
                values["Conduta Clínica - Via do antifúngico * (cond.)"] = "Via intravenosa"
        if value_to_text(values.get("Conduta Clínica - Uso de antiviral? *")).lower().startswith("s"):
            if is_blank(values.get("Conduta Clínica - Via do antiviral * (cond.)")):
                values["Conduta Clínica - Via do antiviral * (cond.)"] = "Via intravenosa"

        surgery_yn = value_to_text(
            values.get("Conduta Clínica - Realizado procedimento cirúrgico? *")
        ).lower()
        if surgery_yn.startswith("s"):
            required_surgery_fields = [
                "Conduta Clínica - TUSS + Nome do Procedimento * (cond.)",
                "Conduta Clínica - Quantidade Solicitada * (cond.)",
                "Conduta Clínica - Quantidade Autorizada * (cond.)",
                "Conduta Clínica - Quantidade Realizada * (cond.)",
                "Conduta Clínica - Tipo de anestesia * (cond.)",
            ]
            if any(is_blank(values.get(field)) for field in required_surgery_fields):
                values["Conduta Clínica - Realizado procedimento cirúrgico? *"] = "Não"
                for header in list(values):
                    if header.startswith((
                        "Conduta Clínica - TUSS +",
                        "Conduta Clínica - Quantidade",
                        "Conduta Clínica - Anvisa +",
                        "Conduta Clínica - Tipo de anestesia",
                        "Conduta Clínica - Houve intercorrências",
                        "Conduta Clínica - Tipo de intercorrência",
                        "Conduta Clínica - Justifique",
                    )):
                        values[header] = ""

        # Regra operacional fixa: a automação nunca marca condição adquirida.
        # Qualquer exceção deve ser lançada manualmente fora do robô.
        values["Condição Adquirida - Paciente adquiriu alguma condição? *"] = "Não"
        for header in list(values):
            if header.startswith("Condição Adquirida - ") and header != "Condição Adquirida - Paciente adquiriu alguma condição? *":
                values[header] = ""

        auditor_defaults = {
            "Parecer do Auditor - Pertinência Técnica da Internação *": "Sim",
            "Parecer do Auditor - Pertinência Técnica da permanência hospitalar *": "Sim",
            "Parecer do Auditor - Paciente permanece internado? *": "Sim",
            "Parecer do Auditor - Programação de alta * (cond.)": "Sem programação de alta",
        }
        for field, default in auditor_defaults.items():
            if is_blank(values.get(field)):
                values[field] = default

        death = extract_completed_death(evolution_text)
        discharge = extract_completed_discharge(evolution_text)
        incomplete_discharge = extract_incomplete_discharge(evolution_text)
        census_discharge = (
            parse_census_discharge(ws.cell(row, census_discharge_col).value)
            if census_discharge_col
            else None
        )
        if death:
            discharge_date, hour = death
            values["Parecer do Auditor - Paciente permanece internado? *"] = "Não"
            values["Parecer do Auditor - Selecione o desfecho assistencial * (cond.)"] = "Óbito"
            values["Parecer do Auditor - Data do desfecho * (cond.)"] = discharge_date
            values["Parecer do Auditor - Hora do desfecho * (cond.)"] = hour
        elif discharge:
            discharge_date, hour = discharge
            values["Parecer do Auditor - Paciente permanece internado? *"] = "Não"
            values["Parecer do Auditor - Selecione o desfecho assistencial * (cond.)"] = "Alta melhorada"
            values["Parecer do Auditor - Data do desfecho * (cond.)"] = discharge_date
            values["Parecer do Auditor - Hora do desfecho * (cond.)"] = hour
        elif census_discharge:
            discharge_date, hour = census_discharge
            values["Parecer do Auditor - Paciente permanece internado? *"] = "Não"
            values["Parecer do Auditor - Selecione o desfecho assistencial * (cond.)"] = "Alta melhorada"
            values["Parecer do Auditor - Data do desfecho * (cond.)"] = discharge_date
            values["Parecer do Auditor - Hora do desfecho * (cond.)"] = hour
        elif incomplete_discharge:
            discharge_date, _invalid_hour = incomplete_discharge
            values["Parecer do Auditor - Paciente permanece internado? *"] = "Não"
            values["Parecer do Auditor - Selecione o desfecho assistencial * (cond.)"] = "Alta melhorada"
            values["Parecer do Auditor - Data do desfecho * (cond.)"] = discharge_date
            values["Parecer do Auditor - Hora do desfecho * (cond.)"] = TECHNICAL_DISCHARGE_TIME
        elif value_to_text(values.get("Parecer do Auditor - Paciente permanece internado? *")).lower().startswith("n"):
            required_discharge = [
                "Parecer do Auditor - Selecione o desfecho assistencial * (cond.)",
                "Parecer do Auditor - Data do desfecho * (cond.)",
                "Parecer do Auditor - Hora do desfecho * (cond.)",
            ]
            if any(is_blank(values.get(field)) for field in required_discharge):
                values["Parecer do Auditor - Paciente permanece internado? *"] = "Sim"
                for field in required_discharge:
                    values[field] = ""
        patient = ClinicalPatient(
            row_number=row,
            senha=senha,
            nome=value_to_text(ws.cell(row, nome_col).value) if nome_col else "",
            iniciais=value_to_text(ws.cell(row, iniciais_col).value) if iniciais_col else "",
            dias_internado=value_to_text(ws.cell(row, dias_col).value) if dias_col else "",
            id_internacao=value_to_text(ws.cell(row, id_col).value) if id_col else "",
            values=values,
        )
        by_password.setdefault(senha, []).append(patient)
    return by_password, meta, field_headers


def read_successful_passwords(report_path: Path) -> set[str]:
    if not report_path.exists():
        return set()
    wb = load_workbook(report_path, data_only=True)
    ws = wb.active
    headers = worksheet_headers(ws)
    senha_col = find_column(headers, "Senha")
    status_col = find_column(headers, "Status")
    if not senha_col or not status_col:
        return set()
    successful: set[str] = set()
    for row in range(2, ws.max_row + 1):
        senha = value_to_text(ws.cell(row, senha_col).value)
        status = value_to_text(ws.cell(row, status_col).value)
        if senha and status in SUCCESS_STATUSES:
            successful.add(senha)
    return successful


def equivalent(a: str, b: str) -> bool:
    return not a or not b or normalize(a) == normalize(b)


def canonical_option(value: str, options: list[str]) -> str | None:
    value_norm = normalize(value)
    for option in options:
        if normalize(option) == value_norm:
            return option
    return None


def is_multiple_field(meta: FieldMeta) -> bool:
    return meta.tipo in MULTIPLE_TYPES or meta.controle in MULTIPLE_CONTROLS


def prepare_field(column: str, raw_value: Any, meta: FieldMeta | None) -> PreparedField:
    field_meta = meta or FieldMeta(coluna=column)
    value = value_to_text(raw_value)
    prepared = PreparedField(coluna=column, valor_original=raw_value, valor=value, meta=field_meta)

    if not field_meta.tipo and not field_meta.controle and not field_meta.opcoes:
        prepared.status = "OK"
        return prepared

    if field_meta.tipo == "DATA":
        if isinstance(raw_value, (dt.date, dt.datetime)):
            prepared.valor = value_to_text(raw_value)
            prepared.status = "OK"
            return prepared
        normalized_date = normalize_date_text(value)
        if normalized_date:
            prepared.valor = normalized_date
            prepared.status = "OK"
            return prepared
        prepared.status = "ERRO"
        prepared.mensagem = "Data deve estar no formato dd/mm/aaaa."
        return prepared

    if field_meta.tipo in {"NUMERO", "NÚMERO", "INTEGER", "DECIMAL"}:
        try:
            float(str(value).replace(".", "").replace(",", "."))
        except ValueError:
            prepared.status = "ERRO"
            prepared.mensagem = "Valor numerico invalido."
            return prepared
        prepared.status = "OK"
        return prepared

    if field_meta.tipo == "BOOLEANO":
        normalized = normalize(value)
        if normalized in YES_VALUES:
            prepared.valor = "Sim"
            prepared.status = "OK"
            return prepared
        if normalized in NO_VALUES:
            prepared.valor = "Não"
            prepared.status = "OK"
            return prepared
        prepared.status = "ERRO"
        prepared.mensagem = "Campo Sim/Nao aceita apenas Sim ou Nao."
        return prepared

    if field_meta.opcoes and is_multiple_field(field_meta):
        chosen = [part.strip() for part in value.split(";") if part.strip()]
        invalid = [part for part in chosen if canonical_option(part, field_meta.opcoes) is None]
        if invalid:
            prepared.status = "ERRO"
            prepared.mensagem = "Opcao multipla invalida: " + "; ".join(invalid)
            return prepared
        prepared.valor = "; ".join(canonical_option(part, field_meta.opcoes) or part for part in chosen)
        prepared.status = "OK"
        return prepared

    if field_meta.opcoes:
        if ";" in value:
            prepared.status = "ERRO"
            prepared.mensagem = "Campo de escolha unica recebeu mais de uma opcao."
            return prepared
        canonical = canonical_option(value, field_meta.opcoes)
        if canonical is None:
            prepared.status = "ERRO"
            prepared.mensagem = "Opcao invalida para lista: " + value
            return prepared
        prepared.valor = canonical
        prepared.status = "OK"
        return prepared

    prepared.status = "OK"
    return prepared


class DryRunExecutor:
    def lancar(self, queue_patient: QueuePatient, clinical_patient: ClinicalPatient, fields: list[PreparedField]) -> list[PreparedField]:
        for field in fields:
            if field.status == "PENDENTE":
                field.status = "OK"
                field.mensagem = "Simulado; nada foi lancado no Salus."
        return fields


class SalusExecutor:
    """Executor real do Salus.

    A regra de negocio ja chama esta classe com paciente e campos prontos.
    O preenchimento real usa a tela HTML do Salus via Chrome DevTools remoto,
    clicando e digitando nos campos como um usuario.
    """

    def __init__(self, usar_defaults_obrigatorios: bool = False, confirmar: bool = True) -> None:
        self.usar_defaults_obrigatorios = usar_defaults_obrigatorios
        self.confirmar = confirmar

    def lancar(self, queue_patient: QueuePatient, clinical_patient: ClinicalPatient, fields: list[PreparedField]) -> list[PreparedField]:
        from lancar_evolucao_html_salus import run_html_fill

        remains_admitted = value_to_text(
            clinical_patient.values.get("Parecer do Auditor - Paciente permanece internado? *")
        ).lower()
        if remains_admitted.startswith("n"):
            required_discharge = [
                "Parecer do Auditor - Selecione o desfecho assistencial * (cond.)",
                "Parecer do Auditor - Data do desfecho * (cond.)",
                "Parecer do Auditor - Hora do desfecho * (cond.)",
            ]
            missing = [
                field for field in required_discharge
                if is_blank(clinical_patient.values.get(field))
            ]
            if missing:
                raise AwaitingDischargeError(
                    "Alta identificada no final da evolução, mas falta data ou horário válido do desfecho."
                )

        result = run_html_fill(
            clinical_patient,
            confirmar=self.confirmar,
            usar_defaults_obrigatorios=self.usar_defaults_obrigatorios,
        )
        if result.get("missingCid"):
            raise AwaitingCidError(
                "Demais páginas preenchidas; falta apenas definir o CID antes da confirmação."
            )
        if not result.get("confirmEnabled"):
            raise RuntimeError("Tela HTML preenchida, mas Confirmar evolucao permaneceu desabilitado.")
        for field in fields:
            if field.status == "PENDENTE":
                field.status = "OK"
                field.mensagem = "Preenchido pela tela HTML do Salus."
        return fields


def build_patient_result(
    queue_patient: QueuePatient,
    clinical_patient: ClinicalPatient,
    prepared_fields: list[PreparedField],
    ignored_fields: list[str],
    alerts: list[str],
    dry_run: bool,
) -> PatientResult:
    error_fields = [
        f"{field.coluna}: {field.mensagem or 'erro'}"
        for field in prepared_fields
        if field.status == "ERRO"
    ]
    filled_fields = [
        field.coluna
        for field in prepared_fields
        if field.status == "OK"
    ]

    if error_fields:
        status = "ERRO_VALIDACAO" if dry_run else "ERRO"
    elif dry_run:
        status = "DRY_RUN_COM_ALERTA" if alerts else "DRY_RUN"
    else:
        status = "SUCESSO_COM_ALERTA" if alerts else "SUCESSO"

    messages = []
    if dry_run and not error_fields:
        messages.append("Simulacao concluida; nenhum dado lancado no Salus.")
    if alerts:
        messages.extend(alerts)
    if error_fields:
        messages.append("Existem campos com erro; paciente nao deve ser lancado antes de corrigir.")

    return PatientResult(
        senha=queue_patient.senha,
        nome=queue_patient.nome or clinical_patient.nome,
        iniciais=queue_patient.iniciais or clinical_patient.iniciais,
        status=status,
        mensagem=" | ".join(messages),
        campos_preenchidos=filled_fields,
        campos_ignorados=ignored_fields,
        campos_com_erro=error_fields,
    )


def process_patients(
    queue_patients: list[QueuePatient],
    clinical_by_password: dict[str, list[ClinicalPatient]],
    field_meta: dict[str, FieldMeta],
    field_headers: list[str],
    successful_passwords: set[str],
    dry_run: bool,
    only_password: str | None = None,
    limit: int | None = None,
    usar_defaults_obrigatorios: bool = False,
    stop_on_error: bool = False,
    confirmar: bool = True,
    progress_callback: Callable[[str, QueuePatient, PatientResult | None], None] | None = None,
) -> list[PatientResult]:
    executor = DryRunExecutor() if dry_run else SalusExecutor(
        usar_defaults_obrigatorios=usar_defaults_obrigatorios,
        confirmar=confirmar,
    )
    results: list[PatientResult] = []
    processed = 0

    for queue_patient in queue_patients:
        if only_password and queue_patient.senha != only_password:
            continue
        if limit is not None and processed >= limit:
            break
        processed += 1
        if progress_callback:
            progress_callback("inicio", queue_patient, None)

        if queue_patient.senha in successful_passwords:
            result = PatientResult(
                senha=queue_patient.senha,
                nome=queue_patient.nome,
                iniciais=queue_patient.iniciais,
                status="JA_LANCADO",
                mensagem="Senha ja consta no relatorio anterior com sucesso; paciente pulado.",
            )
            results.append(result)
            if progress_callback:
                progress_callback("fim", queue_patient, result)
            continue

        matches = clinical_by_password.get(queue_patient.senha, [])
        if not matches:
            result = PatientResult(
                senha=queue_patient.senha,
                nome=queue_patient.nome,
                iniciais=queue_patient.iniciais,
                status="PULADO",
                mensagem="Senha nao encontrada na planilha clinica.",
            )
            results.append(result)
            if progress_callback:
                progress_callback("fim", queue_patient, result)
            continue

        if len(matches) > 1:
            result = PatientResult(
                senha=queue_patient.senha,
                nome=queue_patient.nome,
                iniciais=queue_patient.iniciais,
                status="ERRO",
                mensagem="Senha duplicada na planilha clinica.",
            )
            results.append(result)
            if progress_callback:
                progress_callback("fim", queue_patient, result)
            continue

        clinical_patient = matches[0]
        alerts: list[str] = []
        if not equivalent(queue_patient.nome, clinical_patient.nome):
            alerts.append(
                f"Nome divergente: Salus='{queue_patient.nome}' Excel='{clinical_patient.nome}'."
            )
        if not equivalent(queue_patient.iniciais, clinical_patient.iniciais):
            alerts.append(
                f"Iniciais divergentes: Salus='{queue_patient.iniciais}' Excel='{clinical_patient.iniciais}'."
            )

        ignored_fields: list[str] = []
        prepared_fields: list[PreparedField] = []
        for column in field_headers:
            raw_value = clinical_patient.values.get(column)
            if is_blank(raw_value):
                ignored_fields.append(column)
                continue
            prepared_fields.append(prepare_field(column, raw_value, field_meta.get(column)))

        if not any(field.status == "ERRO" for field in prepared_fields):
            try:
                prepared_fields = executor.lancar(queue_patient, clinical_patient, prepared_fields)
            except AwaitingDischargeError as exc:
                result = PatientResult(
                    senha=queue_patient.senha,
                    nome=queue_patient.nome or clinical_patient.nome,
                    iniciais=queue_patient.iniciais or clinical_patient.iniciais,
                    status="AGUARDANDO",
                    mensagem=str(exc),
                    campos_preenchidos=[],
                    campos_ignorados=ignored_fields,
                    campos_com_erro=[],
                )
                results.append(result)
                if progress_callback:
                    progress_callback("fim", queue_patient, result)
                continue
            except AwaitingCidError as exc:
                result = PatientResult(
                    senha=queue_patient.senha,
                    nome=queue_patient.nome or clinical_patient.nome,
                    iniciais=queue_patient.iniciais or clinical_patient.iniciais,
                    status="AGUARDANDO_CID",
                    mensagem=str(exc),
                    campos_preenchidos=[],
                    campos_ignorados=ignored_fields,
                    campos_com_erro=[],
                )
                results.append(result)
                if progress_callback:
                    progress_callback("fim", queue_patient, result)
                continue
            except Exception as exc:
                result = PatientResult(
                    senha=queue_patient.senha,
                    nome=queue_patient.nome or clinical_patient.nome,
                    iniciais=queue_patient.iniciais or clinical_patient.iniciais,
                    status="ERRO",
                    mensagem=str(exc),
                    campos_preenchidos=[],
                    campos_ignorados=ignored_fields,
                    campos_com_erro=[str(exc)],
                )
                results.append(result)
                if progress_callback:
                    progress_callback("fim", queue_patient, result)
                if stop_on_error:
                    break
                continue

        result = build_patient_result(
            queue_patient,
            clinical_patient,
            prepared_fields,
            ignored_fields,
            alerts,
            dry_run=dry_run,
        )
        if not dry_run and not confirmar and result.status in {"SUCESSO", "SUCESSO_COM_ALERTA"}:
            result.status = "PRE_LANCADO"
            result.mensagem = "Pré-lançamento preenchido e validado; aguardando confirmação."
        results.append(result)
        if progress_callback:
            progress_callback("fim", queue_patient, result)
        if stop_on_error and result.status == "ERRO":
            break

    return results


def write_report(results: list[PatientResult], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"
    headers = [
        "Senha",
        "Nome",
        "Iniciais",
        "Status",
        "Mensagem",
        "Data/hora",
        "Campos preenchidos",
        "Campos ignorados",
        "Campos com erro",
    ]
    ws.append(headers)
    now = dt.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    for result in results:
        ws.append(
            [
                result.senha,
                result.nome,
                result.iniciais,
                result.status,
                result.mensagem,
                now,
                "\n".join(result.campos_preenchidos),
                "\n".join(result.campos_ignorados),
                "\n".join(result.campos_com_erro),
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [16, 38, 12, 20, 60, 20, 50, 50, 50]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def print_summary(results: list[PatientResult]) -> None:
    counts: dict[str, int] = {}
    for result in results:
        counts[result.status] = counts.get(result.status, 0) + 1
    print("Resumo:")
    for status, count in sorted(counts.items()):
        print(f"- {status}: {count}")
    print(f"Total processado: {len(results)}")


def default_report_name() -> str:
    return f"exports/relatorio_lancamentos_{dt.date.today().isoformat()}.xlsx"


def main() -> int:
    parser = argparse.ArgumentParser(description="Executa a Etapa 2 do Robo Sallus.")
    parser.add_argument("--fila", default="exports/pacientes_sirio_libanes_2026-07-15.xlsx")
    parser.add_argument("--clinica", default="exports/preenchimento_evolucao_clinica_LP_K3ZAVM6_2026-07-15.xlsx")
    parser.add_argument("--saida", default=default_report_name())
    parser.add_argument("--somente-senha", help="Processa apenas uma senha especifica.")
    parser.add_argument("--limite", type=int, help="Limita a quantidade de pacientes processados.")
    parser.add_argument("--sem-retomada", action="store_true", help="Nao pula senhas ja lancadas com sucesso em relatorio anterior.")
    parser.add_argument(
        "--executar-salus",
        action="store_true",
        help="Tenta executar no Salus. Sem esta flag, roda apenas simulacao/dry-run.",
    )
    parser.add_argument(
        "--preencher-obrigatorios-medios",
        action="store_true",
        help="Ao executar no Salus, preenche obrigatorios ausentes com defaults medios explicitamente autorizados.",
    )
    args = parser.parse_args()

    queue_path = Path(args.fila)
    clinical_path = Path(args.clinica)
    report_path = Path(args.saida)

    queue_patients = read_queue(queue_path)
    clinical_by_password, field_meta, field_headers = read_clinical(clinical_path)
    successful_passwords = set() if args.sem_retomada else read_successful_passwords(report_path)

    results = process_patients(
        queue_patients=queue_patients,
        clinical_by_password=clinical_by_password,
        field_meta=field_meta,
        field_headers=field_headers,
        successful_passwords=successful_passwords,
        dry_run=not args.executar_salus,
        only_password=args.somente_senha,
        limit=args.limite,
        usar_defaults_obrigatorios=args.preencher_obrigatorios_medios,
    )
    write_report(results, report_path)
    print_summary(results)
    print(f"Relatorio gerado: {report_path}")
    if not args.executar_salus:
        print("Modo: dry-run. Nenhum dado foi lancado no Salus.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
