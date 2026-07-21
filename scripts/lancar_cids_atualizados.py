#!/usr/bin/env python3
"""Dispara, um por vez, os registros cujo CID foi recuperado nesta revisão."""

import json
import time
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "exports" / "data_base_lancar_20_07_2026.xlsx"
API = "http://127.0.0.1:8765"


def request(path: str, payload=None):
    data = None if payload is None else json.dumps(payload).encode()
    req = urllib.request.Request(
        API + path,
        data=data,
        headers={"Content-Type": "application/json"},
        method="GET" if data is None else "POST",
    )
    with urllib.request.urlopen(req, timeout=15) as response:
        return json.load(response)


def passwords():
    ws = load_workbook(BASE, data_only=True, read_only=True).active
    headers = {str(cell.value).strip(): cell.column for cell in ws[1] if cell.value}
    result = []
    for row in range(2, ws.max_row + 1):
        status = str(ws.cell(row, headers["Lançamento Salus - Status"]).value or "").strip().upper()
        message = str(ws.cell(row, headers["Lançamento Salus - Mensagem"]).value or "")
        cid = str(ws.cell(row, headers["Dados da Internação - CID ajustado *"]).value or "").strip()
        if status == "AGUARDANDO" and cid and "CID " in message:
            result.append(str(ws.cell(row, headers["Senha"]).value).strip())
    return result


def main():
    queue = passwords()
    print(f"Lote CID: {len(queue)} senha(s).", flush=True)
    for index, senha in enumerate(queue, 1):
        while request("/api/status").get("running"):
            time.sleep(2)
        print(f"[{index}/{len(queue)}] Iniciando {senha}", flush=True)
        request("/api/etapa2", {"only_password": senha, "batch_limit": 1, "confirmar": True})
        while request("/api/status").get("running"):
            time.sleep(2)
        state = request("/api/status")
        rows = [row for row in state.get("launch_rows", []) if row.get("senha") == senha]
        situation = rows[-1].get("situacao") if rows else "finalizado pelo servidor"
        print(f"[{index}/{len(queue)}] {senha}: {situation}", flush=True)
    request("/api/refresh", {})
    print("Lote CID concluído.", flush=True)


if __name__ == "__main__":
    main()
