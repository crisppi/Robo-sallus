#!/usr/bin/env python3
"""Arquiva as planilhas ativas e prepara os arquivos de um novo dia.

O comando baixa a fila atual do Salus e gera uma base clinica limpa usando o
modelo versionado em ``templates/data_base_lancamento_modelo.xlsx``.
"""

from __future__ import annotations

import argparse
import datetime as dt
import shutil
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from gerar_lista_pacientes import fetch_patients, pick, save_excel
from preencher_evolucoes_excel import extract
from salus_cdp import SalusCdpError


ROOT = Path(__file__).resolve().parents[1]
EXPORTS = ROOT / "exports"
ARCHIVE = EXPORTS / "arquivo"
DEFAULT_TEMPLATE = ROOT / "templates" / "data_base_lancamento_modelo.xlsx"


def parse_date(value: str) -> dt.date:
    try:
        return dt.datetime.strptime(value, "%d_%m_%Y").date()
    except ValueError as exc:
        raise argparse.ArgumentTypeError("Use a data no formato DD_MM_AAAA.") from exc


def archive_active_exports() -> list[Path]:
    """Move todas as planilhas da raiz de exports para a pasta de arquivo."""
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    moved: list[Path] = []
    for source in sorted(EXPORTS.glob("*.xlsx")):
        destination = ARCHIVE / source.name
        if destination.exists():
            timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            destination = ARCHIVE / f"{source.stem}_{timestamp}{source.suffix}"
        shutil.move(str(source), destination)
        moved.append(destination)
    return moved


def generate_clinical_base(
    patients: list[dict],
    template: Path,
    output: Path,
    previous_workbook=None,
    evolution_date: dt.date | None = None,
) -> dict[str, int]:
    if previous_workbook is None and not template.exists():
        raise FileNotFoundError(f"Modelo da base clinica nao encontrado: {template}")

    workbook = previous_workbook or load_workbook(template)
    if "Preenchimento" not in workbook.sheetnames:
        raise RuntimeError("O modelo precisa conter a aba 'Preenchimento'.")
    sheet = workbook["Preenchimento"]
    if sheet.max_row < 2:
        raise RuntimeError("O modelo precisa conter uma linha formatada abaixo do cabecalho.")

    headers = {str(cell.value): cell.column for cell in sheet[1] if cell.value}
    if "evolucao" not in headers:
        # Modelos antigos não possuíam esta coluna. Acrescentar no fim evita
        # deslocar listas/validações existentes.
        evolution_column = sheet.max_column + 1
        sheet.cell(1, evolution_column).value = "evolucao"
        reference = sheet.cell(1, max(1, evolution_column - 1))
        target = sheet.cell(1, evolution_column)
        target.font = copy(reference.font)
        target.fill = copy(reference.fill)
        target.border = copy(reference.border)
        target.alignment = copy(reference.alignment)
        headers["evolucao"] = evolution_column

    previous_rows: dict[tuple[str, str], dict[str, object]] = {}
    senha_col = headers.get("Senha")
    id_col = headers.get("ID internação")
    if senha_col and id_col:
        for row_number in range(2, sheet.max_row + 1):
            senha = str(sheet.cell(row_number, senha_col).value or "").strip()
            admission_id = str(sheet.cell(row_number, id_col).value or "").strip()
            if senha:
                previous_rows[(senha, admission_id)] = {
                    header: sheet.cell(row_number, column).value
                    for header, column in headers.items()
                }

    row_style = [
        (
            copy(cell.font),
            copy(cell.fill),
            copy(cell.border),
            copy(cell.alignment),
            cell.number_format,
            copy(cell.protection),
        )
        for cell in sheet[2]
    ]

    last_row = max(sheet.max_row, len(patients) + 1)
    for row in sheet.iter_rows(min_row=2, max_row=last_row):
        for cell in row:
            cell.value = None

    reused_evolutions = 0
    derived_rows = 0
    derived_cells = 0
    for row_number, patient in enumerate(patients, 2):
        for column, style in enumerate(row_style, 1):
            cell = sheet.cell(row_number, column)
            cell.font = copy(style[0])
            cell.fill = copy(style[1])
            cell.border = copy(style[2])
            cell.alignment = copy(style[3])
            cell.number_format = style[4]
            cell.protection = copy(style[5])

        values = (
            pick(patient, "nomeCompleto", "Nome", "nomePaciente", "paciente"),
            pick(patient, "nomeIniciais", "Iniciais", "iniciais"),
            pick(patient, "senha", "Senha", "senhaInternacao"),
            pick(patient, "diasInternados", "DiasInternado", "diasInternado", "dias"),
            pick(patient, "idInternacao", "internacao"),
        )
        for column, value in enumerate(values, 1):
            sheet.cell(row_number, column).value = value

        senha = str(values[2] or "").strip()
        admission_id = str(values[4] or "").strip()
        previous = previous_rows.get((senha, admission_id), {})
        # Dados estáveis da mesma internação podem ser reaproveitados. Os
        # blocos de exame, conduta, UTI, auditor e status são sempre diários.
        reusable_headers = {
            "Dados da Internação - Caráter da internação *",
            "Dados da Internação - Tipo da internação *",
            "Dados da Internação - Data da internação *",
            "Dados da Internação - Acomodação *",
            "Dados da Internação - Paciente em isolamento? *",
            "Dados da Internação - Motivo do isolamento * (cond.)",
            "Dados da Internação - CID de internação *",
            "Dados da Internação - CID ajustado *",
            "Dados da Internação - Comorbidades *",
        }
        for header in reusable_headers:
            column = headers.get(header)
            value = previous.get(header)
            if column and value not in (None, ""):
                sheet.cell(row_number, column).value = value

        evolution_cell = sheet.cell(row_number, headers["evolucao"])
        previous_evolution = previous.get("evolucao")
        if previous_evolution not in (None, "") and str(previous_evolution).strip():
            evolution_cell.value = previous_evolution
            evolution_cell.fill = PatternFill("solid", fgColor="C6EFCE")
            reused_evolutions += 1

            evolution_date_column = headers.get("Data da evolução")
            if evolution_date_column and evolution_date:
                sheet.cell(row_number, evolution_date_column).value = evolution_date.strftime("%d/%m/%Y")

            inferred_values = extract(str(previous_evolution), values[3])
            row_writes = 0
            for header, inferred_value in inferred_values.items():
                column = headers.get(header)
                if not column or inferred_value in (None, ""):
                    continue
                current_value = sheet.cell(row_number, column).value
                if current_value in (None, ""):
                    sheet.cell(row_number, column).value = inferred_value
                    row_writes += 1
            if row_writes:
                derived_rows += 1
                derived_cells += row_writes
        else:
            evolution_cell.value = None
            evolution_cell.fill = PatternFill("solid", fgColor="FFC7CE")

    sheet.auto_filter.ref = f"A1:{sheet.cell(1, sheet.max_column).column_letter}{len(patients) + 1}"
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return {
        "evolucoes_reaproveitadas": reused_evolutions,
        "linhas_derivadas": derived_rows,
        "celulas_derivadas": derived_cells,
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Arquiva os arquivos antigos e gera a fila e a base do novo dia."
    )
    parser.add_argument("--data", type=parse_date, default=dt.date.today())
    parser.add_argument("--modelo", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument(
        "--base-anterior",
        type=Path,
        help="Base clinica a reutilizar. Se omitida, usa a base ativa mais recente.",
    )
    parser.add_argument("--cdp-url", default="http://127.0.0.1:9222")
    parser.add_argument("--user-key", type=int, default=49)
    parser.add_argument("--prestador", type=int, default=113)
    parser.add_argument("--empresa-auditoria", type=int, default=6)
    parser.add_argument("--tamanho-pagina", type=int, default=10)
    parser.add_argument(
        "--nao-arquivar",
        action="store_true",
        help="Gera os arquivos sem mover as planilhas que ja estao em exports.",
    )
    args = parser.parse_args()

    date_label = args.data.strftime("%d_%m_%Y")
    queue_output = EXPORTS / f"fila_salus_{date_label}.xlsx"
    clinical_output = EXPORTS / f"data_base_lancar_{date_label}.xlsx"

    # Baixa primeiro: se a sessao do Salus estiver indisponivel, nenhum arquivo
    # existente sera movido.
    patients = fetch_patients(
        user_key=args.user_key,
        prestador=args.prestador,
        empresa_auditoria=args.empresa_auditoria,
        page_size=args.tamanho_pagina,
        cdp_url=args.cdp_url,
    )
    if not patients:
        raise RuntimeError("O Salus retornou uma fila vazia; a virada foi cancelada.")

    active_bases = sorted(
        [
            path
            for path in (
                list(EXPORTS.glob("data_base_lancar*.xlsx"))
                + list(EXPORTS.glob("data_base_lancamento*.xlsx"))
            )
            if "antes_" not in path.name.lower() and not path.name.startswith("~$")
        ],
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    previous_date_label = (args.data - dt.timedelta(days=1)).strftime("%d_%m_%Y")
    canonical_previous_bases = [
        EXPORTS / f"data_base_lancar_{previous_date_label}.xlsx",
        EXPORTS / f"data_base_lancamento_{previous_date_label}.xlsx",
    ]
    canonical_previous_base = next(
        (path for path in canonical_previous_bases if path.exists()),
        None,
    )
    previous_base = args.base_anterior or canonical_previous_base or (
        active_bases[0] if active_bases else None
    )
    if previous_base and not previous_base.exists():
        raise FileNotFoundError(f"Base anterior nao encontrada: {previous_base}")
    previous_workbook = load_workbook(previous_base) if previous_base else None
    moved = [] if args.nao_arquivar else archive_active_exports()
    save_excel(patients, queue_output)
    stats = generate_clinical_base(
        patients,
        args.modelo,
        clinical_output,
        previous_workbook=previous_workbook,
        evolution_date=args.data,
    )

    print(f"Arquivos arquivados: {len(moved)}")
    print(f"Fila gerada: {queue_output}")
    print(f"Base gerada: {clinical_output}")
    print(f"Pacientes: {len(patients)}")
    print(f"Base anterior: {previous_base or 'modelo limpo'}")
    print(f"Evolucoes reaproveitadas: {stats['evolucoes_reaproveitadas']}")
    print(f"Linhas preenchidas pela evolucao: {stats['linhas_derivadas']}")
    print(f"Celulas preenchidas pela evolucao: {stats['celulas_derivadas']}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except SalusCdpError as exc:
        print(f"Erro: {exc}", file=sys.stderr)
        raise SystemExit(1)
