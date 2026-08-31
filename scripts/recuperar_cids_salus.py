#!/usr/bin/env python3
"""Preenche CIDs faltantes da base usando evolucao e detalhes da internacao."""

from __future__ import annotations

import argparse
import datetime as dt
import shutil
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from cid_evolucao import infer_adjusted_cid_from_evolution, infer_cid_suggestion
from etapa2_lancar_evolucao_salus import value_to_text
from salus_cdp import SalusCdpError, call_salus_api


ROOT = Path(__file__).resolve().parents[1]
EXPORTS = ROOT / "exports"
ARCHIVE = EXPORTS / "arquivo"


def newest_base() -> Path:
    matches = sorted(
        EXPORTS.glob("data_base_lancar_*.xlsx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not matches:
        raise FileNotFoundError("Nenhuma base data_base_lancar_*.xlsx encontrada em exports.")
    return matches[0]


def create_backup(path: Path) -> Path:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    timestamp = dt.datetime.now().strftime("%H%M")
    backup = ARCHIVE / f"{path.stem}_antes_recuperar_cids_salus_{timestamp}{path.suffix}"
    if backup.exists():
        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = ARCHIVE / f"{path.stem}_antes_recuperar_cids_salus_{timestamp}{path.suffix}"
    shutil.copy2(path, backup)
    return backup


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def salus_cids(id_internacao: str, cdp_url: str) -> tuple[str, str]:
    details = call_salus_api(
        f"/api/internacoes/{id_internacao}/detalhes-internacao?user_key=49",
        cdp_url=cdp_url,
    )
    admission = details.get("internacao", {}) if isinstance(details, dict) else {}
    return (
        value_to_text(admission.get("cidInicial")),
        value_to_text(admission.get("cidAtual")),
    )


def recover(path: Path, *, cdp_url: str, dry_run: bool = False) -> dict[str, Any]:
    workbook = load_workbook(path)
    sheet = workbook["Preenchimento"] if "Preenchimento" in workbook.sheetnames else workbook.active
    headers = {str(cell.value): cell.column for cell in sheet[1] if cell.value}

    required = [
        "Senha",
        "ID internação",
        "evolucao",
        "Dados da Internação - CID de internação *",
        "Dados da Internação - CID ajustado *",
    ]
    missing_headers = [header for header in required if header not in headers]
    if missing_headers:
        workbook.close()
        raise RuntimeError(f"Colunas ausentes: {', '.join(missing_headers)}")

    rows_missing_before = 0
    rows_changed = 0
    cells_written = 0
    from_evolution = 0
    from_salus = 0
    still_missing: list[tuple[int, str, str]] = []

    for row in range(2, sheet.max_row + 1):
        senha = value_to_text(sheet.cell(row, headers["Senha"]).value)
        if not senha:
            continue
        admission_cid = value_to_text(
            sheet.cell(row, headers["Dados da Internação - CID de internação *"]).value
        )
        adjusted_cid = value_to_text(
            sheet.cell(row, headers["Dados da Internação - CID ajustado *"]).value
        )
        if admission_cid and adjusted_cid:
            continue

        rows_missing_before += 1
        source = ""
        evolution = value_to_text(sheet.cell(row, headers["evolucao"]).value)
        inferred_cid, _reason = infer_cid_suggestion(evolution)
        if not admission_cid and inferred_cid:
            admission_cid = inferred_cid
            source = "evolucao"
        if admission_cid and not adjusted_cid:
            adjusted_cid = infer_adjusted_cid_from_evolution(evolution) or admission_cid
            source = source or "evolucao"

        if not admission_cid or not adjusted_cid:
            id_internacao = value_to_text(sheet.cell(row, headers["ID internação"]).value)
            if id_internacao:
                salus_initial, salus_current = salus_cids(id_internacao, cdp_url)
                if not admission_cid and salus_initial:
                    admission_cid = salus_initial
                    source = "salus"
                if not adjusted_cid:
                    adjusted_cid = salus_current or admission_cid
                    if adjusted_cid:
                        source = source or "salus"

        row_writes = 0
        if admission_cid and is_blank(sheet.cell(row, headers["Dados da Internação - CID de internação *"]).value):
            row_writes += 1
            cells_written += 1
            if not dry_run:
                sheet.cell(row, headers["Dados da Internação - CID de internação *"]).value = admission_cid
        if adjusted_cid and is_blank(sheet.cell(row, headers["Dados da Internação - CID ajustado *"]).value):
            row_writes += 1
            cells_written += 1
            if not dry_run:
                sheet.cell(row, headers["Dados da Internação - CID ajustado *"]).value = adjusted_cid

        if row_writes:
            rows_changed += 1
            if source == "salus":
                from_salus += 1
            else:
                from_evolution += 1
        else:
            still_missing.append((row, senha, value_to_text(sheet.cell(row, headers["ID internação"]).value)))

    if not dry_run:
        workbook.save(path)
    workbook.close()
    return {
        "arquivo": path,
        "linhas_sem_cid_antes": rows_missing_before,
        "linhas_alteradas": rows_changed,
        "celulas_preenchidas": cells_written,
        "linhas_por_evolucao": from_evolution,
        "linhas_por_salus": from_salus,
        "ainda_sem_cid": still_missing,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Recupera CIDs faltantes na base clinica.")
    parser.add_argument("base", nargs="?", type=Path, default=None)
    parser.add_argument("--cdp-url", default="http://127.0.0.1:9222")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--sem-backup", action="store_true")
    args = parser.parse_args()

    base = args.base or newest_base()
    backup = None
    if not args.dry_run and not args.sem_backup:
        backup = create_backup(base)
    result = recover(base, cdp_url=args.cdp_url, dry_run=args.dry_run)
    if backup:
        print(f"Backup: {backup}")
    print(f"Arquivo: {result['arquivo']}")
    print(f"Linhas sem algum CID antes: {result['linhas_sem_cid_antes']}")
    print(f"Linhas alteradas: {result['linhas_alteradas']}")
    print(f"Celulas CID preenchidas: {result['celulas_preenchidas']}")
    print(f"Linhas preenchidas pela evolucao: {result['linhas_por_evolucao']}")
    print(f"Linhas preenchidas pelo Salus: {result['linhas_por_salus']}")
    print(f"Linhas ainda sem CID: {len(result['ainda_sem_cid'])}")
    for row, senha, id_internacao in result["ainda_sem_cid"][:30]:
        print(f"SEM_CID linha={row} senha={senha} id_internacao={id_internacao}")
    return 1 if result["ainda_sem_cid"] else 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except SalusCdpError as exc:
        print(f"Erro: {exc}", file=sys.stderr)
        raise SystemExit(1)
