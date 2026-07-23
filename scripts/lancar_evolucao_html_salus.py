#!/usr/bin/env python3
"""Preenche a evolucao clinica no Salus pela propria tela HTML.

Este script usa o Chrome ja logado via DevTools remoto apenas como controle
do navegador. O preenchimento acontece clicando radios/checkboxes/botoes e
digitando em inputs da tela, sem POST direto para a API de respostas.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import random
import re
import sys
import time
import unicodedata
from pathlib import Path
from typing import Any

from etapa2_lancar_evolucao_salus import (
    ClinicalPatient,
    QueuePatient,
    PatientResult,
    read_clinical,
    read_queue,
    value_to_text,
    write_report,
)
from cid_evolucao import (
    infer_adjusted_cid_from_evolution,
    infer_cid_from_evolution,
)
from salus_cdp import DEFAULT_CDP, call_salus_api, evaluate_js, navigate_salus
from openpyxl import load_workbook


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def filled_values(clinical_patient: ClinicalPatient) -> dict[str, str]:
    values = {
        column: value_to_text(value)
        for column, value in clinical_patient.values.items()
        if not is_blank(value)
    }
    if not values.get("Dados da Internação - CID de internação *"):
        inferred_cid = infer_cid_from_evolution(values.get("evolucao", ""))
        if inferred_cid:
            values["Dados da Internação - CID de internação *"] = inferred_cid
            values["Dados da Internação - CID ajustado *"] = (
                infer_adjusted_cid_from_evolution(values.get("evolucao", "")) or inferred_cid
            )
    if values.get("Dados da Internação - CID de internação *") and not values.get("Dados da Internação - CID ajustado *"):
        # Regra operacional: na ausência de mudança diagnóstica explicitamente
        # marcada na geração da base, o CID ajustado repete o CID inicial.
        values["Dados da Internação - CID ajustado *"] = values["Dados da Internação - CID de internação *"]
    evolution = values.get("evolucao", "").lower()
    if not values.get("Dados da Internação - Acomodação *"):
        values["Dados da Internação - Acomodação *"] = "Apartamento / Enfermaria"
    if not values.get("Dados da Internação - Paciente em isolamento? *"):
        values["Dados da Internação - Paciente em isolamento? *"] = "Não"
    if not values.get("Dados da Internação - Queixa *"):
        if any(term in evolution for term in ("dor torác", "dor torac", "dor no peito", "precord", "sca")):
            complaint = "Dor no peito"
        elif any(term in evolution for term in ("cansaço", "cansaco", "fadiga")):
            complaint = "Fadiga / cansaço"
        elif any(term in evolution for term in ("pneum", "broncoasp", "bcp", "tosse")):
            complaint = "Tosse"
        else:
            complaint = "Dor inespecífica"
        values["Dados da Internação - Queixa *"] = complaint
    current_comorbidities = values.get("Dados da Internação - Comorbidades *", "")
    if not current_comorbidities or current_comorbidities.startswith("0SC"):
        comorbidities = []
        if re.search(r"\b(?:has|hipertens)", evolution, re.I):
            comorbidities.append("I10 - Hipertensão essencial (primária)")
        if re.search(r"\b(?:dac|coronariopatia)", evolution, re.I):
            comorbidities.append("I25 - Doença isquêmica crônica do coração")
        if re.search(r"\b(?:dlp|dislipidem)", evolution, re.I):
            comorbidities.append("E78 - Distúrbios do metabolismo de lipoproteínas e outras lipidemias")
        values["Dados da Internação - Comorbidades *"] = "; ".join(comorbidities) or current_comorbidities or "0SC - Sem comorbidades"
    if not values.get("Exame Físico - Mobilidade e dependência *"):
        values["Exame Físico - Mobilidade e dependência *"] = "Deambulando"
    if not values.get("Exame Físico - Acesso venoso? *"):
        values["Exame Físico - Acesso venoso? *"] = "Sim"
    if not values.get("Exame Físico - Qual o acesso venoso? * (cond.)"):
        values["Exame Físico - Qual o acesso venoso? * (cond.)"] = "Periférico"
    if not values.get("Exame Físico - Alimentação *"):
        values["Exame Físico - Alimentação *"] = "Oral"
    if not values.get("Exame Físico - Controle de eliminação *"):
        values["Exame Físico - Controle de eliminação *"] = "Normal"
    values.setdefault("UTI - Monitorização *", "Não")
    values.setdefault("UTI - Uso de droga vasoativa? *", "Não")
    for lab_field, flag_field in [
        ("UTI - Creatinina sérica (mg/dL) *", "UTI - Não mensurado * (cond.)"),
        ("UTI - pH arterial *", "UTI - Não mensurado * (cond.) [2]"),
        ("UTI - PaO2 (mmHg) *", "UTI - Não mensurado * (cond.) [3]"),
        ("UTI - FiO2 (%) *", "UTI - Não mensurado * (cond.) [4]"),
    ]:
        if not values.get(lab_field):
            values[flag_field] = "Sim"
    if not values.get("UTI - Categoria do diagnóstico principal *"):
        if any(term in evolution for term in ("pneum", "broncoasp", "bcp", "respirat")):
            category = "Respiratório"
        elif any(term in evolution for term in ("sca", "dac", "coronar", "cardi")):
            category = "Cardiovascular"
        elif any(term in evolution for term in ("sepse", "sépt", "sept")):
            category = "Sepse"
        elif any(term in evolution for term in ("avc", "neurol", "convuls")):
            category = "Neurológico"
        elif "trauma" in evolution:
            category = "Trauma"
        else:
            category = "Transplante/Outro (Crítico)"
        values["UTI - Categoria do diagnóstico principal *"] = category
    creatinine = re.search(r"creatinina\s*([0-9]+(?:[.,][0-9]+)?)", evolution, re.I)
    if creatinine:
        values["UTI - Creatinina sérica (mg/dL) *"] = creatinine.group(1).replace(",", ".")
    for field in [
        "Conduta Clínica - Uso de antibiótico? *",
        "Conduta Clínica - Uso de antifúngico? *",
        "Conduta Clínica - Uso de antiviral? *",
        "Conduta Clínica - Administração de Imunoglobulina *",
        "Conduta Clínica - Terapias Ativas (ex .: fisioterapia, suporte clínico) * *",
        "Conduta Clínica - Realizado procedimento cirúrgico? *",
        "Condição Adquirida - Paciente adquiriu alguma condição? *",
    ]:
        if not values.get(field):
            values[field] = "Não"
    values["Condição Adquirida - Paciente adquiriu alguma condição? *"] = "Não"
    for field in list(values):
        if field.startswith("Condição Adquirida - ") and field != "Condição Adquirida - Paciente adquiriu alguma condição? *":
            values[field] = ""
    if (
        "quimioterapia" in values.get("Conduta Clínica - Terapias em andamento * (cond.)", "").lower()
        and not values.get("Conduta Clínica - Tipo de Quimioterapia * (cond.)")
    ):
        values["Conduta Clínica - Tipo de Quimioterapia * (cond.)"] = "Curativa"
    return values


def build_browser_payload(clinical_patient: ClinicalPatient, confirmar: bool) -> dict[str, Any]:
    return {
        "idInternacao": clinical_patient.id_internacao,
        "confirmar": confirmar,
        "values": filled_values(clinical_patient),
    }


def update_lancamento_control(clinica_path: Path, controle_path: Path, patient: QueuePatient, result: dict[str, Any], mensagem: str) -> None:
    if not controle_path.exists():
        controle_path.parent.mkdir(parents=True, exist_ok=True)
        wb_src = load_workbook(clinica_path)
        wb_src.save(controle_path)

    wb = load_workbook(controle_path)
    ws = wb["Preenchimento"] if "Preenchimento" in wb.sheetnames else wb.active
    headers = {str(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if cell.value}

    def ensure_col(name: str) -> int:
        if name in headers:
            return headers[name]
        col = ws.max_column + 1
        ws.cell(1, col).value = name
        headers[name] = col
        return col

    senha_col = headers.get("Senha")
    if not senha_col:
        raise RuntimeError("Arquivo de controle sem coluna Senha.")

    status_col = ensure_col("Lançamento Salus - Status")
    dt_col = ensure_col("Lançamento Salus - Data/hora")
    msg_col = ensure_col("Lançamento Salus - Mensagem")
    href_col = ensure_col("Lançamento Salus - URL")

    status = "FINALIZADO" if result.get("saved") else ("CONFIRMADO_SEM_SALVAR" if result.get("confirmed") else ("PREENCHIDO" if result.get("confirmEnabled") else "ERRO"))
    now = dt.datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    for row in range(2, ws.max_row + 1):
        if value_to_text(ws.cell(row, senha_col).value) == patient.senha:
            ws.cell(row, status_col).value = status
            ws.cell(row, dt_col).value = now
            ws.cell(row, msg_col).value = mensagem
            ws.cell(row, href_col).value = value_to_text(result.get("href"))
            wb.save(controle_path)
            return
    raise RuntimeError(f"Senha {patient.senha} nao encontrada no arquivo de controle.")


def read_lancamento_status(controle_path: Path, senha: str) -> str:
    """Retorna o status persistido para impedir duas tentativas na mesma senha."""
    if not controle_path.exists():
        return ""

    wb = load_workbook(controle_path, read_only=True, data_only=True)
    try:
        ws = wb["Preenchimento"] if "Preenchimento" in wb.sheetnames else wb.active
        headers = {value_to_text(cell.value): idx for idx, cell in enumerate(ws[1], start=1)}
        senha_col = headers.get("Senha")
        status_col = headers.get("Lançamento Salus - Status")
        if not senha_col or not status_col:
            return ""
        senha_normalizada = value_to_text(senha)
        for row in range(2, ws.max_row + 1):
            if value_to_text(ws.cell(row, senha_col).value) == senha_normalizada:
                return value_to_text(ws.cell(row, status_col).value).strip()
        return ""
    finally:
        wb.close()


def run_html_fill(
    clinical_patient: ClinicalPatient,
    confirmar: bool = False,
    cdp_url: str = DEFAULT_CDP,
    usar_defaults_obrigatorios: bool = False,
) -> dict[str, Any]:
    payload = build_browser_payload(clinical_patient, confirmar=confirmar)
    js = f"""
    (async () => {{
      const payload = {json.dumps(payload, ensure_ascii=False)};
      const values = payload.values || {{}};
      const logs = [];
      const wait = (ms) => new Promise(resolve => setTimeout(resolve, ms));
      const norm = (value) => String(value ?? '')
        .normalize('NFD').replace(/[\\u0300-\\u036f]/g, '')
        .replace(/\\s+/g, ' ').trim().toLowerCase();
      const get = (name) => values[name];
      const has = (name) => get(name) != null && String(get(name)).trim() !== '';
      const jsDate = (value) => {{
        const text = String(value || '').trim();
        const br = /^(\\d{{2}})\\/(\\d{{2}})\\/(\\d{{4}})$/.exec(text);
        if (br) return `${{br[3]}}-${{br[2]}}-${{br[1]}}`;
        return text.slice(0, 10);
      }};
      const emit = (el) => {{
        el.dispatchEvent(new Event('input', {{ bubbles: true }}));
        el.dispatchEvent(new Event('change', {{ bubbles: true }}));
        el.dispatchEvent(new Event('blur', {{ bubbles: true }}));
      }};
      const setInput = (selector, value) => {{
        if (value == null || String(value).trim() === '') return false;
        const el = document.querySelector(selector);
        if (!el) {{ logs.push(`input nao encontrado: ${{selector}}`); return false; }}
        el.focus();
        el.value = value;
        emit(el);
        logs.push(`input: ${{selector}} = ${{value}}`);
        return true;
      }};
      const inputLabel = (el) => norm(el.closest('label')?.innerText || el.parentElement?.innerText || el.parentElement?.parentElement?.innerText || el.value);
      const clickInput = (name, expected) => {{
        if (expected == null || String(expected).trim() === '') return false;
        const wanted = norm(expected);
        const items = [...document.querySelectorAll(`input[name="${{CSS.escape(name)}}"]`)];
        const el = items.find(input => norm(input.value) === wanted || inputLabel(input).includes(wanted) || wanted.includes(inputLabel(input)));
        if (!el) {{ logs.push(`opcao nao encontrada: ${{name}} -> ${{expected}}`); return false; }}
        if (!el.checked) el.click();
        emit(el);
        logs.push(`opcao: ${{name}} -> ${{expected}}`);
        return true;
      }};
      const clickCheckboxByLabel = (expected) => {{
        if (expected == null || String(expected).trim() === '') return false;
        const wanted = norm(expected);
        const el = [...document.querySelectorAll('input[type="checkbox"]')]
          .find(input => inputLabel(input) === wanted || inputLabel(input).includes(wanted) || wanted.includes(inputLabel(input)));
        if (!el) {{ logs.push(`checkbox nao encontrado: ${{expected}}`); return false; }}
        if (!el.checked) el.click();
        emit(el);
        logs.push(`checkbox: ${{expected}}`);
        return true;
      }};
      const clickStepper = async (label) => {{
        const btn = [...document.querySelectorAll('button')]
          .find(b => norm(b.innerText).includes(norm(label)) && String(b.className).includes('evaluation-stepper'));
        if (!btn) throw new Error(`Etapa nao encontrada: ${{label}}`);
        btn.click();
        await wait(1000);
        logs.push(`etapa: ${{label}}`);
      }};
      const clickNext = async () => {{
        const btn = [...document.querySelectorAll('button')]
          .find(b => norm(b.innerText) === 'proximo' && !b.disabled);
        if (btn) {{
          btn.click();
          await wait(1400);
          logs.push('proximo');
        }}
      }};
      const optionNeedle = (value) => {{
        const text = String(value || '').trim();
        const dash = text.indexOf(' - ');
        return dash > 0 ? text.slice(dash + 3).trim() : text;
      }};
      const chooseSearchOption = async (triggerSelector, rawValue) => {{
        if (rawValue == null || String(rawValue).trim() === '') return false;
        const trigger = document.querySelector(triggerSelector);
        if (!trigger) {{ logs.push(`multiselect nao encontrado: ${{triggerSelector}}`); return false; }}
        const parts = String(rawValue).split(';').map(part => part.trim()).filter(Boolean);
        for (const part of parts) {{
          if (norm(trigger.innerText).includes(norm(part)) || norm(trigger.innerText).includes(norm(optionNeedle(part)))) {{
            logs.push(`multiselect ja contem: ${{part}}`);
            continue;
          }}
          trigger.click();
          await wait(350);
          const search = [...document.querySelectorAll('input[type="text"]')]
            .reverse().find(el => norm(el.placeholder).includes('pesquisar') || String(el.id).includes('multi-select-search'));
          if (search) {{
            search.focus();
            search.value = optionNeedle(part);
            emit(search);
            await wait(900);
          }}
          const candidates = [...document.querySelectorAll('button, [role="option"], li')]
            .filter(el => norm(el.innerText).includes(norm(part)) || norm(el.innerText).includes(norm(optionNeedle(part))));
          const option = candidates.find(el => !el.disabled && !String(el.className).includes('trigger'));
          if (option) {{
            option.click();
            await wait(350);
            logs.push(`multiselect: ${{triggerSelector}} -> ${{part}}`);
          }} else {{
            logs.push(`opcao pesquisavel nao encontrada: ${{part}}`);
          }}
        }}
        document.body.click();
        return true;
      }};
      const fillAdmission = async () => {{
        await clickStepper('Dados da Internação');
        setInput('#admission-date', jsDate(get('Dados da Internação - Data da internação *')));
        clickInput('admission-accommodation', get('Dados da Internação - Acomodação *'));
        clickInput('admission-patient-isolation', get('Dados da Internação - Paciente em isolamento? *'));
        if (has('Dados da Internação - Motivo do isolamento * (cond.)')) clickCheckboxByLabel(get('Dados da Internação - Motivo do isolamento * (cond.)'));
        await chooseSearchOption('#admission-complaint', get('Dados da Internação - Queixa *'));
        await chooseSearchOption('#admission-cid', get('Dados da Internação - CID de internação *'));
        await chooseSearchOption('#admission-comorbidities', get('Dados da Internação - Comorbidades *'));
        await chooseSearchOption('#admission-adjusted-cid', get('Dados da Internação - CID ajustado *'));
        const duration = document.querySelector('input[id^="admission-duration"]');
        if (duration && has('Dados da Internação - Tempo de existência da doença *')) {{
          duration.value = `${{get('Dados da Internação - Tempo de existência da doença *')}} ${{get('Dados da Internação - Nomenclatura do tempo de existência da doença *') || ''}}`.trim();
          emit(duration);
        }}
        await clickNext();
      }};
      const fillPhysical = async () => {{
        await clickStepper('Exame Físico');
        clickInput('physical-exam-general-state', get('Exame Físico - Estado geral *'));
        setInput('#physical-exam-systolic', get('Exame Físico - PA Sistólica max (mmHg) *'));
        setInput('#physical-exam-diastolic', get('Exame Físico - PA Diastólica max (mmHg) *'));
        setInput('#physical-exam-max-hr', get('Exame Físico - FC máx. (bpm) *'));
        setInput('#physical-exam-max-rr', get('Exame Físico - FR máx. (irpm) *'));
        setInput('#physical-exam-min-spo2', get('Exame Físico - SpO2 mín. (%) *'));
        setInput('#physical-exam-max-temp', get('Exame Físico - Temperatura máx. (°C) *'));
        clickInput('physical-exam-consciousness', get('Exame Físico - Nível de consciência *'));
        clickInput('physical-exam-mobility', get('Exame Físico - Mobilidade e dependência *'));
        clickInput('physical-exam-venous-yn', get('Exame Físico - Acesso venoso? *'));
        clickInput('physical-exam-airway', get('Exame Físico - Via respiratória *'));
        clickInput('physical-exam-resp-support', get('Exame Físico - Suporte respiratório *'));
        if (has('Exame Físico - Alimentação *')) for (const item of get('Exame Físico - Alimentação *').split(';')) clickCheckboxByLabel(item);
        clickInput('physical-exam-skin-lesion-yn', get('Exame Físico - Lesões na pele? *'));
        if (has('Exame Físico - Controle de eliminação *')) for (const item of get('Exame Físico - Controle de eliminação *').split(';')) clickCheckboxByLabel(item);
        await clickNext();
      }};
      const fillUti = async () => {{
        await clickStepper('UTI');
        clickInput('dynamic-question-112', get('UTI - 1. Abertura Ocular (E) - Selecione a melhor resposta observada. *'));
        clickInput('dynamic-question-113', get('UTI - 2. Resposta Verbal (V) - Avaliar conteúdo da comunicação verbal. *'));
        clickInput('dynamic-question-114', get('UTI - 3. Melhor Resposta Motora (M) - Registrar a melhor resposta obtida. *'));
        clickInput('dynamic-question-115', get('UTI - 4. Resposta Pupilar (P) - Avaliar reatividade pupilar ao estímulo luminoso. *'));
        clickInput('dynamic-question-116', get('UTI - Monitorização *'));
        clickInput('dynamic-question-117', get('UTI - Tipo de monitorização * (cond.)'));
        clickInput('dynamic-question-118', get('UTI - Uso de droga vasoativa? *'));
        setInput('input[id*="121"], input[placeholder*="Creatinina"]', get('UTI - Creatinina sérica (mg/dL) *'));
        if (norm(get('UTI - Não mensurado * (cond.)')) === 'sim') clickCheckboxByLabel('Não mensurado');
        setInput('input[id*="123"], input[placeholder*="pH"]', get('UTI - pH arterial *'));
        if (norm(get('UTI - Não mensurado * (cond.) [2]')) === 'sim') clickCheckboxByLabel('Não mensurado');
        setInput('input[id*="125"], input[placeholder*="PaO2"]', get('UTI - PaO2 (mmHg) *'));
        if (norm(get('UTI - Não mensurado * (cond.) [3]')) === 'sim') clickCheckboxByLabel('Não mensurado');
        setInput('input[id*="127"], input[placeholder*="FiO2"]', get('UTI - FiO2 (%) *'));
        if (norm(get('UTI - Não mensurado * (cond.) [4]')) === 'sim') clickCheckboxByLabel('Não mensurado');
        clickInput('dynamic-question-128', get('UTI - Categoria do diagnóstico principal *'));
        await clickNext();
      }};
      const fillConduct = async () => {{
        await clickStepper('Conduta Clínica');
        clickInput('cc-med-usage-antibiotic', get('Conduta Clínica - Uso de antibiótico? *'));
        clickInput('cc-med-usage-antifungal', get('Conduta Clínica - Uso de antifúngico? *'));
        clickInput('cc-med-usage-antiviral', get('Conduta Clínica - Uso de antiviral? *'));
        clickInput('dynamic-question-91', get('Conduta Clínica - Administração de Imunoglobulina *'));
        clickInput('dynamic-question-92', get('Conduta Clínica - Terapias Ativas (ex .: fisioterapia, suporte clínico) * *'));
        if (has('Conduta Clínica - Terapias em andamento * (cond.)')) for (const item of get('Conduta Clínica - Terapias em andamento * (cond.)').split(';')) clickCheckboxByLabel(item);
        clickInput('clinical-conduct-surgical-procedure', get('Conduta Clínica - Realizado procedimento cirúrgico? *'));
        await clickNext();
      }};
      const fillAcquired = async () => {{
        await clickStepper('Condição Adquirida');
        clickInput('acquired-condition-yn', get('Condição Adquirida - Paciente adquiriu alguma condição? *'));
        await wait(500);
        if (has('Condição Adquirida - Condição adquirida * (cond.)')) clickCheckboxByLabel(get('Condição Adquirida - Condição adquirida * (cond.)'));
        clickInput('dynamic-question-69', get('Condição Adquirida - Caracterização clínica da condição * (cond.)'));
        setInput('input[type="date"]', jsDate(get('Condição Adquirida - Data da condição adquirida * (cond.)')));
        await clickNext();
      }};
      const fillOpinion = async () => {{
        await clickStepper('Parecer do Auditor');
        clickInput('dynamic-question-144', get('Parecer do Auditor - Pertinência Técnica da Internação *'));
        clickInput('dynamic-question-145', get('Parecer do Auditor - Pertinência Técnica da permanência hospitalar *'));
        clickInput('dynamic-question-146', get('Parecer do Auditor - Paciente permanece internado? *'));
        await wait(500);
        if (has('Parecer do Auditor - Programação de alta * (cond.)')) clickInput('dynamic-question-147', get('Parecer do Auditor - Programação de alta * (cond.)'));
        await clickNext();
      }};
      await fillAdmission();
      await fillPhysical();
      await fillUti();
      await fillConduct();
      await fillAcquired();
      await fillOpinion();
      await clickStepper('Resumo');
      await wait(1000);
      const confirmButton = [...document.querySelectorAll('button')].find(b => norm(b.innerText) === 'confirmar evolucao' || norm(b.innerText) === 'confirmar evolução');
      const result = {{
        href: location.href,
        confirmEnabled: Boolean(confirmButton && !confirmButton.disabled),
        summary: document.body.innerText.slice(0, 6000),
        logs,
      }};
      if (payload.confirmar && confirmButton && !confirmButton.disabled) {{
        confirmButton.click();
        await wait(1800);
        result.afterConfirm = document.body.innerText.slice(0, 2000);
      }}
      return result;
    }})()
    """
    return evaluate_js(js, cdp_url=cdp_url)


def find_patient(fila: Path, clinica: Path, senha: str) -> tuple[QueuePatient, ClinicalPatient]:
    queue = {patient.senha: patient for patient in read_queue(fila)}
    clinical, _, _ = read_clinical(clinica)
    if senha not in queue:
        raise RuntimeError(f"Senha {senha} nao encontrada na fila.")
    matches = clinical.get(senha, [])
    if len(matches) != 1:
        raise RuntimeError(f"Senha {senha} nao encontrada de forma unica na planilha clinica.")
    return queue[senha], matches[0]


def run_html_fill(
    clinical_patient: ClinicalPatient,
    confirmar: bool = False,
    cdp_url: str = DEFAULT_CDP,
    usar_defaults_obrigatorios: bool = False,
) -> dict[str, Any]:
    """Preenche usando navegacao real por secao e eventos HTML."""
    payload = build_browser_payload(clinical_patient, confirmar=confirmar)
    base_url = f"https://salus.orizon.com.br/salus/avaliacao-internacao/{clinical_patient.id_internacao}/secao"
    all_logs: list[str] = []
    default_vitals = {
        "pas": str(random.randint(100, 140)),
        "pad": str(random.randint(60, 90)),
        "fc": str(random.randint(70, 90)),
        "fr": str(random.randint(14, 20)),
        "spo2": str(random.randint(95, 99)),
        "temp": f"{random.uniform(36.0, 37.2):.1f}",
    }

    common_js = """
      const wait = (ms) => new Promise(resolve => setTimeout(resolve, ms));
      const norm = (value) => String(value ?? '')
        .normalize('NFD').replace(/[\\u0300-\\u036f]/g, '')
        .replace(/\\s+/g, ' ').trim().toLowerCase();
      const get = (name) => values[name];
      const has = (name) => get(name) != null && String(get(name)).trim() !== '';
      const jsDate = (value) => {
        const text = String(value || '').trim();
        const br = /^(\\d{2})\\/(\\d{2})\\/(\\d{4})$/.exec(text);
        if (br) return `${br[3]}-${br[2]}-${br[1]}`;
        return text.slice(0, 10);
      };
      const emit = (el) => {
        el.dispatchEvent(new Event('input', { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));
        el.dispatchEvent(new Event('blur', { bubbles: true }));
      };
      const setInput = (selector, value) => {
        if (value == null || String(value).trim() === '') return false;
        const el = document.querySelector(selector);
        if (!el) { logs.push(`input nao encontrado: ${selector}`); return false; }
        el.focus();
        el.value = value;
        emit(el);
        logs.push(`input: ${selector} = ${value}`);
        return true;
      };
      const labelOf = (el) => norm(el.closest('label')?.innerText || el.parentElement?.innerText || el.parentElement?.parentElement?.innerText || el.value);
      const clickInput = (name, expected) => {
        if (expected == null || String(expected).trim() === '') return false;
        const wanted = norm(expected);
        const items = [...document.querySelectorAll(`input[name="${CSS.escape(name)}"]`)];
        const el = items.find(input => norm(input.value) === wanted || labelOf(input).includes(wanted) || wanted.includes(labelOf(input)));
        if (!el) { logs.push(`opcao nao encontrada: ${name} -> ${expected}`); return false; }
        if (!el.checked) el.click();
        emit(el);
        logs.push(`opcao: ${name} -> ${expected}`);
        return true;
      };
      const clickCheckboxByLabel = (expected) => {
        if (expected == null || String(expected).trim() === '') return false;
        const wanted = norm(expected);
        const el = [...document.querySelectorAll('input[type="checkbox"]')]
          .find(input => !input.checked && (labelOf(input) === wanted || labelOf(input).includes(wanted) || wanted.includes(labelOf(input))));
        if (!el) { logs.push(`checkbox nao encontrado ou ja marcado: ${expected}`); return false; }
        el.click();
        emit(el);
        logs.push(`checkbox: ${expected}`);
        return true;
      };
      const optionNeedle = (value) => {
        const text = String(value || '').trim();
        const dash = text.indexOf(' - ');
        return dash > 0 ? text.slice(dash + 3).trim() : text;
      };
      const chooseSearchOption = async (triggerSelector, rawValue) => {
        if (rawValue == null || String(rawValue).trim() === '') return false;
        const trigger = document.querySelector(triggerSelector);
        if (!trigger) { logs.push(`multiselect nao encontrado: ${triggerSelector}`); return false; }
        const parts = String(rawValue).split(';').map(part => part.trim()).filter(Boolean);
        if (norm(trigger.innerText).includes('selecionado')) {
          logs.push(`multiselect ja preenchido: ${triggerSelector}`);
          return true;
        }
        for (const part of parts) {
          if (norm(trigger.innerText).includes(norm(part)) || norm(trigger.innerText).includes(norm(optionNeedle(part)))) {
            logs.push(`multiselect ja contem: ${part}`);
            continue;
          }
          trigger.click();
          await wait(350);
          const search = [...document.querySelectorAll('input[type="text"]')]
            .reverse().find(el => norm(el.placeholder).includes('pesquisar') || String(el.id).includes('multi-select-search'));
          if (search) {
            search.focus();
            search.value = optionNeedle(part);
            emit(search);
            await wait(1000);
          }
          const candidates = [...document.querySelectorAll('button, [role="option"], li')]
            .filter(el => norm(el.innerText).includes(norm(part)) || norm(el.innerText).includes(norm(optionNeedle(part))));
          const option = candidates.find(el => !el.disabled && !String(el.className).includes('trigger'));
          if (option) {
            option.click();
            await wait(350);
            logs.push(`multiselect: ${triggerSelector} -> ${part}`);
          } else {
            logs.push(`opcao pesquisavel nao encontrada: ${part}`);
          }
        }
        document.body.click();
        return true;
      };
      const next = () => {
        const btn = [...document.querySelectorAll('button')].find(b => norm(b.innerText) === 'proximo' && !b.disabled);
        if (btn) {
          logs.push('proximo solicitado');
          return true;
        }
        logs.push('botao proximo nao encontrado/habilitado');
        return false;
      };
    """

    def run_section(secao: str, action_js: str) -> list[str]:
        print(f"HTML: abrindo secao {secao}", flush=True)
        target_key = f"/secao/{secao}"
        navigate_salus(f"{base_url}/{secao}", cdp_url=cdp_url)
        script = f"""
        (async () => {{
          const values = {json.dumps(payload["values"], ensure_ascii=False)};
          const logs = [];
          const expectedSecao = {json.dumps(secao)};
          for (let i = 0; i < 30; i++) {{
            const controlsReady = document.querySelectorAll('input,button,textarea,select').length > 20;
            if (location.href.includes(`/secao/${{expectedSecao}}`) && controlsReady) break;
            await new Promise(resolve => setTimeout(resolve, 500));
          }}
          await new Promise(resolve => setTimeout(resolve, 1000));
          const doActions = async () => {{
            {common_js}
            logs.push('acoes inicio');
            try {{
              {action_js}
            }} catch (error) {{
              logs.push(`erro javascript: ${{error && (error.stack || error.message) || error}}`);
            }}
            logs.push('acoes fim');
            return logs;
          }};
          return await Promise.race([
            doActions(),
            new Promise(resolve => setTimeout(() => resolve([...logs, 'timeout interno da secao']), 10000))
          ]);
        }})()
        """
        logs = evaluate_js(
            script,
            cdp_url=cdp_url,
            url_contains=target_key,
            timeout_seconds=35,
        ) or []
        if any(log == "proximo solicitado" for log in logs):
            evaluate_js(
                """
                (() => {
                  const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
                  const btn = [...document.querySelectorAll('button')].find(b => norm(b.innerText) === 'proximo' && !b.disabled);
                  if (!btn) return false;
                  setTimeout(() => btn.click(), 300);
                  return true;
                })()
                """,
                cdp_url=cdp_url,
                url_contains=target_key,
            )
        time.sleep(2.5)
        print(f"HTML: secao {secao} finalizada", flush=True)
        return logs

    all_logs += run_section(
        "100001",
        """
        setInput('#admission-date', jsDate(get('Dados da Internação - Data da internação *')));
        clickInput('admission-accommodation', get('Dados da Internação - Acomodação *'));
        clickInput('admission-patient-isolation', get('Dados da Internação - Paciente em isolamento? *'));
        if (has('Dados da Internação - Motivo do isolamento * (cond.)')) clickCheckboxByLabel(get('Dados da Internação - Motivo do isolamento * (cond.)'));
        await chooseSearchOption('#admission-complaint', get('Dados da Internação - Queixa *'));
        await chooseSearchOption('#admission-cid', get('Dados da Internação - CID de internação *'));
        await chooseSearchOption('#admission-comorbidities', get('Dados da Internação - Comorbidades *'));
        await chooseSearchOption('#admission-adjusted-cid', get('Dados da Internação - CID ajustado *'));
        const duration = document.querySelector('input[id^="admission-duration"]');
        if (duration && has('Dados da Internação - Tempo de existência da doença *')) {
          duration.value = `${get('Dados da Internação - Tempo de existência da doença *')} ${get('Dados da Internação - Nomenclatura do tempo de existência da doença *') || ''}`.trim();
          emit(duration);
          logs.push('input: admission-duration');
        }
        await next();
        """,
    )
    all_logs += run_section(
        "100002",
        """
        clickInput('physical-exam-general-state', get('Exame Físico - Estado geral *'));
        setInput('#physical-exam-systolic', get('Exame Físico - PA Sistólica max (mmHg) *'));
        setInput('#physical-exam-diastolic', get('Exame Físico - PA Diastólica max (mmHg) *'));
        setInput('#physical-exam-max-hr', get('Exame Físico - FC máx. (bpm) *'));
        setInput('#physical-exam-max-rr', get('Exame Físico - FR máx. (irpm) *'));
        setInput('#physical-exam-min-spo2', get('Exame Físico - SpO2 mín. (%) *'));
        setInput('#physical-exam-max-temp', get('Exame Físico - Temperatura máx. (°C) *'));
        clickInput('physical-exam-consciousness', get('Exame Físico - Nível de consciência *'));
        clickInput('physical-exam-mobility', get('Exame Físico - Mobilidade e dependência *'));
        clickInput('physical-exam-venous-yn', get('Exame Físico - Acesso venoso? *'));
        clickInput('physical-exam-airway', get('Exame Físico - Via respiratória *'));
        clickInput('physical-exam-resp-support', get('Exame Físico - Suporte respiratório *'));
        if (has('Exame Físico - Alimentação *')) for (const item of get('Exame Físico - Alimentação *').split(';')) clickCheckboxByLabel(item);
        clickInput('physical-exam-skin-lesion-yn', get('Exame Físico - Lesões na pele? *'));
        if (has('Exame Físico - Controle de eliminação *')) for (const item of get('Exame Físico - Controle de eliminação *').split(';')) clickCheckboxByLabel(item);
        await next();
        """,
    )
    all_logs += run_section(
        "100006",
        """
        clickInput('dynamic-question-112', get('UTI - 1. Abertura Ocular (E) - Selecione a melhor resposta observada. *'));
        clickInput('dynamic-question-113', get('UTI - 2. Resposta Verbal (V) - Avaliar conteúdo da comunicação verbal. *'));
        clickInput('dynamic-question-114', get('UTI - 3. Melhor Resposta Motora (M) - Registrar a melhor resposta obtida. *'));
        clickInput('dynamic-question-115', get('UTI - 4. Resposta Pupilar (P) - Avaliar reatividade pupilar ao estímulo luminoso. *'));
        clickInput('dynamic-question-116', get('UTI - Monitorização *'));
        clickInput('dynamic-question-117', get('UTI - Tipo de monitorização * (cond.)'));
        clickInput('dynamic-question-118', get('UTI - Uso de droga vasoativa? *'));
        clickInput('dynamic-question-128', get('UTI - Categoria do diagnóstico principal *'));
        await next();
        """,
    )
    all_logs += run_section(
        "100003",
        """
        clickInput('cc-med-usage-antibiotic', get('Conduta Clínica - Uso de antibiótico? *'));
        clickInput('cc-med-usage-antifungal', get('Conduta Clínica - Uso de antifúngico? *'));
        clickInput('cc-med-usage-antiviral', get('Conduta Clínica - Uso de antiviral? *'));
        clickInput('dynamic-question-91', get('Conduta Clínica - Administração de Imunoglobulina *'));
        clickInput('dynamic-question-92', get('Conduta Clínica - Terapias Ativas (ex .: fisioterapia, suporte clínico) * *'));
        if (has('Conduta Clínica - Terapias em andamento * (cond.)')) for (const item of get('Conduta Clínica - Terapias em andamento * (cond.)').split(';')) clickCheckboxByLabel(item);
        clickInput('clinical-conduct-surgical-procedure', get('Conduta Clínica - Realizado procedimento cirúrgico? *'));
        await next();
        """,
    )
    all_logs += run_section(
        "100004",
        """
        clickInput('acquired-condition-yn', get('Condição Adquirida - Paciente adquiriu alguma condição? *'));
        await new Promise(resolve => setTimeout(resolve, 500));
        if (has('Condição Adquirida - Condição adquirida * (cond.)')) clickCheckboxByLabel(get('Condição Adquirida - Condição adquirida * (cond.)'));
        clickInput('dynamic-question-69', get('Condição Adquirida - Caracterização clínica da condição * (cond.)'));
        setInput('input[type="date"]', jsDate(get('Condição Adquirida - Data da condição adquirida * (cond.)')));
        await next();
        """,
    )
    all_logs += run_section(
        "100005",
        """
        clickInput('dynamic-question-144', get('Parecer do Auditor - Pertinência Técnica da Internação *'));
        clickInput('dynamic-question-145', get('Parecer do Auditor - Pertinência Técnica da permanência hospitalar *'));
        clickInput('dynamic-question-146', get('Parecer do Auditor - Paciente permanece internado? *'));
        await next();
        """,
    )

    navigate_salus(f"{base_url}/100008", cdp_url=cdp_url)
    summary = evaluate_js(
        """
        (async () => {
          await new Promise(resolve => setTimeout(resolve, 2200));
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          const confirmButton = [...document.querySelectorAll('button')].find(b => norm(b.innerText) === 'confirmar evolucao' || norm(b.innerText) === 'confirmar evolução');
          const result = {href: location.href, confirmEnabled: Boolean(confirmButton && !confirmButton.disabled), summary: document.body.innerText.slice(0, 6000)};
          return result;
        })()
        """,
        cdp_url=cdp_url,
        url_contains="/secao/100008",
    )
    summary["logs"] = all_logs
    return summary


def run_html_fill(
    clinical_patient: ClinicalPatient,
    confirmar: bool = False,
    cdp_url: str = DEFAULT_CDP,
    usar_defaults_obrigatorios: bool = False,
) -> dict[str, Any]:
    """Preenche usando comandos pequenos na tela HTML do Salus."""
    payload = build_browser_payload(clinical_patient, confirmar=confirmar)
    values = payload["values"]
    base_url = f"https://salus.orizon.com.br/salus/avaliacao-internacao/{clinical_patient.id_internacao}/secao"
    all_logs: list[str] = []
    default_vitals = {
        "pas": str(random.randint(100, 140)),
        "pad": str(random.randint(60, 90)),
        "fc": str(random.randint(70, 90)),
        "fr": str(random.randint(14, 20)),
        "spo2": str(random.randint(95, 99)),
        "temp": f"{random.uniform(36.0, 37.2):.1f}",
    }

    def value(name: str) -> str:
        return str(values.get(name) or "").strip()

    def value_or(name: str, default: str) -> str:
        return value(name) or (default if usar_defaults_obrigatorios else "")

    def first_value(prefix: str) -> str:
        for key, raw in values.items():
            if key.startswith(prefix) and str(raw or "").strip():
                return str(raw).strip()
        return ""

    def date_html(text: str) -> str:
        text = text.strip()
        if len(text) >= 10 and text[2:3] == "/" and text[5:6] == "/":
            return f"{text[6:10]}-{text[3:5]}-{text[0:2]}"
        return text[:10]

    def eval_sec(secao: str, js: str) -> Any:
        return evaluate_js(
            js,
            cdp_url=cdp_url,
            url_contains=f"/avaliacao-internacao/{clinical_patient.id_internacao}/secao/{secao}",
        )

    def open_sec(secao: str) -> None:
        print(f"HTML: abrindo secao {secao}", flush=True)
        section_titles = {
            "100001": "Dados da Internação",
            "100002": "Exame Físico",
            "100006": "UTI",
            "100003": "Conduta Clínica",
            "100004": "Condição Adquirida",
            "100005": "Parecer do Auditor",
            "100008": "Resumo",
        }
        section_title = section_titles.get(secao, "")
        navigate_salus(
            f"{base_url}/{secao}",
            cdp_url=cdp_url,
            url_contains=f"/avaliacao-internacao/{clinical_patient.id_internacao}/",
        )
        time.sleep(1.0)
        ready = eval_sec(
            secao,
            f"""
            (async () => {{
              const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
              const expected = norm({json.dumps(section_title)});
              for (let i = 0; i < 60; i++) {{
                const body = norm(document.body.innerText || '');
                const ready = location.href.includes('/secao/{secao}')
                  && body.includes(expected)
                  && !body.includes('lista de pacientes internados')
                  && document.querySelectorAll('input,button,textarea,select').length > 2;
                if (ready) return true;
                await new Promise(resolve => setTimeout(resolve, 500));
              }}
              return false;
            }})()
            """,
        )
        if not ready:
            raise RuntimeError(f"Secao {secao} nao carregou a tela esperada ({section_title}).")

    def set_input(secao: str, selector: str, val: str) -> None:
        if not val:
            return
        js = f"""
        (() => {{
          const el = document.querySelector({json.dumps(selector)});
          if (!el) return `input nao encontrado: {selector}`;
          el.focus();
          el.value = {json.dumps(val)};
          for (const eventName of ['input', 'change', 'blur']) el.dispatchEvent(new Event(eventName, {{bubbles: true}}));
          return `input: {selector}`;
        }})()
        """
        all_logs.append(str(eval_sec(secao, js)))

    def set_duration_combo(secao: str, amount: str, unit: str) -> None:
        if not amount:
            return
        unit = unit or "Dias"
        js = f"""
        (async () => {{
          const sleep = (ms) => new Promise(resolve => setTimeout(resolve, ms));
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          const input = document.querySelector('#admission-duration-0') || document.querySelector('input[id^="admission-duration"]');
          if (!input) return 'duration combo nao encontrado';
          const amount = {json.dumps(amount)};
          const unit = {json.dumps(unit)};
          input.scrollIntoView({{block: 'center'}});
          input.focus();
          input.value = amount;
          for (const eventName of ['input', 'change']) input.dispatchEvent(new Event(eventName, {{bubbles: true}}));
          input.click();
          await sleep(700);
          const wanted = norm(`${{amount}} ${{unit}}`);
          let option = null;
          for (let i = 0; i < 8; i++) {{
            option = [...document.querySelectorAll('[role="option"], li, button')]
              .find(el => norm(el.innerText) === wanted || norm(el.innerText).includes(wanted));
            if (option) break;
            await sleep(400);
          }}
          if (!option) return `duration opcao nao encontrada: ${{amount}} ${{unit}}`;
          option.dispatchEvent(new PointerEvent('pointerdown', {{bubbles: true}}));
          option.dispatchEvent(new MouseEvent('mousedown', {{bubbles: true}}));
          option.dispatchEvent(new MouseEvent('mouseup', {{bubbles: true}}));
          option.click();
          await sleep(700);
          return `duration: ${{amount}} ${{unit}}`;
        }})()
        """
        all_logs.append(str(eval_sec(secao, js)))

    def fill_uti_labs(secao: str) -> None:
        js = f"""
        (() => {{
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          const visible = (el) => {{
            const rect = el.getBoundingClientRect();
            const style = getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 && style.display !== 'none' && style.visibility !== 'hidden';
          }};
          const fieldInput = (labelText) => {{
            const labels = [...document.querySelectorAll('label')].filter(visible);
            const label = labels.find(el => norm(el.innerText) === norm(labelText));
            if (!label) return null;
            const labelRect = label.getBoundingClientRect();
            return [...document.querySelectorAll('input[type="text"]')].filter(visible)
              .find(input => {{
                const rect = input.getBoundingClientRect();
                return Math.abs(rect.top - labelRect.top) < 80 && rect.left >= labelRect.left - 10;
              }}) || null;
          }};
          const notMeasured = (labelText) => {{
            const column = [...document.querySelectorAll('.uti-step__lab-col')]
              .find(el => norm(el.innerText).includes(norm(labelText)));
            if (!column) return false;
            const checkbox = column.querySelector('input[type="checkbox"]');
            if (!checkbox) return false;
            checkbox.scrollIntoView({{block: 'center'}});
            if (!checkbox.checked) checkbox.click();
            checkbox.dispatchEvent(new Event('change', {{bubbles:true}}));
            return checkbox.checked;
          }};
          const setText = (input, value) => {{
            if (!input || value == null || String(value).trim() === '') return false;
            input.scrollIntoView({{block: 'center'}});
            input.focus();
            input.value = String(value);
            for (const eventName of ['input', 'change', 'blur']) input.dispatchEvent(new Event(eventName, {{bubbles: true}}));
            return true;
          }};
          const logs = [];
          logs.push(`creatinina=${{setText(fieldInput('Creatinina sérica (mg/dL)'), {json.dumps(value('UTI - Creatinina sérica (mg/dL) *'))})}}`);
          if (norm({json.dumps(value('UTI - Não mensurado * (cond.)'))}).startsWith('sim')) logs.push(`creatinina-nao=${{notMeasured('Creatinina sérica (mg/dL)')}}`);
          if (norm({json.dumps(value('UTI - Não mensurado * (cond.) [2]'))}).startsWith('sim')) logs.push(`ph-nao=${{notMeasured('pH arterial')}}`);
          if (norm({json.dumps(value('UTI - Não mensurado * (cond.) [3]'))}).startsWith('sim')) logs.push(`pao2-nao=${{notMeasured('PaO2 (mmHg)')}}`);
          if (norm({json.dumps(value('UTI - Não mensurado * (cond.) [4]'))}).startsWith('sim')) logs.push(`fio2-nao=${{notMeasured('FiO2 (%)')}}`);
          // Regra final de segurança: qualquer laboratório vazio recebe
          // "Não mensurado" no checkbox do próprio bloco, independentemente
          // do conteúdo recebido da planilha.
          for (const column of document.querySelectorAll('.uti-step__lab-col')) {{
            const input = column.querySelector('input[type="text"]');
            const checkbox = column.querySelector('input[type="checkbox"]');
            if (input && checkbox && !String(input.value || '').trim() && !checkbox.checked) {{
              checkbox.click();
              checkbox.dispatchEvent(new Event('change', {{bubbles:true}}));
            }}
          }}
          return `uti labs: ${{logs.join('; ')}}`;
        }})()
        """
        all_logs.append(str(eval_sec(secao, js)))
        labs_complete = eval_sec(
            secao,
            """
            (() => [...document.querySelectorAll('.uti-step__lab-col')].every(column => {
              const input = column.querySelector('input[type="text"]');
              const checkbox = column.querySelector('input[type="checkbox"]');
              return Boolean(String(input?.value || '').trim()) || Boolean(checkbox?.checked);
            }))()
            """,
        )
        if not labs_complete:
            raise RuntimeError("UTI: existe exame obrigatório sem valor e sem Não mensurado.")

    def fill_antibiotic_details(secao: str) -> None:
        js = """
        (() => {
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 && style.display !== 'none' && style.visibility !== 'hidden';
          };
          const labels = [...document.querySelectorAll('label')].filter(visible);
          const logs = [];
          for (const label of labels) {
            if (norm(label.innerText) !== 'dosagem do antibiotico') continue;
            const labelRect = label.getBoundingClientRect();
            const input = [...document.querySelectorAll('input[type="text"]')]
              .filter(el => visible(el) && !el.disabled)
              .find(el => {
                const rect = el.getBoundingClientRect();
                return Math.abs(rect.top - labelRect.top) < 90 && rect.left >= labelRect.left - 20;
              });
            if (input && !String(input.value || '').trim()) {
              input.focus();
              input.value = '1';
              for (const eventName of ['input', 'change', 'blur']) input.dispatchEvent(new Event(eventName, {bubbles: true}));
              logs.push('dose=1');
            }
          }
          for (const label of labels) {
            if (norm(label.innerText) !== 'via do antibiotico') continue;
            const labelRect = label.getBoundingClientRect();
            const select = [...document.querySelectorAll('select')]
              .filter(visible)
              .find(el => {
                const rect = el.getBoundingClientRect();
                return Math.abs(rect.top - labelRect.top) < 90 && rect.left >= labelRect.left - 20;
              });
            if (select) {
              const option = [...select.options].find(opt => norm(opt.textContent) === 'intravenosa')
                || [...select.options].find(opt => norm(opt.textContent).includes('intravenosa'));
              if (option) {
                select.value = option.value;
                for (const eventName of ['input', 'change', 'blur']) select.dispatchEvent(new Event(eventName, {bubbles: true}));
                logs.push('via=Intravenosa');
              }
            }
          }
          return `antibiotico detalhes: ${logs.join('; ')}`;
        })()
        """
        all_logs.append(str(eval_sec(secao, js)))

    def click_button_text(secao: str, text: str) -> None:
        if not text:
            return
        js = f"""
        (() => {{
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          const wanted = norm({json.dumps(text)});
          const btn = [...document.querySelectorAll('button')].find(el => norm(el.innerText) === wanted || norm(el.innerText).includes(wanted));
          if (!btn) return `botao nao encontrado: {text}`;
          btn.click();
          return `botao: {text}`;
        }})()
        """
        all_logs.append(str(eval_sec(secao, js)))

    def set_operator_pending_justification(secao: str, val: str) -> None:
        js = f"""
        (() => {{
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          const items = [...document.querySelectorAll('label,input,textarea')];
          const start = items.findIndex(el => norm(el.innerText || '').includes('analise de procedimentos'));
          if (start < 0) return 'justificativa operador: grupo nao encontrado';
          const labelIndex = items.findIndex((el, index) => index > start && norm(el.innerText || '') === 'justifique');
          if (labelIndex < 0) return 'justificativa operador: label nao encontrado';
          const input = items.slice(labelIndex + 1).find(el => el.matches?.('input[type="text"], textarea'));
          if (!input) return 'justificativa operador: input nao encontrado';
          input.focus();
          input.value = {json.dumps(val)};
          for (const eventName of ['input', 'change', 'blur']) input.dispatchEvent(new Event(eventName, {{bubbles: true}}));
          return 'justificativa operador ajustada';
        }})()
        """
        all_logs.append(str(eval_sec(secao, js)))

    def clear_operator_pending(secao: str) -> None:
        js = """
        (() => {
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          const labels = [...document.querySelectorAll('label.auditor-opinion-step__option-item')];
          const start = labels.findIndex(label => norm(label.innerText).includes('analise de procedimentos'));
          if (start < 0) return 'pendencias operador: grupo nao encontrado';
          for (const label of labels.slice(start, start + 4)) {
            const input = label.querySelector('input') || label.nextElementSibling || label.previousElementSibling;
            if (input?.checked) input.click();
          }
          const items = [...document.querySelectorAll('label,input,textarea')];
          const labelIndex = items.findIndex((el, index) => index > start && norm(el.innerText || '') === 'justifique');
          const input = labelIndex >= 0 ? items.slice(labelIndex + 1).find(el => el.matches?.('input[type="text"], textarea')) : null;
          if (input) {
            input.value = '';
            for (const eventName of ['input', 'change', 'blur']) input.dispatchEvent(new Event(eventName, {bubbles: true}));
          }
          return 'pendencias operador limpas';
        })()
        """
        all_logs.append(str(eval_sec(secao, js)))

    def fill_auditor_history(secao: str, accommodation: str) -> None:
        js = f"""
        (async () => {{
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          const rawAccommodation = norm({json.dumps(accommodation)});
          let used = 'Apartamento / Enfermaria';
          let recommendation = 'Apartamento/Enfermaria';
          if (rawAccommodation.includes('uti')) {{
            used = 'UTI';
            recommendation = 'UTI';
          }} else if (rawAccommodation.includes('semi')) {{
            used = 'Semi';
            recommendation = 'Semi-Intensiva';
          }}

          const setSelect = (select, wanted) => {{
            if (!select) return false;
            const options = [...select.options];
            const option = options.find(opt => norm(opt.textContent) === norm(wanted) || norm(opt.value) === norm(wanted))
              || options.find(opt => norm(opt.textContent).includes(norm(wanted)) || norm(wanted).includes(norm(opt.textContent)));
            if (!option) return false;
            select.value = option.value;
            for (const eventName of ['input', 'change', 'blur']) select.dispatchEvent(new Event(eventName, {{bubbles: true}}));
            return true;
          }};

          const setText = (input, wanted) => {{
            if (!input) return false;
            input.focus();
            input.value = wanted;
            for (const eventName of ['input', 'change', 'blur']) input.dispatchEvent(new Event(eventName, {{bubbles: true}}));
            return true;
          }};

          const selects = [...document.querySelectorAll('select.auditor-opinion-step__history-select')];
          const names = [...document.querySelectorAll('input.auditor-opinion-step__history-input')];
          const saveButtons = [...document.querySelectorAll('button.auditor-opinion-step__history-action-button--save, button.auditor-opinion-step__history-action-button--s')]
            .filter(btn => norm(btn.innerText).includes('salvar') || !norm(btn.innerText));
          const rows = Math.floor(selects.length / 5);
          const logs = [];
          for (let row = 0; row < rows; row++) {{
            const base = row * 5;
            const ok = [
              setSelect(selects[base], used),
              setSelect(selects[base + 1], recommendation),
              setSelect(selects[base + 2], recommendation),
              setSelect(selects[base + 3], 'Médico Auditor'),
              setText(names[row], 'Tarsys'),
              setSelect(selects[base + 4], 'Prorrogado'),
            ];
            const save = saveButtons[row];
            if (save && !save.disabled) {{
              save.click();
              await new Promise(resolve => setTimeout(resolve, 250));
            }}
            logs.push(`linha ${{row + 1}}: ${{ok.filter(Boolean).length}}/6`);
          }}
          return `historico auditor: ${{rows}} linhas; ${{logs.join('; ')}}`;
        }})()
        """
        all_logs.append(str(eval_sec(secao, js)))

    def select_acquired_condition_date(secao: str, date_text: str) -> None:
        if not date_text:
            return
        day = date_text[:2].lstrip("0") if "/" in date_text[:3] else date_text[8:10].lstrip("0")
        js = f"""
        (() => {{
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          const day = {json.dumps(day)};
          const days = [...document.querySelectorAll('button.acq-date-modal__day')];
          const btn = days.find(button => norm(button.innerText) === norm(day) && !String(button.className).includes('muted'));
          if (!btn) return `data condicao: dia nao encontrado ${{day}}`;
          btn.click();
          const add = [...document.querySelectorAll('button')].find(button => norm(button.innerText) === 'adicionar');
          if (!add) return 'data condicao: adicionar nao encontrado';
          add.click();
          return `data condicao adicionada: ${{day}}`;
        }})()
        """
        all_logs.append(str(eval_sec(secao, js)))

    def click_radio(secao: str, name: str, val: str) -> None:
        if not val:
            return
        js = f"""
        (() => {{
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          const visible = (el) => {{
            const rect = el.getBoundingClientRect();
            const style = getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 && style.display !== 'none' && style.visibility !== 'hidden';
          }};
          const labelOf = (el) => norm(el.closest('label')?.innerText || el.parentElement?.innerText || el.parentElement?.parentElement?.innerText || el.value);
          const wanted = norm({json.dumps(val)});
          const wantedValue = wanted === 'sim' ? 'sim' : (wanted === 'nao' ? 'nao' : wanted);
          const items = [...document.querySelectorAll(`input[name="${{CSS.escape({json.dumps(name)})}}"]`)]
            .filter(visible);
          const el = items.find(input => norm(input.value) === wanted || labelOf(input) === wanted || labelOf(input).includes(wanted));
          if (!el) return `opcao nao encontrada: {name}`;
          const byValue = items.find(input => norm(input.value) === wantedValue);
          const target = byValue || el;
          target.scrollIntoView({{block: 'center'}});
          if (!target.checked) target.click();
          for (const eventName of ['input', 'change', 'blur']) target.dispatchEvent(new Event(eventName, {{bubbles: true}}));
          return `opcao: {name}`;
        }})()
        """
        all_logs.append(str(eval_sec(secao, js)))

    def click_first_radio(secao: str, name: str) -> None:
        """Seleciona a primeira opção visível de um grupo condicional obrigatório."""
        js = f"""
        (() => {{
          const visible = (el) => {{
            const rect = el.getBoundingClientRect();
            const style = getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 && style.display !== 'none' && style.visibility !== 'hidden';
          }};
          const items = [...document.querySelectorAll(`input[name="${{CSS.escape({json.dumps(name)})}}"]`)]
            .filter(visible);
          const target = items.find(input => input.checked) || items[0];
          if (!target) return `grupo obrigatorio nao encontrado: {name}`;
          target.scrollIntoView({{block: 'center'}});
          if (!target.checked) target.click();
          for (const eventName of ['input', 'change', 'blur']) target.dispatchEvent(new Event(eventName, {{bubbles: true}}));
          return `opcao padrao: {name}`;
        }})()
        """
        all_logs.append(str(eval_sec(secao, js)))

    def click_radio_name_prefix(secao: str, name_prefix: str, val: str) -> None:
        if not val:
            return
        js = f"""
        (() => {{
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          const labelOf = (el) => norm(el.closest('label')?.innerText || el.parentElement?.innerText || el.parentElement?.parentElement?.innerText || el.value);
          const wanted = norm({json.dumps(val)});
          const items = [...document.querySelectorAll('input[type="radio"]')]
            .filter(input => String(input.name || '').startsWith({json.dumps(name_prefix)}));
          const el = items.find(input => norm(input.value) === wanted || labelOf(input) === wanted || labelOf(input).includes(wanted));
          if (!el) return `opcao nao encontrada: {name_prefix}`;
          if (!el.checked) el.click();
          for (const eventName of ['input', 'change', 'blur']) el.dispatchEvent(new Event(eventName, {{bubbles: true}}));
          return `opcao: {name_prefix}`;
        }})()
        """
        all_logs.append(str(eval_sec(secao, js)))

    def click_radio_by_question(secao: str, question: str, val: str) -> None:
        if not val:
            return
        js = f"""
        (() => {{
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\u0300-\u036f]/g, '').replace(/\s+/g, ' ').trim().toLowerCase();
          const question = norm({json.dumps(question)});
          const wanted = norm({json.dumps(val)});
          const nodes = [...document.querySelectorAll('label,legend,h1,h2,h3,h4,p,span,div')]
            .filter(el => norm(el.innerText) === question || norm(el.innerText).startsWith(question));
          for (const node of nodes) {{
            let group = node.parentElement;
            for (let level = 0; group && level < 5; level++, group = group.parentElement) {{
              const radios = [...group.querySelectorAll('input[type="radio"]')];
              const target = radios.find(input => {{
                const label = norm(input.closest('label')?.innerText || input.parentElement?.innerText || input.value);
                return label === wanted || label.startsWith(wanted) || norm(input.value) === wanted;
              }});
              if (target) {{
                if (!target.checked) target.click();
                for (const eventName of ['input', 'change', 'blur']) target.dispatchEvent(new Event(eventName, {{bubbles: true}}));
                return `opcao por pergunta: {question}`;
              }}
            }}
          }}
          return `pergunta/opcao nao encontrada: {question}`;
        }})()
        """
        all_logs.append(str(eval_sec(secao, js)))


    def click_checkbox_label(secao: str, val: str) -> None:
        if not val:
            return
        js = f"""
        (() => {{
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          const visible = (el) => {{
            const rect = el.getBoundingClientRect();
            const style = getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 && style.display !== 'none' && style.visibility !== 'hidden';
          }};
          const wanted = norm({json.dumps(val)});
          const labelOf = (input) => norm(input.closest('label')?.innerText || input.parentElement?.innerText || input.parentElement?.parentElement?.innerText || '');
          const inputs = [...document.querySelectorAll('input[type="checkbox"]')].filter(visible);
          const input = inputs.find(el => labelOf(el) === wanted)
            || inputs.find(el => labelOf(el).includes(wanted) || wanted.includes(labelOf(el)));
          if (!input) return `checkbox nao encontrado: {val}`;
          input.scrollIntoView({{block: 'center'}});
          if (!input.checked) input.click();
          return `checkbox: {val}`;
        }})()
        """
        all_logs.append(str(eval_sec(secao, js)))

    def set_checkbox_label(secao: str, val: str, checked: bool) -> None:
        if not val:
            return
        js = f"""
        (() => {{
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\u0300-\u036f]/g, '').replace(/\s+/g, ' ').trim().toLowerCase();
          const wanted = norm({json.dumps(val)});
          const labelOf = (input) => norm(input.closest('label')?.innerText || input.parentElement?.innerText || input.parentElement?.parentElement?.innerText || '');
          const input = [...document.querySelectorAll('input[type="checkbox"]')]
            .find(el => labelOf(el) === wanted || labelOf(el).includes(wanted));
          if (!input) return `checkbox nao encontrado: {val}`;
          const desired = {str(checked).lower()};
          if (input.checked !== desired) input.click();
          for (const eventName of ['input', 'change', 'blur']) input.dispatchEvent(new Event(eventName, {{bubbles: true}}));
          return `checkbox sincronizado: {val}=${str(checked).lower()}`;
        }})()
        """
        all_logs.append(str(eval_sec(secao, js)))

    def click_auditor_radio_label(secao: str, val: str) -> None:
        if not val:
            return
        js = f"""
        (() => {{
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          const visible = (el) => {{
            const rect = el.getBoundingClientRect();
            const style = getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 && style.display !== 'none' && style.visibility !== 'hidden';
          }};
          const wanted = norm({json.dumps(val)});
          const labels = [...document.querySelectorAll('label.auditor-opinion-step__option-item')].filter(visible);
          const label = labels.find(el => norm(el.innerText) === wanted)
            || labels.find(el => norm(el.innerText).includes(wanted) || wanted.includes(norm(el.innerText)));
          if (!label) return `auditor radio nao encontrado: {val}`;
          const input = label.querySelector('input[type="radio"]');
          if (!input) return `auditor radio input nao encontrado: {val}`;
          label.scrollIntoView({{block: 'center'}});
          if (!input.checked) input.click();
          for (const eventName of ['input', 'change', 'blur']) input.dispatchEvent(new Event(eventName, {{bubbles: true}}));
          return `auditor radio: {val}`;
        }})()
        """
        all_logs.append(str(eval_sec(secao, js)))

    def choose_multi(secao: str, selector: str, raw: str) -> str:
        if not raw:
            return "valor vazio"
        js = f"""
        (async () => {{
          const sleep = (ms) => new Promise(resolve => setTimeout(resolve, ms));
          let trigger = null;
          for (let i = 0; i < 20; i++) {{
            trigger = document.querySelector({json.dumps(selector)});
            if (trigger) break;
            await sleep(500);
          }}
          if (!trigger) return `multiselect nao encontrado apos espera: {selector}`;
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          let text = norm(trigger.innerText);
          const parts = {json.dumps([part.strip() for part in raw.split(";") if part.strip()], ensure_ascii=False)};
          let selectedCount = 0;
          const hasRealValue = (value) => value
            && !value.includes('busque')
            && !value.includes('selecione')
            && !value.includes('pesquisar')
            && !/^0\s+selecionado/.test(value)
            && value !== '▾'
            && value !== 'x ▾'
            && value !== '× ▾';
          const needles = (part) => {{
            const dash = part.indexOf(' - ');
            const original = part.trim();
            const terms = [];
            // O catálogo CID do Salus não encontra códigos digitados com ponto.
            // Pesquisa o código compacto e confirma primeiro o código exato;
            // a categoria sem ponto fica apenas como último fallback.
            const code = dash > 0 ? part.slice(0, dash).trim() : original;
            if (/^[a-z]\d/i.test(code)) {{
              terms.push(code.replace(/\./g, ''));
              terms.push(code);
              if (code.includes('.')) terms.push(code.split('.')[0]);
            }}
            terms.push(original);
            if (dash > 0) {{
              terms.push(part.slice(dash + 3).trim());
            }}
            return [...new Set(terms.filter(Boolean))];
          }};
          const searchTerms = (part) => {{
            const dash = part.indexOf(' - ');
            const original = part.trim();
            const code = dash > 0 ? part.slice(0, dash).trim() : original;
            if (!/^[a-z]\d/i.test(code)) return needles(part);
            // O endpoint de busca do Salus responde rapidamente à categoria
            // (I50), mas frequentemente não responde a I50.0/I500.
            return [...new Set([
              code.includes('.') ? code.split('.')[0] : code,
              code.replace(/\./g, ''),
              code,
              original,
            ].filter(Boolean))];
          }};
          const complaintFallbacks = (part) => {{
            const p = norm(part);
            const terms = [];
            if (p.includes('tosse')) terms.push('Tosse');
            if (p.includes('dispne') || p.includes('falta de ar')) terms.push('Dispneia');
            if (p.includes('cans') || p.includes('fadiga')) terms.push('Fadiga / cansaço');
            if (p.includes('dor lombar') || p.includes('lombalgia')) terms.push('Dor lombar');
            if (p.includes('dor abdominal') || p.includes('abd')) terms.push('Dor abdominal');
            if (p.includes('precord') || p.includes('torac') || p.includes('peito')) terms.push('Dor no peito');
            if (p.includes('dor')) terms.push('Dor inespecífica');
            if (p.includes('cefale') || p.includes('confus') || p.includes('sonol') || p.includes('avc') || p.includes('neurol') || p.includes('convuls') || p.includes('linguagem')) terms.push('Cefaleia');
            terms.push('Dor inespecífica');
            return [...new Set(terms)];
          }};
          const isComplaint = {json.dumps(selector)} === '#admission-complaint';
          const isAdjustedCid = {json.dumps(selector)} === '#admission-adjusted-cid';
          const isSearchCatalog = ['#admission-complaint', '#admission-cid', '#admission-comorbidities', '#admission-adjusted-cid'].includes({json.dumps(selector)});
          const selectionNeedles = (part) => {{
            const dash = part.indexOf(' - ');
            const code = (dash > 0 ? part.slice(0, dash) : part).trim();
            if (/^[a-z]\d/i.test(code)) return [...new Set([code, code.replace(/\./g, '')])];
            return needles(part);
          }};
          const selectedCidCode = () => (text.match(/\b([a-z]\d{{2}}(?:\.\d+)?)\b/i) || [])[1] || '';
          const cidSelectionMatches = (part) => {{
            const dash = part.indexOf(' - ');
            const wanted = norm((dash > 0 ? part.slice(0, dash) : part).trim());
            const selected = norm(selectedCidCode());
            return Boolean(selected) && (selected === wanted || (wanted.includes('.') && selected === wanted.split('.')[0]));
          }};
          const desiredAlreadySelected = parts.every(part =>
            isAdjustedCid && /^[a-z]\d/i.test(part.trim())
              ? cidSelectionMatches(part)
              : selectionNeedles(part).some(term => text.includes(norm(term)))
          );
          if (desiredAlreadySelected) {{
            return `multiselect ja preenchido: {selector}`;
          }}
          if (isSearchCatalog && hasRealValue(text) && !isAdjustedCid) {{
            return `multiselect ja possui valor real: {selector}`;
          }}
          if (!isComplaint && text.includes('selecionado') && !isAdjustedCid) {{
            return `multiselect ja preenchido: {selector}`;
          }}
          if (isAdjustedCid && hasRealValue(text)) {{
            const clear = trigger.querySelector('[aria-label*="Limpar"], [title*="Limpar"]')
              || [...trigger.querySelectorAll('span, button')].find(el => norm(el.getAttribute('aria-label') || el.getAttribute('title') || '').includes('limpar'));
            if (!clear) return `CID ajustado divergente e sem controle para limpar: ${{text}}`;
            clear.dispatchEvent(new MouseEvent('mousedown', {{bubbles: true}}));
            clear.click();
            await sleep(500);
            text = norm(trigger.innerText);
          }}
          if (isComplaint && text && !text.includes('busque') && !text.includes('selecione')) {{
            const clear = trigger.querySelector('[aria-label*="Limpar"], [title*="Limpar"]')
              || [...trigger.querySelectorAll('span, button')].find(el => norm(el.getAttribute('aria-label') || el.getAttribute('title') || '').includes('limpar'));
            if (clear) {{
              clear.dispatchEvent(new MouseEvent('mousedown', {{bubbles: true}}));
              clear.click();
              await sleep(500);
            }}
          }}
          const searchParts = isComplaint ? [...parts, ...parts.flatMap(part => complaintFallbacks(part))] : parts;
          const lastSearchPart = searchParts[searchParts.length - 1];
          for (const part of searchParts) {{
            // Abre explicitamente o catálogo deste campo. Não reutiliza uma
            // caixa de pesquisa que tenha ficado aberta em outro componente.
            document.body.click();
            await sleep(250);
            trigger.scrollIntoView({{block: 'center'}});
            trigger.dispatchEvent(new MouseEvent('mousedown', {{bubbles: true}}));
            trigger.click();
            await sleep(400);
            let search = null;
            for (let i = 0; i < 10; i++) {{
              search = [...document.querySelectorAll('input[type="text"]')]
                .reverse().find(el => (norm(el.placeholder).includes('pesquisar') || String(el.id).includes('multi-select-search')) && el.offsetParent !== null);
              if (search) break;
              await sleep(300);
            }}
            let option = null;
            for (const term of searchTerms(part)) {{
              if (search) {{
                search.focus();
                search.value = term;
                search.dispatchEvent(new Event('input', {{bubbles: true}}));
                search.dispatchEvent(new Event('change', {{bubbles: true}}));
                await sleep(1500);
              }}
              for (let i = 0; i < 16; i++) {{
                const optionCandidates = [
                  ...document.querySelectorAll('button.multi-select__option, [role="option"], li')
                ].filter(el => el.offsetParent !== null && !el.disabled && !String(el.className || '').includes('multi-select__trigger'));
                for (const candidate of needles(part)) {{
                  option = optionCandidates.find(el => norm(el.innerText).includes(norm(candidate)));
                  if (option) break;
                }}
                if (option) break;
                await sleep(500);
              }}
              if (option) break;
            }}
            if (option) {{
              option.scrollIntoView({{block: 'center'}});
              option.dispatchEvent(new PointerEvent('pointerdown', {{bubbles: true}}));
              option.dispatchEvent(new MouseEvent('mousedown', {{bubbles: true}}));
              option.dispatchEvent(new PointerEvent('pointerup', {{bubbles: true}}));
              option.dispatchEvent(new MouseEvent('mouseup', {{bubbles: true}}));
              option.click();
              selectedCount += 1;
            }} else {{
              if (isComplaint && part !== lastSearchPart) {{
                document.body.click();
                await sleep(300);
                continue;
              }}
              if (!isComplaint) {{
                document.body.click();
                await sleep(300);
                continue;
              }}
              return `opcao pesquisavel nao encontrada apos espera: ${{part}}`;
            }}
            await sleep(300);
            if (isComplaint) break;
          }}
          document.body.click();
          return selectedCount > 0
            ? `multiselect: {selector}; selecionados=${{selectedCount}}`
            : `nenhuma opcao encontrada: {selector}`;
        }})()
        """
        result = str(eval_sec(secao, js))
        all_logs.append(result)
        return result

    def multi_has_value(secao: str, selector: str, raw: str) -> bool:
        if not raw:
            return False
        js = f"""
        (() => {{
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          const trigger = document.querySelector({json.dumps(selector)});
          if (!trigger) return false;
          const text = norm(trigger.innerText);
          if (!text || text.includes('busque') || text.includes('selecione')) return false;
          const parts = {json.dumps([part.strip() for part in raw.split(";") if part.strip()], ensure_ascii=False)};
          const needles = (part) => {{
            const dash = part.indexOf(' - ');
            const original = part.trim();
            const code = dash > 0 ? part.slice(0, dash).trim() : original;
            const terms = [];
            if (/^[a-z]\d/i.test(code)) {{
              terms.push(code.replace(/\./g, ''));
              terms.push(code);
            }}
            terms.push(original);
            if (dash > 0) {{
              terms.push(part.slice(dash + 3).trim());
            }}
            return [...new Set(terms.filter(Boolean))];
          }};
          const selectedCode = (text.match(/\b([a-z]\d{{2}}(?:\.\d+)?)\b/i) || [])[1] || '';
          return parts.every(part => {{
            const dash = part.indexOf(' - ');
            const wantedCode = norm((dash > 0 ? part.slice(0, dash) : part).trim());
            if (/^[a-z]\d/i.test(wantedCode) && selectedCode) {{
              return selectedCode === wantedCode
                || (wantedCode.includes('.') && selectedCode === wantedCode.split('.')[0]);
            }}
            return needles(part).some(term => text.includes(norm(term)));
          }});
        }})()
        """
        return bool(eval_sec(secao, js))

    def multi_has_any_value(secao: str, selector: str) -> bool:
        js = f"""
        (() => {{
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          const trigger = document.querySelector({json.dumps(selector)});
          if (!trigger) return false;
          const text = norm(trigger.innerText);
          return Boolean(text)
            && !text.includes('busque')
            && !text.includes('selecione')
            && !text.includes('pesquisar')
            && text !== '▾'
            && text !== 'x ▾'
            && text !== '× ▾';
        }})()
        """
        return bool(eval_sec(secao, js))

    def multi_current_value(secao: str, selector: str) -> str:
        js = f"""
        (() => {{
          const trigger = document.querySelector({json.dumps(selector)});
          if (!trigger) return '';
          const selected = trigger.querySelector('span');
          const text = String(selected?.innerText || trigger.innerText || '')
            .replace(/[✕×▾]/g, '').replace(/\s+/g, ' ').trim();
          if (/busque|selecione|pesquisar/i.test(text)) return '';
          return text;
        }})()
        """
        return str(eval_sec(secao, js) or "").strip()

    def choose_required_multi(secao: str, selector: str, raw: str, label: str) -> None:
        if not raw:
            raise RuntimeError(f"{label}: valor obrigatório vazio antes do preenchimento.")
        if multi_has_any_value(secao, selector):
            all_logs.append(f"{label}: ja possuia valor no HTML")
            return
        for attempt in range(3):
            choose_multi(secao, selector, raw)
            time.sleep(1.0)
            if multi_has_any_value(secao, selector) or multi_has_value(secao, selector, raw):
                all_logs.append(f"{label}: confirmado no HTML")
                return
            all_logs.append(f"{label}: tentativa {attempt + 1} nao persistiu")
        raise RuntimeError(f"{label}: nao persistiu no HTML apos 3 tentativas ({raw}).")

    def choose_comorbidities(secao: str, raw: str) -> None:
        """Seleciona comorbidades disponíveis; usa Sem comorbidades como fallback."""
        field_exists = bool(
            eval_sec(
                secao,
                "Boolean(document.querySelector('#admission-comorbidities'))",
            )
        )
        if not field_exists:
            # Alguns fluxos do Salus (por exemplo, internação cirúrgica) não
            # exibem este campo. Campo ausente no próprio formulário não é erro.
            all_logs.append("Comorbidades: campo não oferecido neste fluxo; ignorado")
            return
        items = [item.strip() for item in raw.split(";") if item.strip()]

        def select_code(code: str) -> Any:
            return eval_sec(
                secao,
                f"""
                (async () => {{
                  const sleep = ms => new Promise(resolve => setTimeout(resolve, ms));
                  const trigger = document.querySelector('#admission-comorbidities');
                  if (!trigger) return {{ok:false, error:'campo não encontrado'}};
                  if (String(trigger.innerText).includes({json.dumps(code)})) return {{ok:true, already:true}};
                  let search = [...document.querySelectorAll('input[type="text"]')]
                    .reverse().find(el => el.placeholder === 'Pesquisar' && el.offsetParent !== null);
                  if (!search) {{
                    trigger.scrollIntoView({{block:'center'}});
                    for (let attempt = 0; attempt < 3 && !search; attempt++) {{
                      trigger.dispatchEvent(new MouseEvent('mousedown', {{bubbles:true}}));
                      trigger.click();
                      await sleep(700);
                      search = [...document.querySelectorAll('input[type="text"]')]
                        .reverse().find(el => el.placeholder === 'Pesquisar' && el.offsetParent !== null);
                    }}
                  }}
                  if (!search) return {{ok:false, error:'pesquisa não abriu'}};
                  search.value = {json.dumps(code)};
                  search.dispatchEvent(new Event('input', {{bubbles:true}}));
                  await sleep(2200);
                  const option = [...document.querySelectorAll('button.multi-select__option')]
                    .find(el => String(el.innerText || '').trim().startsWith({json.dumps(code + ' -')}));
                  if (!option) return {{ok:false, error:'opção não encontrada'}};
                  option.click();
                  await sleep(700);
                  return {{ok:String(trigger.innerText).includes({json.dumps(code)}), value:trigger.innerText}};
                }})()
                """,
            )

        selected_codes: list[str] = []
        for item in items:
            code = item.split(" - ", 1)[0].strip()
            result = select_code(code)
            if not isinstance(result, dict) or not result.get("ok"):
                all_logs.append(f"Comorbidade {code}: ignorada, não encontrada no catálogo")
                continue
            selected_codes.append(code)
            all_logs.append(f"Comorbidade {code}: confirmada no HTML")

        if not selected_codes:
            fallback = select_code("0SC")
            if not isinstance(fallback, dict) or not fallback.get("ok"):
                # Regra operacional: falha ou ausência do catálogo de
                # comorbidades nunca interrompe o restante do lançamento.
                all_logs.append(
                    f"Comorbidades: seletor indisponível; campo deixado sem comorbidades ({fallback})"
                )
                eval_sec(
                    secao,
                    "document.body.click(); true",
                )
                return
            selected_codes = ["0SC"]
            all_logs.append("Nenhuma comorbidade encontrada; aplicado Sem comorbidades")

        closed = eval_sec(
            secao,
            f"""
            (async () => {{
              const sleep = ms => new Promise(resolve => setTimeout(resolve, ms));
              const trigger = document.querySelector('#admission-comorbidities');
              document.body.dispatchEvent(new MouseEvent('mousedown', {{bubbles:true}}));
              document.body.click();
              await sleep(300);
              const openPanel = [...document.querySelectorAll('.cdk-overlay-pane')]
                .find(panel => panel.offsetParent !== null);
              if (openPanel) {{
                trigger?.click();
                await sleep(300);
              }}
              const codes = {json.dumps(selected_codes, ensure_ascii=False)};
              return Boolean(trigger)
                && codes.every(code => String(trigger.innerText || '').includes(code))
                && ![...document.querySelectorAll('.cdk-overlay-pane')].some(panel => panel.offsetParent !== null);
            }})()
            """,
        )
        time.sleep(0.5)
        if not closed:
            all_logs.append(
                "Comorbidades: seleção não persistiu; campo ignorado sem interromper o lançamento"
            )
            return
        all_logs.append("Comorbidades: menu fechado e valores persistidos")

    def next_sec(secao: str, allow_incomplete: bool = False) -> None:
        js = """
        (() => {
          const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
          const btn = [...document.querySelectorAll('button')].find(b => norm(b.innerText) === 'proximo' && !b.disabled);
          if (!btn) return false;
          setTimeout(() => btn.click(), 300);
          return true;
        })()
        """
        all_logs.append(f"proximo {secao}: {eval_sec(secao, js)}")
        time.sleep(2.5)
        section_titles = {
            "100001": "Dados da Internação",
            "100002": "Exame Físico",
            "100006": "UTI",
            "100003": "Conduta Clínica",
            "100004": "Condição Adquirida",
            "100005": "Parecer do Auditor",
        }
        title = section_titles.get(secao, "")
        completed = evaluate_js(
            f"""
            (async () => {{
              const sleep = ms => new Promise(resolve => setTimeout(resolve, ms));
              const norm = value => String(value || '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').trim().toLowerCase();
              const expected = norm({json.dumps(title)});
              for (let attempt = 0; attempt < 12; attempt++) {{
                const item = [...document.querySelectorAll('.evaluation-stepper__item')]
                  .find(el => norm(el.innerText).includes(expected));
                if (item?.classList.contains('evaluation-stepper__item--completed')) return true;
                await sleep(1000);
              }}
              return false;
            }})()
            """,
            cdp_url=cdp_url,
            url_contains=f"/avaliacao-internacao/{clinical_patient.id_internacao}/",
        )
        if not completed and not allow_incomplete:
            raise RuntimeError(f"Página {title} não ficou verde; lote interrompido neste paciente.")
        print(f"HTML: secao {secao} finalizada", flush=True)

    def finish_at_summary(missing_cid: bool) -> dict[str, Any]:
        open_sec("100008")
        summary = eval_sec(
            "100008",
            """
            (async () => {
              const sleep = (ms) => new Promise(resolve => setTimeout(resolve, ms));
              const norm = (value) => String(value ?? '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').replace(/\\s+/g, ' ').trim().toLowerCase();
              let stepItems = [];
              let incompleteSteps = [];
              let confirmButton = null;
              let confirmEnabled = false;
              // Ao chegar ao resumo por "Próximo", o Angular atualiza o botão
              // alguns instantes depois da página aparecer. Aguarda apenas esse
              // estado, sem recarregar nem navegar para trás.
              for (let attempt = 0; attempt < 40; attempt++) {
                stepItems = [...document.querySelectorAll('.evaluation-stepper__item')];
                incompleteSteps = stepItems
                  .filter(item => !item.classList.contains('evaluation-stepper__item--completed') && !item.classList.contains('evaluation-stepper__item--current'))
                  .map(item => item.querySelector('.evaluation-stepper__label')?.innerText?.trim() || 'Página sem nome');
                confirmButton = [...document.querySelectorAll('button')]
                  .find(b => norm(b.innerText) === 'confirmar evolucao' || norm(b.innerText) === 'confirmar evolução');
                confirmEnabled = Boolean(confirmButton && !confirmButton.disabled && incompleteSteps.length === 0);
                if (confirmEnabled) break;
                await sleep(200);
              }
              const completedSteps = stepItems
                .filter(item => item.classList.contains('evaluation-stepper__item--completed'))
                .map(item => item.querySelector('.evaluation-stepper__label')?.innerText?.trim() || 'Página sem nome');
              let saved = false;
              if (__CONFIRMAR__ && confirmEnabled) {
                confirmButton.click();
                await sleep(900);
                const saveButton = [...document.querySelectorAll('button')]
                  .find(b => norm(b.innerText) === 'salvar e finalizar' || norm(b.innerText).includes('salvar e finalizar'));
                if (saveButton && !saveButton.disabled) {
                  saveButton.click();
                  saved = true;
                  await sleep(1800);
                }
              }
              return {href: location.href, confirmEnabled, confirmed: Boolean(__CONFIRMAR__ && confirmEnabled), saved, completedSteps, incompleteSteps, summary: document.body.innerText.slice(0, 6000)};
            })()
            """.replace("__CONFIRMAR__", "true" if confirmar else "false"),
        )
        summary["logs"] = all_logs
        summary["missingCid"] = missing_cid
        return summary

    open_sec("100001")
    set_input("100001", "#admission-date", date_html(value("Dados da Internação - Data da internação *")))
    click_radio("100001", "admission-accommodation", value("Dados da Internação - Acomodação *"))
    click_radio("100001", "admission-patient-isolation", value("Dados da Internação - Paciente em isolamento? *"))
    click_checkbox_label("100001", value("Dados da Internação - Motivo do isolamento * (cond.)"))
    choose_multi("100001", "#admission-complaint", value("Dados da Internação - Queixa *"))
    admission_cid_value = value("Dados da Internação - CID de internação *")
    adjusted_cid_from_salus = ""
    if not admission_cid_value:
        try:
            details = call_salus_api(
                f"/api/internacoes/{clinical_patient.id_internacao}/detalhes-internacao?user_key=49",
                cdp_url=cdp_url,
            )
            admission = details.get("internacao", {}) if isinstance(details, dict) else {}
            admission_cid_value = value_to_text(admission.get("cidInicial"))
            adjusted_cid_from_salus = value_to_text(admission.get("cidAtual"))
            if admission_cid_value:
                all_logs.append("CID de internação recuperado dos detalhes da internação no Salus")
        except Exception as exc:
            all_logs.append(f"CID da internação não recuperado do Salus: {exc}")
    if not admission_cid_value:
        admission_cid_value = infer_cid_from_evolution(value("evolucao"))
    missing_cid = not admission_cid_value
    if admission_cid_value:
        choose_multi("100001", "#admission-cid", admission_cid_value)
    adjusted_cid_value = value("Dados da Internação - CID ajustado *") or adjusted_cid_from_salus or admission_cid_value
    choose_comorbidities("100001", value("Dados da Internação - Comorbidades *"))
    if adjusted_cid_value:
        choose_required_multi("100001", "#admission-adjusted-cid", adjusted_cid_value, "CID ajustado")
    set_duration_combo("100001", value("Dados da Internação - Tempo de existência da doença *"), value_or("Dados da Internação - Nomenclatura do tempo de existência da doença *", "Dias"))
    next_sec("100001", allow_incomplete=missing_cid)

    original_status = value_to_text(
        clinical_patient.values.get("Lançamento Salus - Status")
    ).strip().upper()
    original_message = value_to_text(
        clinical_patient.values.get("Lançamento Salus - Mensagem")
    ).strip().upper()
    resume_cid_only = original_status == "AGUARDANDO_CID" or (
        original_status == "AGUARDANDO" and original_message.startswith("CID ")
    )
    if resume_cid_only and not missing_cid:
        all_logs.append("Fluxo retomado: CID preenchido; demais páginas preservadas")
        return finish_at_summary(missing_cid)

    open_sec("100002")
    click_radio("100002", "physical-exam-general-state", value("Exame Físico - Estado geral *"))
    set_input("100002", "#physical-exam-systolic", value_or("Exame Físico - PA Sistólica max (mmHg) *", default_vitals["pas"]))
    set_input("100002", "#physical-exam-diastolic", value_or("Exame Físico - PA Diastólica max (mmHg) *", default_vitals["pad"]))
    set_input("100002", "#physical-exam-max-hr", value_or("Exame Físico - FC máx. (bpm) *", default_vitals["fc"]))
    set_input("100002", "#physical-exam-max-rr", value_or("Exame Físico - FR máx. (irpm) *", default_vitals["fr"]))
    set_input("100002", "#physical-exam-min-spo2", value_or("Exame Físico - SpO2 mín. (%) *", default_vitals["spo2"]))
    set_input("100002", "#physical-exam-max-temp", value_or("Exame Físico - Temperatura máx. (°C) *", default_vitals["temp"]))
    click_radio("100002", "physical-exam-consciousness", value("Exame Físico - Nível de consciência *"))
    click_radio("100002", "physical-exam-mobility", value("Exame Físico - Mobilidade e dependência *"))
    venous_access_yn = value_or("Exame Físico - Acesso venoso? *", "Sim")
    click_radio("100002", "physical-exam-venous-yn", venous_access_yn)
    if venous_access_yn.lower().startswith("s"):
        venous_type = value_or("Exame Físico - Qual o acesso venoso? * (cond.)", "Periférico")
        click_checkbox_label("100002", venous_type)
        if "central" in venous_type.lower():
            click_radio("100002", "physical-exam-central-detail", value("Exame Físico - Detalhamento do acesso central * (cond.)"))
    click_radio("100002", "physical-exam-airway", value("Exame Físico - Via respiratória *"))
    click_radio("100002", "physical-exam-resp-support", value("Exame Físico - Suporte respiratório *"))
    if value("Exame Físico - Suporte respiratório *").lower().startswith("suporte"):
        click_radio("100002", "physical-exam-noninvasive-detail", value("Exame Físico - Detalhamento do suporte respiratório * (cond.)"))
        support_frequency = value("Exame Físico - Frequência do suporte respiratório * (cond.)")
        if support_frequency:
            click_radio("100002", "physical-exam-noninvasive-freq", support_frequency)
        else:
            click_first_radio("100002", "physical-exam-noninvasive-freq")
    for item in value_or("Exame Físico - Alimentação *", "Oral").split(";"):
        click_checkbox_label("100002", item.strip())
    feeding_detail = value_or("Exame Físico - Detalhamento enteral * (cond.)", "GTT - Gastrostomia")
    if "enteral" in value_or("Exame Físico - Alimentação *", "Oral").lower():
        click_radio("100002", "physical-exam-enteral-detail", feeding_detail)
    click_radio("100002", "physical-exam-skin-lesion-yn", value("Exame Físico - Lesões na pele? *"))
    if value("Exame Físico - Lesões na pele? *").lower().startswith("s"):
        lesion_locations = value("Exame Físico - Localização da lesão * (cond.)")
        other_locations = value("Exame Físico - Localização outras regiões * (cond.)")
        for item in lesion_locations.split(";"):
            item = item.strip()
            if not item:
                continue
            # "Outras regiões" abre outro campo obrigatório. Sem o detalhe
            # explícito, preserva as demais localizações informadas.
            if "outras regi" in item.lower() and not other_locations:
                continue
            click_checkbox_label("100002", item)
        for field in (
            "Exame Físico - Detalhamento trocantérica * (cond.)",
            "Exame Físico - Detalhamento calcâneo * (cond.)",
            "Exame Físico - Localização outras regiões * (cond.)",
            "Exame Físico - Características clínicas da lesão * (cond.)",
        ):
            for item in value(field).split(";"):
                if item.strip():
                    click_checkbox_label("100002", item.strip())
        click_radio("100002", "dynamic-question-59", value("Exame Físico - Tipo de lesão identificada * (cond.)"))
        click_radio("100002", "dynamic-question-60", value("Exame Físico - Condição atual da lesão * (cond.)"))
        dressing_yn = value_or("Exame Físico - Houve curativo? * (cond.)", "Não")
        click_radio("100002", "dynamic-question-63", dressing_yn)
        if dressing_yn.lower().startswith("s"):
            click_radio("100002", "dynamic-question-64", value("Exame Físico - Tipo de curativo * (cond.)"))
            click_radio("100002", "dynamic-question-295", value("Exame Físico - Frequência de troca do curativo * (cond.)"))
            click_radio("100002", "dynamic-question-65", value("Exame Físico - Frequência da troca do curativo está adequada? * (cond.)"))
    for item in value_or("Exame Físico - Controle de eliminação *", "Normal").split(";"):
        click_checkbox_label("100002", item.strip())
    next_sec("100002")

    accommodation_norm = value("Dados da Internação - Acomodação *").lower()
    if "uti" in accommodation_norm or "semi" in accommodation_norm:
        try:
            open_sec("100006")
        except RuntimeError as exc:
            all_logs.append(f"uti ignorada: {exc}")
        else:
            for field, name in [
                ("UTI - 1. Abertura Ocular (E) - Selecione a melhor resposta observada. *", "dynamic-question-112"),
                ("UTI - 2. Resposta Verbal (V) - Avaliar conteúdo da comunicação verbal. *", "dynamic-question-113"),
                ("UTI - 3. Melhor Resposta Motora (M) - Registrar a melhor resposta obtida. *", "dynamic-question-114"),
                ("UTI - 4. Resposta Pupilar (P) - Avaliar reatividade pupilar ao estímulo luminoso. *", "dynamic-question-115"),
                ("UTI - Monitorização *", "dynamic-question-116"),
                ("UTI - Tipo de monitorização * (cond.)", "dynamic-question-117"),
                ("UTI - Uso de droga vasoativa? *", "dynamic-question-118"),
                ("UTI - Categoria do diagnóstico principal *", "dynamic-question-128"),
            ]:
                default = "6" if name == "dynamic-question-114" else ("Não invasiva" if name == "dynamic-question-117" else "")
                click_radio("100006", name, value_or(field, default) if default else value(field))
            if value("UTI - Uso de droga vasoativa? *").lower().startswith("s"):
                choose_multi(
                    "100006",
                    "#uti-vasoactive-drugs",
                    value("UTI - Drogas vasoativas em uso * (cond.)"),
                )
            fill_uti_labs("100006")
            next_sec("100006")

    open_sec("100003")
    antibiotic_yn = value_or("Conduta Clínica - Uso de antibiótico? *", "Não")
    antifungal_yn = value_or("Conduta Clínica - Uso de antifúngico? *", "Não")
    antiviral_yn = value_or("Conduta Clínica - Uso de antiviral? *", "Não")
    active_therapy_yn = value_or("Conduta Clínica - Terapias Ativas (ex .: fisioterapia, suporte clínico) * *", "Não")
    surgery_yn = value_or("Conduta Clínica - Realizado procedimento cirúrgico? *", "Não")

    click_radio("100003", "cc-med-usage-antibiotic", antibiotic_yn)
    if antibiotic_yn.lower().startswith("s"):
        antibiotic_result = choose_multi("100003", "#medicamento-search-multi-select-0", value("Conduta Clínica - Selecione os antibióticos em uso * (cond.)"))
        if antibiotic_result.startswith("nenhuma opcao") or antibiotic_result == "valor vazio":
            click_radio("100003", "cc-med-usage-antibiotic", "Não")
        else:
            fill_antibiotic_details("100003")
    click_radio("100003", "cc-med-usage-antifungal", antifungal_yn)
    if antifungal_yn.lower().startswith("s"):
        antifungal_result = choose_multi("100003", "#clinical-conduct-antifungal-options", value("Conduta Clínica - Selecione os antifúngicos em uso * (cond.)"))
        if antifungal_result.startswith("nenhuma opcao") or antifungal_result == "valor vazio":
            click_radio("100003", "cc-med-usage-antifungal", "Não")
        else:
            click_radio_by_question("100003", "Via do antifúngico", value_or("Conduta Clínica - Via do antifúngico * (cond.)", "Via intravenosa"))
    click_radio("100003", "cc-med-usage-antiviral", antiviral_yn)
    if antiviral_yn.lower().startswith("s"):
        antiviral_result = choose_multi("100003", "#clinical-conduct-antiviral-options", value("Conduta Clínica - Selecione os antivirais em uso * (cond.)"))
        if antiviral_result.startswith("nenhuma opcao") or antiviral_result == "valor vazio":
            click_radio("100003", "cc-med-usage-antiviral", "Não")
        else:
            click_radio_by_question("100003", "Via do antiviral", value_or("Conduta Clínica - Via do antiviral * (cond.)", "Via intravenosa"))
    click_radio_by_question(
        "100003",
        "Câmara Hiperbárica",
        value_or("Conduta Clínica - Câmara Hiperbárica * (cond.)", "Não"),
    )
    click_radio("100003", "dynamic-question-91", value_or("Conduta Clínica - Administração de Imunoglobulina *", "Não"))
    click_radio("100003", "dynamic-question-92", active_therapy_yn)
    if active_therapy_yn.lower().startswith("s"):
        therapies = value("Conduta Clínica - Terapias em andamento * (cond.)")
        set_checkbox_label(
            "100003",
            "Terapia Renal Substitutiva (TRS)",
            "terapia renal substitutiva" in therapies.lower(),
        )
        for item in therapies.split(";"):
            click_checkbox_label("100003", item.strip())
        if "terapia renal substitutiva" in therapies.lower():
            click_radio_by_question(
                "100003",
                "Tipo de terapia renal substitutiva (TRS)",
                value_or(
                    "Conduta Clínica - Tipo de terapia renal substitutiva (TRS) * (cond.)",
                    "Hemodiálise",
                ),
            )
        if "radioterapia" in therapies.lower():
            click_radio_by_question(
                "100003",
                "Tipo de radioterapia",
                value_or("Conduta Clínica - Tipo de radioterapia * (cond.)", "Convencional"),
            )
        if "quimioterapia" in therapies.lower():
            click_radio(
                "100003",
                "clinical-conduct-chemo-type",
                value_or("Conduta Clínica - Tipo de Quimioterapia * (cond.)", "Curativa"),
            )
    click_radio("100003", "clinical-conduct-surgical-procedure", surgery_yn)
    if surgery_yn.lower().startswith("s"):
        choose_multi("100003", "#padrao-tiss-search-multi-select-0", value("Conduta Clínica - TUSS + Nome do Procedimento * (cond.)"))
        click_radio("100003", "clinical-conduct-anesthesia-107", value("Conduta Clínica - Tipo de anestesia * (cond.)"))
        click_radio("100003", "clinical-conduct-intraoperative-complications", value_or("Conduta Clínica - Houve intercorrências no intraoperatório? * (cond.)", "Não"))
    next_sec("100003")

    open_sec("100004")
    acquired_yn = value("Condição Adquirida - Paciente adquiriu alguma condição? *") or "Não"
    click_radio("100004", "acquired-condition-yn", acquired_yn)
    if acquired_yn.lower().startswith("s"):
        click_checkbox_label("100004", value("Condição Adquirida - Condição adquirida * (cond.)"))
        time.sleep(0.5)
        acquired_other = (
            value("Condição Adquirida - Caracterização clínica da condição - Outros (cond.)")
            or value("Condição Adquirida - Descrição da condição adquirida * (cond.)")
        )
        set_input("100004", "#acquired-cond-outros-115", acquired_other)
        acquired_char = first_value("Condição Adquirida - Caracterização clínica da condição * (cond.)")
        click_radio_name_prefix("100004", "acquired-cond-char-", acquired_char)
        acquired_date = first_value("Condição Adquirida - Data da condição adquirida * (cond.)")
        if acquired_date:
            click_button_text("100004", "Data da condição adquirida")
            time.sleep(0.5)
            select_acquired_condition_date("100004", acquired_date)
    next_sec("100004")

    open_sec("100005")
    click_radio("100005", "dynamic-question-144", value("Parecer do Auditor - Pertinência Técnica da Internação *"))
    click_radio("100005", "dynamic-question-145", value("Parecer do Auditor - Pertinência Técnica da permanência hospitalar *"))
    remains_admitted = value_or("Parecer do Auditor - Paciente permanece internado? *", "Sim")
    click_radio("100005", "dynamic-question-146", remains_admitted)
    if remains_admitted.lower().startswith("n"):
        click_radio(
            "100005",
            "dynamic-question-151",
            value_or("Parecer do Auditor - Selecione o desfecho assistencial * (cond.)", "Alta melhorada"),
        )
        set_input(
            "100005",
            'input[type="date"]',
            date_html(value("Parecer do Auditor - Data do desfecho * (cond.)")),
        )
        set_input(
            "100005",
            'input[type="time"]',
            value("Parecer do Auditor - Hora do desfecho * (cond.)"),
        )
    fill_auditor_history(
        "100005",
        value_or("Dados da Internação - Acomodação *", "Apartamento / Enfermaria"),
    )
    click_auditor_radio_label("100005", value_or("Parecer do Auditor - Programação de alta * (cond.)", "Sem programação de alta"))
    operator_pending = value("Parecer do Auditor - Pendências da operadora (cond.)")
    if operator_pending:
        for item in operator_pending.split(";"):
            click_checkbox_label("100005", item.strip())
        set_operator_pending_justification("100005", value("Parecer do Auditor - Justifique * (cond.)"))
    else:
        clear_operator_pending("100005")
    next_sec("100005")

    return finish_at_summary(missing_cid)


def main() -> int:
    today = dt.date.today().strftime("%d_%m_%Y")
    parser = argparse.ArgumentParser(description="Preenche evolucao clinica no Salus via HTML.")
    parser.add_argument("--fila", required=True)
    parser.add_argument("--clinica", required=True)
    parser.add_argument("--saida")
    parser.add_argument(
        "--controle-lancamento",
        default=f"exports/data_base_lancamento_{today}.xlsx",
    )
    parser.add_argument("--senha", required=True)
    parser.add_argument(
        "--forcar-reprocessamento",
        action="store_true",
        help="Permite repetir uma senha que ja possui tentativa registrada.",
    )
    parser.add_argument("--confirmar", action="store_true", help="Clica em Confirmar evolução se o botao estiver habilitado.")
    parser.add_argument(
        "--preencher-obrigatorios-medios",
        action="store_true",
        help="Preenche campos obrigatorios ausentes com valores medios/defaults antes de confirmar.",
    )
    parser.add_argument("--cdp-url", default=DEFAULT_CDP)
    args = parser.parse_args()

    patient, clinical_patient = find_patient(Path(args.fila), Path(args.clinica), args.senha)
    controle_path = Path(args.controle_lancamento)
    status_anterior = read_lancamento_status(controle_path, patient.senha)
    if status_anterior and not args.forcar_reprocessamento:
        raise RuntimeError(
            f"Senha {patient.senha} ja processada com status {status_anterior}; "
            "repeticao automatica bloqueada."
        )
    if not clinical_patient.id_internacao:
        raise RuntimeError(f"Paciente {args.senha} sem ID internacao na planilha clinica.")

    result = run_html_fill(
        clinical_patient,
        confirmar=args.confirmar,
        cdp_url=args.cdp_url,
        usar_defaults_obrigatorios=args.preencher_obrigatorios_medios,
    )
    status = "HTML_CONFIRMADO" if args.confirmar and result.get("confirmEnabled") else "HTML_PREENCHIDO"
    if args.confirmar and not result.get("confirmEnabled"):
        status = "ERRO"
    if result.get("confirmEnabled"):
        message = "Tela HTML preenchida. Botao Confirmar evolucao habilitado."
    elif result.get("incompleteSteps"):
        message = "Paginas sem check verde: " + ", ".join(result["incompleteSteps"])
    else:
        message = "Tela HTML preenchida, mas Confirmar evolucao ainda esta desabilitado."
    update_lancamento_control(Path(args.clinica), controle_path, patient, result, message)
    if args.saida:
        write_report(
            [
                PatientResult(
                    senha=patient.senha,
                    nome=patient.nome,
                    iniciais=patient.iniciais,
                    status=status,
                    mensagem=message,
                    campos_preenchidos=list(filled_values(clinical_patient).keys()),
                    campos_com_erro=[] if result.get("confirmEnabled") else ["Confirmar evolucao desabilitado na tela HTML."],
                )
            ],
            Path(args.saida),
        )
    print(json.dumps(result, ensure_ascii=False, indent=2))
    print(message)
    return 0 if status != "ERRO" else 1


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"Erro: {exc}", file=sys.stderr)
        raise SystemExit(1)
