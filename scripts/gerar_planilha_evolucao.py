#!/usr/bin/env python3
"""Gera uma planilha de preenchimento da evolucao clinica com campos em colunas."""

from __future__ import annotations

import argparse
import datetime as dt
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from salus_cdp import call_salus_api


def first_text(item: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = item.get(key)
        if value not in (None, ""):
            return str(value)
    return ""


def is_question(node: dict[str, Any]) -> bool:
    return bool(
        first_text(node, "descricao", "pergunta", "nome", "label", "titulo")
        and (
            "idFormularioPergunta" in node
            or "idPergunta" in node
            or "tipoPergunta" in node
            or "controle" in node
            or "tipoComponente" in node
        )
    )


def children_of(node: dict[str, Any]) -> list[Any]:
    children: list[Any] = []
    for key in (
        "perguntas",
        "campos",
        "itens",
        "filhos",
        "perguntasFilhas",
        "opcoes",
        "alternativas",
        "subPerguntas",
    ):
        value = node.get(key)
        if isinstance(value, list):
            children.extend(value)
    return children


def option_label(item: Any) -> str:
    if isinstance(item, dict):
        return first_text(item, "descricao", "nome", "label", "valor", "texto")
    return str(item)


def collect_fields(definition: Any) -> list[dict[str, Any]]:
    fields: list[dict[str, Any]] = []

    def walk(node: Any, section: str, parent: str = "") -> None:
        if isinstance(node, list):
            for child in node:
                walk(child, section, parent)
            return
        if not isinstance(node, dict):
            return

        current_section = first_text(node, "nomeSecao", "secao", "tituloSecao", "nome") or section
        if is_question(node):
            label = first_text(node, "descricao", "pergunta", "nome", "label", "titulo")
            required = bool(node.get("obrigatorio") or node.get("required") or node.get("obrigatoria"))
            control = first_text(node, "controle", "tipoControle", "tipoComponente", "componente")
            field_type = first_text(node, "tipo", "tipoPergunta", "tipoCampo")
            options = [
                option_label(option)
                for option in children_of({"opcoes": node.get("opcoes") or node.get("alternativas") or []})
            ]
            fields.append(
                {
                    "section": current_section,
                    "label": label,
                    "required": required,
                    "control": control,
                    "type": field_type,
                    "parent": parent,
                    "id": node.get("idFormularioPergunta") or node.get("idPergunta") or node.get("id"),
                    "options": [option for option in options if option],
                }
            )
            parent = label
        for child in children_of(node):
            walk(child, current_section, parent)

    sections = definition.get("secoes") if isinstance(definition, dict) else definition
    walk(sections, "")
    unique: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    for field in fields:
        key = (field["section"], field["label"], str(field["id"]))
        if key not in seen and field["label"]:
            unique.append(field)
            seen.add(key)
    return unique


def header_for(field: dict[str, Any]) -> str:
    header = f"{field['section']} - {field['label']}"
    if field["required"]:
        header += " *"
    if field.get("parent"):
        header += " (cond.)"
    return header


def save_workbook(fields: list[dict[str, Any]], output: Path, nome: str, iniciais: str, senha: str, internacao: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Preenchimento"
    patient_headers = ["Nome", "Iniciais", "Senha", "Dias internado", "ID internação"]
    headers = patient_headers + [header_for(field) for field in fields]
    ws.append(headers)
    ws.append([nome, iniciais, senha, "", internacao] + [""] * len(fields))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 48
    ws.freeze_panes = "F2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24

    campos = wb.create_sheet("Campos")
    campos.append(["Coluna", "Seção", "Campo", "Obrigatório", "Tipo", "Controle", "Campo pai", "ID formulário pergunta"])
    for field in fields:
        campos.append(
            [
                header_for(field),
                field["section"],
                field["label"],
                "Sim" if field["required"] else "Não",
                field["type"],
                field["control"],
                field["parent"],
                field["id"],
            ]
        )

    opcoes = wb.create_sheet("Opcoes")
    opcoes.append(["Seção", "Campo", "Opção"])
    for field in fields:
        for option in field["options"]:
            opcoes.append([field["section"], field["label"], option])

    resumo = wb.create_sheet("Resumo")
    resumo.append(["Item", "Valor"])
    resumo.append(["Campos", len(fields)])
    resumo.append(["Paciente", f"{nome} ({iniciais})"])
    resumo.append(["Internação", internacao])

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    today = dt.date.today().isoformat()
    parser = argparse.ArgumentParser(description="Gera planilha de evolucao clinica com campos em colunas.")
    parser.add_argument("--internacao", default="3956")
    parser.add_argument("--user-key", default="49")
    parser.add_argument("--nome", default="LUCIANA PELLACANI")
    parser.add_argument("--iniciais", default="LP")
    parser.add_argument("--senha", default="K3ZAVM6")
    parser.add_argument("--cdp-url", default="http://127.0.0.1:9222")
    parser.add_argument("--saida", default=f"exports/preenchimento_evolucao_clinica_LP_K3ZAVM6_{today}.xlsx")
    args = parser.parse_args()

    definition = call_salus_api(
        f"/api/formularios/auditoria/definicao?user_key={args.user_key}&idInternacao={args.internacao}",
        cdp_url=args.cdp_url,
    )
    fields = collect_fields(definition)
    save_workbook(fields, Path(args.saida), args.nome, args.iniciais, args.senha, args.internacao)
    print(f"Arquivo gerado: {args.saida}")
    print(f"Campos: {len(fields)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
