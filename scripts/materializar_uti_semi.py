#!/usr/bin/env python3
"""Materializa defaults obrigatorios de UTI para pacientes em UTI ou SEMI.

O lancador HTML tambem aplica esses defaults em memoria, mas gravar os valores
na planilha deixa a revisao humana e a validacao pre-lancamento mais claras.
"""

from __future__ import annotations

import argparse
import datetime as dt
import shutil
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from etapa2_lancar_evolucao_salus import read_clinical, value_to_text
from lancar_evolucao_html_salus import filled_values


ROOT = Path(__file__).resolve().parents[1]
EXPORTS = ROOT / "exports"
ARCHIVE = EXPORTS / "arquivo"

UTI_REQUIRED_FIELDS = [
    "UTI - 1. Abertura Ocular (E) - Selecione a melhor resposta observada. *",
    "UTI - 2. Resposta Verbal (V) - Avaliar conteúdo da comunicação verbal. *",
    "UTI - 3. Melhor Resposta Motora (M) - Registrar a melhor resposta obtida. *",
    "UTI - 4. Resposta Pupilar (P) - Avaliar reatividade pupilar ao estímulo luminoso. *",
    "UTI - Monitorização *",
    "UTI - Uso de droga vasoativa? *",
    "UTI - Categoria do diagnóstico principal *",
]

UTI_LAB_PAIRS = [
    ("UTI - Creatinina sérica (mg/dL) *", "UTI - Não mensurado * (cond.)"),
    ("UTI - pH arterial *", "UTI - Não mensurado * (cond.) [2]"),
    ("UTI - PaO2 (mmHg) *", "UTI - Não mensurado * (cond.) [3]"),
    ("UTI - FiO2 (%) *", "UTI - Não mensurado * (cond.) [4]"),
]


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def newest_base() -> Path:
    matches = sorted(
        EXPORTS.glob("data_base_lancar_*.xlsx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not matches:
        raise FileNotFoundError("Nenhuma base data_base_lancar_*.xlsx encontrada em exports.")
    return matches[0]


def headers_for(sheet) -> dict[str, int]:
    return {str(cell.value): cell.column for cell in sheet[1] if cell.value}


def is_uti_or_semi(value: Any) -> bool:
    text = value_to_text(value).casefold()
    return "uti" in text or "semi" in text


def missing_uti_fields(sheet, headers: dict[str, int], row: int) -> list[str]:
    missing: list[str] = []
    for field in UTI_REQUIRED_FIELDS:
        column = headers.get(field)
        if not column or is_blank(sheet.cell(row, column).value):
            missing.append(field)
    for lab_field, flag_field in UTI_LAB_PAIRS:
        lab_column = headers.get(lab_field)
        flag_column = headers.get(flag_field)
        lab_value = sheet.cell(row, lab_column).value if lab_column else None
        flag_value = sheet.cell(row, flag_column).value if flag_column else None
        if is_blank(lab_value) and value_to_text(flag_value).casefold() != "sim":
            missing.append(f"{lab_field} ou {flag_field}")
    return missing


def create_backup(path: Path) -> Path:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    timestamp = dt.datetime.now().strftime("%H%M")
    backup = ARCHIVE / f"{path.stem}_antes_materializar_uti_semi_{timestamp}{path.suffix}"
    if backup.exists():
        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = ARCHIVE / f"{path.stem}_antes_materializar_uti_semi_{timestamp}{path.suffix}"
    shutil.copy2(path, backup)
    return backup


def materialize(path: Path, *, overwrite: bool = False, dry_run: bool = False) -> dict[str, Any]:
    clinical_by_password, _, _ = read_clinical(path)
    clinical_unique = {
        senha: rows[0]
        for senha, rows in clinical_by_password.items()
        if len(rows) == 1
    }

    workbook = load_workbook(path)
    sheet = workbook["Preenchimento"] if "Preenchimento" in workbook.sheetnames else workbook.active
    headers = headers_for(sheet)
    senha_col = headers.get("Senha")
    accommodation_col = headers.get("Dados da Internação - Acomodação *")
    if not senha_col or not accommodation_col:
        workbook.close()
        raise RuntimeError("A base precisa conter as colunas Senha e Dados da Internação - Acomodação *.")

    uti_headers = [header for header in headers if header.startswith("UTI -")]
    rows_seen = 0
    rows_changed = 0
    cells_written = 0
    missing_after: list[tuple[str, list[str]]] = []

    for row in range(2, sheet.max_row + 1):
        senha = value_to_text(sheet.cell(row, senha_col).value)
        if not senha or not is_uti_or_semi(sheet.cell(row, accommodation_col).value):
            continue
        rows_seen += 1
        clinical_patient = clinical_unique.get(senha)
        if not clinical_patient:
            missing_after.append((senha, ["Senha sem linha clinica unica."]))
            continue
        defaults = filled_values(clinical_patient)
        row_writes = 0
        for header in uti_headers:
            value = defaults.get(header)
            if is_blank(value):
                continue
            cell = sheet.cell(row, headers[header])
            if overwrite or is_blank(cell.value):
                if cell.value != value:
                    row_writes += 1
                    cells_written += 1
                    if not dry_run:
                        cell.value = value
        if row_writes:
            rows_changed += 1
        missing = missing_uti_fields(sheet, headers, row)
        if missing:
            missing_after.append((senha, missing))

    if not dry_run:
        workbook.save(path)
    workbook.close()

    return {
        "arquivo": path,
        "linhas_uti_semi": rows_seen,
        "linhas_alteradas": rows_changed,
        "celulas_preenchidas": cells_written,
        "faltas": missing_after,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Preenche campos obrigatorios de UTI em pacientes com acomodacao UTI ou SEMI."
    )
    parser.add_argument(
        "base",
        nargs="?",
        type=Path,
        default=None,
        help="Base clinica .xlsx. Se omitida, usa a data_base_lancar_*.xlsx mais recente em exports.",
    )
    parser.add_argument(
        "--sobrescrever",
        action="store_true",
        help="Atualiza tambem campos UTI ja preenchidos. Por padrao, so preenche vazios.",
    )
    parser.add_argument(
        "--sem-backup",
        action="store_true",
        help="Nao cria copia em exports/arquivo antes de salvar.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Mostra o que seria alterado sem salvar a planilha.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    base = args.base or newest_base()
    if not base.exists():
        raise FileNotFoundError(f"Base nao encontrada: {base}")

    backup = None
    if not args.dry_run and not args.sem_backup:
        backup = create_backup(base)

    result = materialize(base, overwrite=args.sobrescrever, dry_run=args.dry_run)
    if backup:
        print(f"Backup: {backup}")
    print(f"Arquivo: {result['arquivo']}")
    print(f"Linhas UTI/SEMI: {result['linhas_uti_semi']}")
    print(f"Linhas alteradas: {result['linhas_alteradas']}")
    print(f"Celulas UTI preenchidas: {result['celulas_preenchidas']}")
    print(f"Linhas UTI/SEMI com falta: {len(result['faltas'])}")
    for senha, missing in result["faltas"][:20]:
        print(f"FALTA {senha}: {'; '.join(missing)}")
    return 1 if result["faltas"] else 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"Erro: {exc}", file=sys.stderr)
        raise SystemExit(1)
