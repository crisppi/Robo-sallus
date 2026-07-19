#!/usr/bin/env python3
"""Tela web local do Robo Sallus.

Este app nao depende de tkinter. Ele abre uma pagina local no navegador com:
- botao Novo dia, Etapa 1 e lancamento automatico no Salus
- cards de contagem
- paciente/senha em processamento
- log de execucao
"""

from __future__ import annotations

import datetime as dt
import json
import subprocess
import sys
import threading
import time
import webbrowser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse
from openpyxl import load_workbook

from etapa2_lancar_evolucao_salus import (
    PatientResult,
    QueuePatient,
    process_patients,
    read_clinical,
    read_queue,
    read_successful_passwords,
    value_to_text,
    write_report,
)
from salus_cdp import SalusCdpError, navigate_salus, start_salus_chrome


ROOT = Path(__file__).resolve().parents[1]
EXPORTS = ROOT / "exports"
HOST = "127.0.0.1"
PORT = 8765


def newest(patterns: str | tuple[str, ...], fallback: str) -> Path:
    if isinstance(patterns, str):
        patterns = (patterns,)
    matches = []
    for pattern in patterns:
        matches.extend(EXPORTS.glob(pattern))
    matches = sorted(set(matches), key=lambda p: p.stat().st_mtime, reverse=True)
    return matches[0] if matches else EXPORTS / fallback


def default_state() -> dict:
    today = dt.date.today().strftime("%d_%m_%Y")
    return {
        "running": False,
        "status": "Pronto",
        "current_patient": "Nenhum paciente em execucao",
        "current_password": "-",
        "processed_now": 0,
        "launch_rows": [],
        "pending_rows": [],
        "logs": [],
        "files": {
            "fila": str(newest(("fila_salus_*.xlsx", "pacientes_sirio_libanes_*.xlsx"), "fila_salus_DATA.xlsx")),
            "clinica": str(newest(("data_base_lancar_*.xlsx", "data_base_lancar-*.xlsx", "data_base_lancamento_*.xlsx", "data_base_lantamento_*.xlsx", "preenchimento_evolucao_clinica_*_colorido.xlsx"), "data_base_lancamento_DATA.xlsx")),
            "relatorio": str(EXPORTS / f"relatorio_lancamentos_{today}.xlsx"),
        },
        "cards": {
            "salus": "-",
            "excel": "-",
            "encontrados": "-",
            "faltam": "-",
        },
    }


STATE = default_state()
LOCK = threading.Lock()


def log(message: str) -> None:
    timestamp = dt.datetime.now().strftime("%H:%M:%S")
    with LOCK:
        STATE["logs"].append(f"[{timestamp}] {message}")
        STATE["logs"] = STATE["logs"][-300:]


def set_state(**kwargs) -> None:
    with LOCK:
        STATE.update(kwargs)


def persist_patient_status(clinica: Path, patient: QueuePatient, result: PatientResult) -> None:
    """Grava cada resultado imediatamente para impedir relançamento da senha."""
    if result.status in {"PULADO", "JA_LANCADO"}:
        return
    wb = load_workbook(clinica)
    ws = wb["Preenchimento"] if "Preenchimento" in wb.sheetnames else wb.active
    headers = {str(cell.value): cell.column for cell in ws[1] if cell.value}
    senha_col = headers.get("Senha")
    status_col = headers.get("Lançamento Salus - Status")
    data_col = headers.get("Lançamento Salus - Data/hora")
    mensagem_col = headers.get("Lançamento Salus - Mensagem")
    if not senha_col or not status_col:
        wb.close()
        raise RuntimeError("Base sem colunas de controle do lançamento.")
    if result.status in {"SUCESSO", "SUCESSO_COM_ALERTA", "SUCESSO_MANUAL"}:
        status = "FINALIZADO"
    elif result.status in {"AGUARDANDO", "AGUARDANDO_CID"}:
        status = result.status
    elif result.status == "PRE_LANCADO":
        status = "PRE_LANCADO"
    else:
        status = "ERRO"
    for row in range(2, ws.max_row + 1):
        if value_to_text(ws.cell(row, senha_col).value) == patient.senha:
            ws.cell(row, status_col).value = status
            if data_col:
                ws.cell(row, data_col).value = dt.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            if mensagem_col:
                ws.cell(row, mensagem_col).value = result.mensagem or result.status
            wb.save(clinica)
            wb.close()
            return
    wb.close()
    raise RuntimeError(f"Senha {patient.senha} não encontrada na base de controle.")


def calculate_cards() -> dict:
    files = STATE["files"]
    fila = Path(files["fila"])
    clinica = Path(files["clinica"])
    relatorio = Path(files["relatorio"])

    queue_patients = read_queue(fila) if fila.exists() else []
    clinical_by_password, _, _ = read_clinical(clinica) if clinica.exists() else ({}, {}, [])
    clinical_unique = {senha for senha, rows in clinical_by_password.items() if len(rows) == 1}
    queue_passwords = [patient.senha for patient in queue_patients]
    found = sum(1 for senha in queue_passwords if senha in clinical_unique)
    finalized_statuses = {
        "FINALIZADO", "SUCESSO", "SUCESSO_COM_ALERTA", "SUCESSO_MANUAL", "JA_LANCADO"
    }
    eligible_passwords = {
        senha
        for senha in queue_passwords
        if senha in clinical_by_password
        and len(clinical_by_password[senha]) == 1
        and value_to_text(clinical_by_password[senha][0].values.get("evolucao")).strip()
    }
    finalized_passwords = {
        senha
        for senha in eligible_passwords
        if value_to_text(
            clinical_by_password[senha][0].values.get("Lançamento Salus - Status")
        ).strip().upper() in finalized_statuses
    }
    processed = len(finalized_passwords)
    missing_to_launch = len(eligible_passwords - finalized_passwords)

    return {
        "salus": len(queue_patients),
        "excel": sum(len(rows) for rows in clinical_by_password.values()),
        "encontrados": found,
        "faltam": missing_to_launch,
        "processados": processed,
    }


def refresh_cards(update_status: bool = True) -> dict:
    cards = calculate_cards()
    files = STATE["files"]
    fila = Path(files["fila"])
    clinica = Path(files["clinica"])
    queue_by_password = {patient.senha: patient for patient in read_queue(fila)} if fila.exists() else {}
    clinical_by_password, _, _ = read_clinical(clinica) if clinica.exists() else ({}, {}, [])
    pending_rows = []
    for senha, rows in clinical_by_password.items():
        if len(rows) != 1:
            continue
        values = rows[0].values
        status = value_to_text(values.get("Lançamento Salus - Status")).strip()
        if status not in {"AGUARDANDO", "AGUARDANDO_CID", "ERRO", "PRE_LANCADO"}:
            continue
        patient = queue_by_password.get(senha)
        pending_rows.append({
            "nome": (patient.nome if patient else rows[0].nome) or "-",
            "senha": senha,
            "iniciais": (patient.iniciais if patient else rows[0].iniciais) or "-",
            "status": status,
            "mensagem": value_to_text(values.get("Lançamento Salus - Mensagem")) or "Revisão necessária.",
        })
    with LOCK:
        STATE["cards"] = cards
        STATE["processed_now"] = cards.get("processados", 0)
        STATE["pending_rows"] = pending_rows
        if update_status:
            STATE["status"] = "Cards atualizados"
    return cards


def calculate_daily_control() -> dict:
    """Monta o painel diário sem alterar a planilha clínica."""
    files = STATE["files"]
    fila = Path(files["fila"])
    clinica = Path(files["clinica"])
    queue = read_queue(fila) if fila.exists() else []
    clinical, _, _ = read_clinical(clinica) if clinica.exists() else ({}, {}, [])
    queue_by_password = {patient.senha: patient for patient in queue}
    passwords = list(queue_by_password)
    passwords.extend(senha for senha in clinical if senha not in queue_by_password)
    today_br = dt.date.today().strftime("%d/%m/%Y")
    rows = []
    counts = {"vermelho": 0, "amarelo": 0, "azul": 0, "verde": 0, "cinza": 0}
    for senha in passwords:
        base_rows = clinical.get(senha, [])
        row = base_rows[0] if len(base_rows) == 1 else None
        values = row.values if row else {}
        patient = queue_by_password.get(senha)
        evolution = value_to_text(values.get("evolucao")).strip()
        evolution_date = value_to_text(values.get("Data da evolução")).strip()
        responsible = value_to_text(values.get("Responsável pelo preenchimento")).strip()
        launch_status = value_to_text(values.get("Lançamento Salus - Status")).strip().upper()
        launch_date = value_to_text(values.get("Lançamento Salus - Data/hora")).strip()
        if patient is None:
            color, label = "cinza", "Não está mais internado"
        elif launch_status in {"FINALIZADO", "SUCESSO", "SUCESSO_MANUAL", "SUCESSO_COM_ALERTA"} and launch_date.startswith(today_br):
            color, label = "verde", "Confirmado hoje no Salus"
        elif launch_status == "PRE_LANCADO":
            color, label = "azul", "Pré-lançado"
        elif evolution and evolution_date == today_br:
            color, label = "amarelo", "Evolução de hoje aguardando revisão"
        else:
            color, label = "vermelho", "Evolução de hoje ausente"
        counts[color] += 1
        rows.append({
            "nome": (patient.nome if patient else (row.nome if row else "")) or "-",
            "senha": senha,
            "iniciais": (patient.iniciais if patient else (row.iniciais if row else "")) or "-",
            "data_evolucao": evolution_date or "Sem data",
            "responsavel": responsible or "-",
            "status": label,
            "cor": color,
        })
    return {"data": today_br, "counts": counts, "rows": rows}


def run_etapa1_worker() -> None:
    try:
        output = EXPORTS / f"pacientes_sirio_libanes_{dt.date.today().isoformat()}.xlsx"
        cmd = [sys.executable, str(ROOT / "scripts" / "gerar_lista_pacientes.py"), "--saida", str(output)]
        log("Etapa 1 iniciada: baixar fila Salus.")
        completed = subprocess.run(cmd, cwd=ROOT, capture_output=True, text=True, check=False)
        if completed.stdout.strip():
            log(completed.stdout.strip())
        if completed.stderr.strip():
            log(completed.stderr.strip())
        if completed.returncode != 0:
            set_state(status=f"Etapa 1 falhou com codigo {completed.returncode}")
            log(f"ERRO: Etapa 1 falhou com codigo {completed.returncode}.")
        else:
            with LOCK:
                STATE["files"]["fila"] = str(output)
            refresh_cards()
            set_state(status="Etapa 1 concluida")
            log(f"Etapa 1 concluida. Arquivo: {output}")
    except Exception as exc:
        set_state(status="Erro na Etapa 1")
        log(f"ERRO: {exc}")
    finally:
        set_state(running=False)


def run_new_day_worker() -> None:
    try:
        date_label = dt.date.today().strftime("%d_%m_%Y")
        queue_output = EXPORTS / f"fila_salus_{date_label}.xlsx"
        clinical_output = EXPORTS / f"data_base_lancar_{date_label}.xlsx"
        report_output = EXPORTS / f"relatorio_lancamentos_{date_label}.xlsx"
        cmd = [sys.executable, str(ROOT / "scripts" / "atualizar_novo_dia.py")]

        log("Novo dia iniciado: baixando a fila antes de arquivar os arquivos atuais.")
        completed = subprocess.run(cmd, cwd=ROOT, capture_output=True, text=True, check=False)
        if completed.stdout.strip():
            for line in completed.stdout.strip().splitlines():
                log(line)
        if completed.stderr.strip():
            lines = [line.strip() for line in completed.stderr.splitlines() if line.strip()]
            log(lines[-1] if lines else "Erro ao preparar o novo dia.")
        if completed.returncode != 0:
            set_state(status=f"Novo dia falhou com codigo {completed.returncode}")
            log("ERRO: nenhum arquivo foi substituido se a fila nao foi baixada.")
            return

        with LOCK:
            STATE["files"] = {
                "fila": str(queue_output),
                "clinica": str(clinical_output),
                "relatorio": str(report_output),
            }
            STATE["processed_now"] = 0
            STATE["current_patient"] = "Nenhum paciente em execucao"
            STATE["current_password"] = "-"
            STATE["launch_rows"] = []
        refresh_cards()
        set_state(status="Novo dia concluido")
        log(f"Novo dia pronto: {len(read_queue(queue_output))} pacientes.")
    except Exception as exc:
        set_state(status="Erro ao iniciar novo dia")
        log(f"ERRO: {exc}")
    finally:
        set_state(running=False)


def run_etapa2_worker(
    only_password: str | None,
    batch_limit: int | None,
    retry_errors: bool = False,
    confirmar: bool = True,
) -> None:
    try:
        files = STATE["files"].copy()
        fila = Path(files["fila"])
        clinica = Path(files["clinica"])
        relatorio = Path(files["relatorio"])
        log("Etapa 2 iniciada: lancamento automatico direto no Salus.")

        queue_patients = read_queue(fila)
        clinical_by_password, field_meta, field_headers = read_clinical(clinica)
        successful_passwords = read_successful_passwords(relatorio)
        attempted_passwords: set[str] = set(successful_passwords)
        error_passwords: set[str] = set()
        for senha, rows in clinical_by_password.items():
            statuses = {
                value_to_text(row.values.get("Lançamento Salus - Status")).strip().upper()
                for row in rows
            }
            if "ERRO" in statuses:
                error_passwords.add(senha)
            if any(statuses):
                attempted_passwords.add(senha)

        if retry_errors:
            # Um erro anterior não significa lançamento concluído. Retira apenas
            # esses registros da trava de duplicidade e mantém finalizados protegidos.
            attempted_passwords.difference_update(error_passwords)

        # Somente pacientes com evolução textual, ainda não tentados, entram no lote.
        eligible_passwords = {
            senha
            for senha, rows in clinical_by_password.items()
            if len(rows) == 1
            and value_to_text(rows[0].values.get("evolucao")).strip()
            and senha not in attempted_passwords
            and (not retry_errors or senha in error_passwords)
        }
        queue_patients = [
            patient for patient in queue_patients
            if patient.senha in eligible_passwords
            and (not only_password or patient.senha == only_password)
        ]
        if batch_limit is not None:
            queue_patients = queue_patients[:batch_limit]
        if not queue_patients:
            raise RuntimeError(
                "Nenhum paciente pendente com evolução foi encontrado na Planilha Clínica selecionada."
            )
        target_passwords = {
            patient.senha
            for patient in queue_patients
            if not only_password or patient.senha == only_password
        }
        missing_ids = sorted(
            senha
            for senha in target_passwords
            if senha in clinical_by_password
            and len(clinical_by_password[senha]) == 1
            and not clinical_by_password[senha][0].id_internacao
        )
        if missing_ids:
            sample = ", ".join(missing_ids[:5])
            raise RuntimeError(
                f"A base clinica possui {len(missing_ids)} paciente(s) sem ID de internacao "
                f"({sample}). Clique em 'Novo dia' com o Salus aberto e autenticado para "
                "regenerar a base antes do lancamento."
            )
        processed_count = 0
        started_at: dict[str, dt.datetime] = {}

        def progress(event: str, patient: QueuePatient, result: PatientResult | None) -> None:
            nonlocal processed_count
            if event == "inicio":
                started_at[patient.senha] = dt.datetime.now()
                with LOCK:
                    STATE["current_patient"] = patient.nome or "-"
                    STATE["current_password"] = patient.senha
                    STATE["launch_rows"].append({
                        "nome": patient.nome or "-",
                        "senha": patient.senha,
                        "iniciais": patient.iniciais or "-",
                        "inicio": started_at[patient.senha].isoformat(),
                        "tempo": "0s",
                        "situacao": "Em lançamento",
                    })
            elif event == "fim" and result:
                # Ritmo operacional mínimo por paciente. Durante a espera o
                # painel permanece como "Em lançamento" e mostra o cronômetro.
                started = started_at.get(patient.senha, dt.datetime.now())
                elapsed_before_finish = (dt.datetime.now() - started).total_seconds()
                if elapsed_before_finish < 50:
                    time.sleep(50 - elapsed_before_finish)
                persist_patient_status(clinica, patient, result)
                processed_count += 1
                # Recalcula também a lista de pendentes imediatamente. Assim,
                # um erro corrigido desaparece da tela sem refresh manual.
                current_cards = refresh_cards(update_status=False)
                elapsed = dt.datetime.now() - started_at.get(patient.senha, dt.datetime.now())
                total_seconds = max(0, int(elapsed.total_seconds()))
                elapsed_text = f"{total_seconds // 60}m {total_seconds % 60:02d}s" if total_seconds >= 60 else f"{total_seconds}s"
                with LOCK:
                    STATE["cards"] = current_cards
                    STATE["processed_now"] = current_cards.get("processados", 0)
                    for row in reversed(STATE["launch_rows"]):
                        if row["senha"] == patient.senha and row["situacao"] == "Em lançamento":
                            row["tempo"] = elapsed_text
                            row["situacao"] = (
                                "Pré-lançado" if result.status == "PRE_LANCADO"
                                else ("Concluído" if result.status not in {"ERRO", "PULADO"} else result.status)
                            )
                            break
                log(f"{result.senha} - {result.nome} - {result.status}: {result.mensagem}")

        results = process_patients(
            queue_patients=queue_patients,
            clinical_by_password=clinical_by_password,
            field_meta=field_meta,
            field_headers=field_headers,
            successful_passwords=attempted_passwords,
            dry_run=False,
            only_password=only_password,
            usar_defaults_obrigatorios=True,
            # O lote nunca para por erro individual: registra ERRO no paciente
            # e segue para o próximo. Finalizados continuam protegidos.
            stop_on_error=False,
            confirmar=confirmar,
            progress_callback=progress,
        )
        write_report(results, relatorio)
        refresh_cards()
        set_state(status=f"Etapa 2 concluida. Relatorio: {relatorio}")
        log(f"Etapa 2 concluida. Relatorio: {relatorio}")
    except Exception as exc:
        set_state(status="Erro na Etapa 2")
        log(f"ERRO: {exc}")
    finally:
        set_state(running=False)


HTML = r"""<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Robo Sallus</title>
  <style>
    :root {
      --bg: #f4f6f8;
      --panel: #ffffff;
      --ink: #1f2937;
      --muted: #667085;
      --blue: #1f4e78;
      --line: #d8dee6;
      --danger: #9b1c1c;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      color: var(--ink);
      background: var(--bg);
    }
    main { max-width: 1220px; margin: 0 auto; padding: 24px; }
    .page-head { display: flex; align-items: center; justify-content: space-between; gap: 16px; }
    h1 { margin: 0; font-size: 32px; letter-spacing: 0; }
    .robot-stamp {
      border: 3px solid #6b7280;
      border-radius: 7px;
      color: #6b7280;
      background: #fff;
      padding: 8px 15px;
      font-size: 14px;
      font-weight: 900;
      letter-spacing: 1.4px;
      text-transform: uppercase;
      transform: rotate(-2deg);
      box-shadow: inset 0 0 0 2px #fff, 0 2px 6px rgba(0,0,0,.08);
    }
    .robot-stamp.running {
      color: #166534;
      border-color: #16a34a;
      background: #f0fdf4;
      animation: stampPulse 1.15s ease-in-out infinite;
    }
    @keyframes stampPulse {
      0%, 100% { opacity: 1; transform: rotate(-2deg) scale(1); }
      50% { opacity: .68; transform: rotate(-2deg) scale(1.035); }
    }
    .sub { color: var(--muted); margin: 6px 0 20px; }
    .nav-link { color: var(--blue); font-weight: 800; text-decoration: none; }
    .panel {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 16px;
      margin-bottom: 16px;
    }
    .files {
      display: grid;
      grid-template-columns: 140px 1fr;
      gap: 5px 10px;
      align-items: center;
      padding: 9px 14px;
      margin-bottom: 10px;
    }
    .files label { font-size: 13px; }
    .files input[type="text"] {
      padding: 6px 10px;
      font-size: 13px;
      height: 32px;
    }
    label { font-weight: 650; color: #374151; }
    input[type="text"] {
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 10px 12px;
      font-size: 14px;
      background: #fff;
    }
    .actions { display: flex; gap: 10px; align-items: center; flex-wrap: wrap; }
    button {
      border: 0;
      border-radius: 6px;
      padding: 11px 16px;
      font-size: 15px;
      font-weight: 750;
      cursor: pointer;
      color: #fff;
      background: var(--blue);
    }
    button.secondary { background: #4b5563; }
    button:disabled { opacity: .55; cursor: wait; }
    .check { color: var(--danger); font-weight: 750; display: flex; gap: 8px; align-items: center; }
    .cards { display: grid; grid-template-columns: repeat(5, 1fr); gap: 12px; }
    .card {
      background: #fff;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 14px;
      min-height: 92px;
    }
    .card .title { color: var(--muted); font-size: 13px; }
    .card .value { color: var(--blue); font-weight: 800; font-size: 30px; margin-top: 8px; }
    .current {
      background: #eaf2f8;
      border: 1px solid #c7d9e8;
      border-radius: 8px;
      padding: 12px 14px;
      margin-bottom: 16px;
    }
    .launch-title { color: #506070; font-size: 13px; font-weight: 750; margin-bottom: 7px; }
    .launch-row { display: grid; grid-template-columns: minmax(260px, 1fr) 120px 100px 120px 110px; gap: 10px; padding: 7px 4px; border-top: 1px solid #c7d9e8; align-items: center; }
    .launch-row:first-child { border-top: 0; }
    .launch-name { color: #17324d; font-weight: 750; }
    .launch-meta { color: #506070; font-size: 13px; }
    .launch-status { color: #166534; font-weight: 800; }
    .launch-status.running { color: #1f4e78; }
    .launch-status.error { color: var(--danger); }
    .pending-list { margin-bottom: 16px; }
    .pending-row { display: grid; grid-template-columns: minmax(220px, 1fr) 110px 85px 150px minmax(260px, 2fr); gap: 10px; padding: 8px 4px; border-top: 1px solid var(--line); align-items: center; font-size: 13px; }
    .pending-row:first-child { border-top: 0; }
    .pending-status { color: #9a6700; font-weight: 800; }
    .status { color: var(--muted); margin-bottom: 8px; }
    pre {
      margin: 0;
      min-height: 300px;
      max-height: 420px;
      overflow: auto;
      background: #111827;
      color: #e5e7eb;
      border-radius: 8px;
      padding: 14px;
      white-space: pre-wrap;
      font: 13px/1.45 Menlo, Consolas, monospace;
    }
    @media (max-width: 900px) {
      .cards { grid-template-columns: repeat(2, 1fr); }
      .launch-row { grid-template-columns: 1fr 1fr; }
      .files { grid-template-columns: 1fr; }
    }
  </style>
</head>
<body>
  <main>
    <div class="page-head">
      <h1>Robo Sallus</h1>
      <div id="robot_stamp" class="robot-stamp">ROBÔ PARADO</div>
    </div>
    <p class="sub">Novo dia arquiva o lote anterior e prepara as listas. Etapa 2 lança automaticamente as evoluções no Salus. &nbsp; <a class="nav-link" href="/pre-lancamentos">Pré-lançamentos</a> &nbsp; <a class="nav-link" href="/controle-diario">Controle diário →</a></p>

    <section class="panel files">
      <label>Fila Salus</label>
      <input id="fila" type="text">
      <label>Planilha Clínica</label>
      <input id="clinica" type="text">
      <label>Relatório</label>
      <input id="relatorio" type="text">
    </section>

    <section class="panel actions">
      <button id="novoDia" onclick="startNewDay()">Novo dia</button>
      <button id="etapa1" onclick="startEtapa1()">Etapa 1: Baixar Fila Salus</button>
      <button id="etapa2" onclick="startEtapa2()">Etapa 2: Lançar Automaticamente no Salus</button>
      <button id="limparSalus" class="secondary" onclick="clearSalus()">Limpar tela do Salus</button>
      <button class="secondary" onclick="refreshCards()">Atualizar Cards</button>
      <label style="margin-left: 8px;">Somente senha</label>
      <input id="senha" type="text" style="width: 140px;">
      <label>Lote</label>
      <select id="batch_limit" style="height: 39px; border: 1px solid var(--line); border-radius: 6px; padding: 0 10px; background: #fff; font-weight: 700;">
        <option value="3">3 pacientes</option>
        <option value="10">10 pacientes</option>
        <option value="20">20 pacientes</option>
        <option value="all">Todos</option>
      </select>
    </section>

    <section class="cards">
      <div class="card"><div class="title">Pacientes no Salus</div><div id="salus" class="value">-</div></div>
      <div class="card"><div class="title">Pacientes no Excel</div><div id="excel" class="value">-</div></div>
      <div class="card"><div class="title">Senhas encontradas</div><div id="encontrados" class="value">-</div></div>
      <div class="card"><div class="title">Faltam lançar</div><div id="faltam" class="value">-</div></div>
      <div class="card"><div class="title">Processados agora</div><div id="processed" class="value">0</div></div>
    </section>

    <section class="current">
      <div class="launch-title">Pacientes em lançamento</div>
      <div id="launch_rows"><div class="launch-meta">Nenhum paciente em execução</div></div>
    </section>

    <section class="panel pending-list">
      <div class="launch-title">Erros para revisar</div>
      <div id="error_rows"><div class="launch-meta">Nenhum erro registrado</div></div>
    </section>

    <div id="status" class="status">Pronto</div>
    <pre id="logs"></pre>
  </main>

  <script>
    async function api(path, options = {}) {
      const response = await fetch(path, options);
      if (!response.ok) throw new Error(await response.text());
      return await response.json();
    }

    function readFiles() {
      return {
        fila: document.getElementById('fila').value,
        clinica: document.getElementById('clinica').value,
        relatorio: document.getElementById('relatorio').value
      };
    }

    async function saveFiles() {
      await api('/api/files', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify(readFiles())
      });
    }

    async function refreshCards() {
      await saveFiles();
      await api('/api/refresh', {method: 'POST'});
      await poll();
    }

    async function startEtapa1() {
      await saveFiles();
      await api('/api/etapa1', {method: 'POST'});
      await poll();
    }

    async function startNewDay() {
      if (!confirm('Iniciar novo dia? As planilhas atuais serão movidas para a pasta de arquivo após a nova fila ser baixada.')) return;
      await api('/api/novo-dia', {method: 'POST'});
      await poll();
    }

    async function startEtapa2() {
      await saveFiles();
      const senha = document.getElementById('senha').value.trim();
      const batchValue = document.getElementById('batch_limit').value;
      const batchLimit = batchValue === 'all' ? null : Number(batchValue);
      const escopo = senha ? ` somente para a senha ${senha}` : ' para todos os pacientes pendentes';
      const lote = batchLimit ? `, limitado a ${batchLimit} paciente(s)` : '';
      if (!confirm(`Confirma o lançamento automático direto no Salus${escopo}${lote}?`)) return;
      await api('/api/etapa2', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({
          only_password: senha,
          batch_limit: batchLimit
        })
      });
      await poll();
    }

    async function clearSalus() {
      if (!confirm('Limpar a tela do paciente atual e voltar para a lista de internações?')) return;
      await api('/api/limpar-salus', {method: 'POST'});
      await poll();
    }

    async function poll() {
      const state = await api('/api/status');
      document.getElementById('fila').value = state.files.fila;
      document.getElementById('clinica').value = state.files.clinica;
      document.getElementById('relatorio').value = state.files.relatorio;
      document.getElementById('salus').textContent = state.cards.salus;
      document.getElementById('excel').textContent = state.cards.excel;
      document.getElementById('encontrados').textContent = state.cards.encontrados;
      document.getElementById('faltam').textContent = state.cards.faltam;
      document.getElementById('processed').textContent = state.processed_now;
      const robotStamp = document.getElementById('robot_stamp');
      robotStamp.textContent = state.running ? 'ROBÔ EM EXECUÇÃO' : 'ROBÔ PARADO';
      robotStamp.classList.toggle('running', state.running);
      const launchRows = document.getElementById('launch_rows');
      if (!state.launch_rows.length) {
        launchRows.innerHTML = '<div class="launch-meta">Nenhum paciente em execução</div>';
      } else {
        const now = Date.now();
        launchRows.innerHTML = state.launch_rows.map(row => {
          let tempo = row.tempo;
          if (row.situacao === 'Em lançamento') {
            const seconds = Math.max(0, Math.floor((now - new Date(row.inicio).getTime()) / 1000));
            tempo = seconds >= 60 ? `${Math.floor(seconds / 60)}m ${String(seconds % 60).padStart(2, '0')}s` : `${seconds}s`;
          }
          const statusClass = row.situacao === 'Em lançamento' ? 'running' : (row.situacao === 'Concluído' ? '' : 'error');
          return `<div class="launch-row"><div class="launch-name">${escapeHtml(row.nome)}</div><div class="launch-meta">Senha: ${escapeHtml(row.senha)}</div><div class="launch-meta">Iniciais: ${escapeHtml(row.iniciais)}</div><div class="launch-meta">${tempo}</div><div class="launch-status ${statusClass}">${row.situacao}</div></div>`;
        }).join('');
      }
      const errors = state.pending_rows.filter(row => row.status === 'ERRO');
      const errorRows = document.getElementById('error_rows');
      errorRows.innerHTML = errors.length ? errors.map(row =>
        `<div class="pending-row"><div class="launch-name">${escapeHtml(row.nome)}</div><div>Senha: ${escapeHtml(row.senha)}</div><div>${escapeHtml(row.iniciais)}</div><div class="launch-status error">ERRO</div><div>${escapeHtml(row.mensagem)}</div></div>`
      ).join('') : '<div class="launch-meta">Nenhum erro registrado</div>';
      document.getElementById('status').textContent = state.status;
      document.getElementById('logs').textContent = state.logs.join('\n');
      document.getElementById('novoDia').disabled = state.running;
      document.getElementById('etapa1').disabled = state.running;
      document.getElementById('etapa2').disabled = state.running;
      document.getElementById('limparSalus').disabled = state.running;
    }

    function escapeHtml(value) {
      return String(value ?? '').replace(/[&<>"']/g, char => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#039;'}[char]));
    }

    setInterval(poll, 1200);
    poll();
  </script>
</body>
</html>
"""

PRE_HTML = r"""<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Pré-lançamentos - Robo Sallus</title>
  <style>
    * { box-sizing: border-box; }
    body { margin: 0; font-family: -apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif; color:#1f2937; background:#f4f6f8; }
    main { max-width:1220px; margin:0 auto; padding:24px; }
    .head { display:flex; justify-content:space-between; align-items:center; gap:16px; margin-bottom:20px; }
    h1 { margin:0; font-size:30px; }
    a { color:#1f4e78; font-weight:800; text-decoration:none; }
    .summary { display:flex; gap:12px; margin-bottom:16px; }
    .actions { display:flex; gap:10px; align-items:center; flex-wrap:wrap; margin-bottom:16px; background:#fff; border:1px solid #d8dee6; border-radius:8px; padding:14px; }
    button { border:0; border-radius:6px; padding:11px 16px; color:#fff; background:#1f4e78; font-weight:800; cursor:pointer; }
    button:disabled { opacity:.55; cursor:wait; }
    input,select { height:39px; border:1px solid #d8dee6; border-radius:6px; padding:0 10px; background:#fff; }
    .count { background:#fff; border:1px solid #d8dee6; border-radius:8px; padding:12px 16px; font-weight:800; }
    .panel { background:#fff; border:1px solid #d8dee6; border-radius:8px; padding:16px; margin-bottom:16px; }
    h2 { margin:0 0 10px; font-size:19px; }
    .row { display:grid; grid-template-columns:minmax(230px,1fr) 120px 100px 170px minmax(280px,2fr); gap:10px; padding:10px 4px; border-top:1px solid #e5e7eb; align-items:center; font-size:13px; }
    .row:first-child { border-top:0; }
    .name { font-weight:800; color:#17324d; }
    .status { font-weight:900; color:#9a6700; }
    .error .status { color:#9b1c1c; }
    .empty { color:#667085; padding:8px 0; }
    @media(max-width:800px){ .row{grid-template-columns:1fr 1fr}.summary{flex-wrap:wrap} }
  </style>
</head>
<body><main>
  <div class="head"><h1>Pré-lançamentos e erros</h1><div><a href="/">← Lançamentos</a> &nbsp; <a href="/controle-diario">Controle diário →</a></div></div>
  <div class="actions">
    <button id="preButton" onclick="startPreLaunch()">Iniciar pré-lançamento</button>
    <label>Somente senha</label><input id="preSenha" type="text" style="width:145px">
    <label>Lote</label><select id="preLimit"><option value="3">3 pacientes</option><option value="10">10 pacientes</option><option value="20">20 pacientes</option><option value="all">Todos</option></select>
  </div>
  <div class="summary"><div class="count">Aguardando finalização: <span id="waiting_count">0</span></div><div class="count">Com erro: <span id="error_count">0</span></div></div>
  <section class="panel"><h2>Aguardando finalização</h2><div id="waiting"></div></section>
  <section class="panel"><h2>Erros para revisar</h2><div id="errors"></div></section>
</main><script>
  const esc = value => String(value ?? '').replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#039;'}[c]));
  const render = (rows, css='') => rows.length ? rows.map(row => `<div class="row ${css}"><div class="name">${esc(row.nome)}</div><div>Senha: ${esc(row.senha)}</div><div>${esc(row.iniciais)}</div><div class="status">${esc(row.status)}</div><div>${esc(row.mensagem)}</div></div>`).join('') : '<div class="empty">Nenhum registro.</div>';
  async function poll(){
    const state = await fetch('/api/status').then(r => r.json());
    const waiting = state.pending_rows.filter(r => r.status !== 'ERRO');
    const errors = state.pending_rows.filter(r => r.status === 'ERRO');
    document.getElementById('waiting_count').textContent = waiting.length;
    document.getElementById('error_count').textContent = errors.length;
    document.getElementById('waiting').innerHTML = render(waiting);
    document.getElementById('errors').innerHTML = render(errors,'error');
    document.getElementById('preButton').disabled = state.running;
  }
  async function startPreLaunch(){
    const senha=document.getElementById('preSenha').value.trim();
    const raw=document.getElementById('preLimit').value;
    const limit=raw==='all'?null:Number(raw);
    if(!confirm('Preencher e validar sem confirmar a evolução no Salus?')) return;
    const response=await fetch('/api/etapa2',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({only_password:senha,batch_limit:limit,confirmar:false})});
    if(!response.ok) alert(await response.text());
    await poll();
  }
  setInterval(poll,1200); poll();
</script></body></html>"""

CONTROL_HTML = r"""<!doctype html>
<html lang="pt-BR"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Controle diário - Robo Sallus</title><style>
*{box-sizing:border-box}body{margin:0;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;color:#1f2937;background:#f4f6f8}main{max-width:1320px;margin:0 auto;padding:24px}.head{display:flex;justify-content:space-between;align-items:center;gap:16px;margin-bottom:18px}h1{margin:0;font-size:30px}a{color:#1f4e78;font-weight:800;text-decoration:none}.cards{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:16px}.card{background:#fff;border:1px solid #d8dee6;border-radius:8px;padding:12px}.card b{font-size:26px;display:block;margin-top:4px}.filters{display:flex;gap:10px;flex-wrap:wrap;background:#fff;border:1px solid #d8dee6;border-radius:8px;padding:12px;margin-bottom:12px}input,select{height:38px;border:1px solid #d8dee6;border-radius:6px;padding:0 10px;background:#fff}.table{background:#fff;border:1px solid #d8dee6;border-radius:8px;overflow:hidden}.row{display:grid;grid-template-columns:minmax(230px,1.4fr) 110px 90px 130px 150px minmax(230px,1.4fr);gap:10px;align-items:center;padding:10px 12px;border-top:1px solid #e5e7eb;font-size:13px}.row:first-child{border-top:0}.header{font-weight:800;background:#eef2f6}.name{font-weight:800}.row.vermelho{border-left:7px solid #dc2626;background:#fef2f2}.row.amarelo{border-left:7px solid #d97706;background:#fffbeb}.row.azul{border-left:7px solid #2563eb;background:#eff6ff}.row.verde{border-left:7px solid #16a34a;background:#f0fdf4}.row.cinza{border-left:7px solid #6b7280;background:#f3f4f6;color:#6b7280}.legend{font-size:13px;color:#667085;margin:0 0 12px}@media(max-width:850px){.cards{grid-template-columns:repeat(2,1fr)}.row{grid-template-columns:1fr 1fr}}
</style></head><body><main>
<div class="head"><h1>Controle diário <span id="date"></span></h1><div><a href="/">← Lançamentos</a> &nbsp; <a href="/pre-lancamentos">Pré-lançamentos →</a></div></div>
<p class="legend">Vermelho: evolução de hoje ausente · Amarelo: aguardando revisão · Azul: pré-lançado · Verde: confirmado hoje · Cinza: fora da fila atual.</p>
<div class="cards"><div class="card">Sem evolução hoje<b id="c_vermelho">0</b></div><div class="card">Aguardando revisão<b id="c_amarelo">0</b></div><div class="card">Pré-lançados<b id="c_azul">0</b></div><div class="card">Confirmados hoje<b id="c_verde">0</b></div><div class="card">Fora da fila<b id="c_cinza">0</b></div></div>
<div class="filters"><input id="search" placeholder="Buscar nome ou senha" oninput="render()"><select id="filter" onchange="render()"><option value="">Todos os status</option><option value="vermelho">Vermelho</option><option value="amarelo">Amarelo</option><option value="azul">Azul</option><option value="verde">Verde</option><option value="cinza">Cinza</option></select></div>
<div class="table"><div class="row header"><div>Paciente</div><div>Senha</div><div>Iniciais</div><div>Data evolução</div><div>Responsável</div><div>Status</div></div><div id="rows"></div></div>
</main><script>
let data={rows:[],counts:{}};const esc=v=>String(v??'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#039;'}[c]));
function render(){const q=document.getElementById('search').value.toLowerCase();const f=document.getElementById('filter').value;const rows=data.rows.filter(r=>(!f||r.cor===f)&&(!q||(r.nome+' '+r.senha).toLowerCase().includes(q)));document.getElementById('rows').innerHTML=rows.length?rows.map(r=>`<div class="row ${r.cor}"><div class="name">${esc(r.nome)}</div><div>${esc(r.senha)}</div><div>${esc(r.iniciais)}</div><div>${esc(r.data_evolucao)}</div><div>${esc(r.responsavel)}</div><div>${esc(r.status)}</div></div>`).join(''):'<div style="padding:18px;color:#667085">Nenhum paciente neste filtro.</div>'}
async function poll(){data=await fetch('/api/controle-diario').then(r=>r.json());document.getElementById('date').textContent='- '+data.data;for(const c of ['vermelho','amarelo','azul','verde','cinza'])document.getElementById('c_'+c).textContent=data.counts[c]||0;render()}setInterval(poll,2500);poll();
</script></body></html>"""


class Handler(BaseHTTPRequestHandler):
    def _send_json(self, payload: dict, status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def do_GET(self) -> None:
        path = urlparse(self.path).path
        if path == "/":
            body = HTML.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        if path == "/pre-lancamentos":
            body = PRE_HTML.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        if path == "/controle-diario":
            body = CONTROL_HTML.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        if path == "/api/controle-diario":
            self._send_json(calculate_daily_control())
            return
        if path == "/api/status":
            with LOCK:
                payload = json.loads(json.dumps(STATE, ensure_ascii=False))
            self._send_json(payload)
            return
        self._send_json({"error": "Nao encontrado"}, 404)

    def do_POST(self) -> None:
        path = urlparse(self.path).path
        try:
            if path == "/api/files":
                data = self._read_json()
                with LOCK:
                    for key in ("fila", "clinica", "relatorio"):
                        if data.get(key):
                            STATE["files"][key] = data[key]
                self._send_json({"ok": True})
                return

            if path == "/api/refresh":
                cards = refresh_cards()
                self._send_json({"ok": True, "cards": cards})
                return

            if path == "/api/novo-dia":
                with LOCK:
                    if STATE["running"]:
                        self._send_json({"error": "Ja existe uma etapa em execucao."}, 409)
                        return
                    STATE["running"] = True
                    STATE["processed_now"] = 0
                    STATE["status"] = "Iniciando novo dia"
                threading.Thread(target=run_new_day_worker, daemon=True).start()
                self._send_json({"ok": True})
                return

            if path == "/api/etapa1":
                with LOCK:
                    if STATE["running"]:
                        self._send_json({"error": "Ja existe uma etapa em execucao."}, 409)
                        return
                    STATE["running"] = True
                    STATE["processed_now"] = 0
                    STATE["status"] = "Executando Etapa 1"
                threading.Thread(target=run_etapa1_worker, daemon=True).start()
                self._send_json({"ok": True})
                return

            if path == "/api/etapa2":
                data = self._read_json()
                with LOCK:
                    if STATE["running"]:
                        self._send_json({"error": "Ja existe uma etapa em execucao."}, 409)
                        return
                    STATE["running"] = True
                    STATE["processed_now"] = 0
                    STATE["status"] = "Executando Etapa 2"
                    STATE["current_patient"] = "Iniciando..."
                    STATE["current_password"] = "-"
                    STATE["launch_rows"] = []
                threading.Thread(
                    target=run_etapa2_worker,
                    args=(
                        data.get("only_password") or None,
                        int(data["batch_limit"]) if data.get("batch_limit") is not None else None,
                        bool(data.get("retry_errors")),
                        data.get("confirmar", True) is not False,
                    ),
                    daemon=True,
                ).start()
                self._send_json({"ok": True})
                return

            if path == "/api/limpar-salus":
                with LOCK:
                    if STATE["running"]:
                        self._send_json({"error": "Aguarde o lançamento atual terminar antes de limpar a tela."}, 409)
                        return
                navigate_salus("https://salus.orizon.com.br/salus/gestao-internacao")
                with LOCK:
                    STATE["current_patient"] = "Nenhum paciente em execução"
                    STATE["current_password"] = "-"
                    STATE["launch_rows"] = []
                    STATE["status"] = "Tela do Salus limpa"
                log("Tela do Salus limpa; retorno à lista de internações.")
                self._send_json({"ok": True})
                return

            self._send_json({"error": "Nao encontrado"}, 404)
        except Exception as exc:
            self._send_json({"error": str(exc)}, 500)

    def log_message(self, format: str, *args) -> None:
        return


def main() -> int:
    EXPORTS.mkdir(parents=True, exist_ok=True)
    try:
        refresh_cards()
    except Exception as exc:
        log(f"Aviso ao carregar cards: {exc}")

    server = ThreadingHTTPServer((HOST, PORT), Handler)
    url = f"http://{HOST}:{PORT}"
    print(f"Robo Sallus aberto em {url}")
    print("Mantenha esta janela aberta enquanto usa a tela.")
    webbrowser.open(url)

    def open_salus_after_interface() -> None:
        try:
            time.sleep(1.0)
            start_salus_chrome()
            log("Portal Salus aberto no Chrome do robo.")
        except SalusCdpError as exc:
            log(f"ATENCAO: {exc}")

    threading.Thread(target=open_salus_after_interface, daemon=True).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nEncerrando Robo Sallus.")
    finally:
        server.server_close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
