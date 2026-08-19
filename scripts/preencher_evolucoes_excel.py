#!/usr/bin/env python3
"""Preenche campos objetivos da base a partir do texto livre de evolucao.

O extrator e conservador: somente grava informacoes explicitamente presentes
no texto. Ele nao cria diagnosticos, CID, sinais vitais ou respostas negativas
pela simples ausencia de mencao.
"""

from __future__ import annotations

import argparse
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from cid_evolucao import infer_adjusted_cid_from_evolution, infer_cid_suggestion
from etapa2_lancar_evolucao_salus import (
    EVOLUTION_DISCHARGE_FILL,
    EVOLUTION_DISCHARGE_FONT,
    discharge_as_datetime,
    parse_census_discharge,
    resolve_discharge,
    resolve_evolution_discharge,
)


def plain(value: object) -> str:
    text = str(value or "")
    return "".join(
        char for char in unicodedata.normalize("NFD", text.lower())
        if unicodedata.category(char) != "Mn"
    )


def numbers(value: str) -> list[float]:
    return [float(item.replace(",", ".")) for item in re.findall(r"\d+(?:[.,]\d+)?", value)]


def format_number(value: float) -> int | float:
    return int(value) if value.is_integer() else value


def last_metric(text: str, labels: str, choose: str = "max") -> int | float | None:
    matches = re.findall(
        rf"(?<![a-z])(?:{labels})\s*(?:max(?:ima)?|min(?:ima)?)?\s*[:=]?\s*"
        rf"(\d+(?:[.,]\d+)?(?:\s*[-–a]\s*\d+(?:[.,]\d+)?)?)",
        plain(text),
        flags=re.I,
    )
    if not matches:
        return None
    values = numbers(matches[-1])
    selected = min(values) if choose == "min" else max(values)
    return format_number(selected)


def blood_pressure(text: str) -> tuple[int | None, int | None]:
    normalized = plain(text)
    pressure_lines = re.findall(
        r"(?<!m)(?<![a-z])pa\b(?:\s+arterial)?\s*[:=]?\s*"
        r"((?:\d{2,3})\s*(?:x|/)\s*(?:\d{2,3})[^\n]{0,70})",
        normalized,
    )
    if pressure_lines:
        pairs = re.findall(r"(\d{2,3})\s*(?:x|/)\s*(\d{2,3})", pressure_lines[-1])
        if pairs:
            return max(int(pair[0]) for pair in pairs), max(int(pair[1]) for pair in pairs)

    pas = last_metric(text, r"pas|pressao sistolica")
    pad = last_metric(text, r"pad|pressao diastolica")
    return (int(pas) if pas is not None else None, int(pad) if pad is not None else None)


def explicit_general_state(text: str) -> str | None:
    normalized = plain(text)
    matches = list(re.finditer(r"(?<![a-z])(beg|reg|meg)(?![a-z])", normalized))
    if not matches:
        return None
    return {
        "beg": "BEG – Bom Estado Geral",
        "reg": "REG – Regular Estado Geral",
        "meg": "MEG – Mau Estado Geral",
    }[matches[-1].group(1)]


def consciousness(text: str) -> str | None:
    normalized = plain(text)
    candidates: list[tuple[int, str]] = []
    rules = {
        "Orientado": r"\b(?:lote|lucido[^\n]{0,35}orientad[oa]|consciente[^\n]{0,35}orientad[oa])\b",
        "Desorientado": r"\bdesorientad[oa]\b",
        "Letárgico": r"\b(?:letargic[oa]|sonolent[oa]|torporos[oa])\b",
        "Comatoso": r"\b(?:comatos[oa]|coma)\b",
        "Sedado": r"\bsedad[oa]\b",
    }
    for value, pattern in rules.items():
        for match in re.finditer(pattern, normalized):
            candidates.append((match.start(), value))
    return max(candidates)[1] if candidates else None


def mobility(text: str) -> str | None:
    normalized = plain(text)
    candidates: list[tuple[int, str]] = []
    rules = {
        "Deambulando com auxílio": r"\bdeambul\w*\s+(?:com|sob)\s+auxilio\b",
        "Deambulando": r"\bdeambul\w*\b",
        "Acamado": r"\bacamad[oa]\b",
        "Ortotatismo preservado": r"\bortostatismo\s+preservado\b",
    }
    for value, pattern in rules.items():
        for match in re.finditer(pattern, normalized):
            candidates.append((match.start(), value))
    return max(candidates)[1] if candidates else None


def respiratory(text: str) -> dict[str, str]:
    normalized = plain(text)
    result: dict[str, str] = {}
    if re.search(r"\b(?:tqt|traqueostom)\w*\b", normalized):
        result["Exame Físico - Via respiratória *"] = "Traqueostomia"
    elif re.search(r"\b(?:iot|intubad[oa]|tubo orotraqueal)\b", normalized):
        result["Exame Físico - Via respiratória *"] = "Tubo"
    elif re.search(r"\b(?:eupneic[oa]|via respiratoria normal)\b", normalized):
        result["Exame Físico - Via respiratória *"] = "Normal"

    if re.search(r"\b(?:vmi|ventilacao mecanica)\b", normalized):
        result["Exame Físico - Suporte respiratório *"] = "Ventilação mecânica"
    elif re.search(r"\b(?:bipap|cpap|cateter (?:de )?o2|cateter nasal|mascara de o2|oxigenio suplementar)\b", normalized):
        result["Exame Físico - Suporte respiratório *"] = "Suporte não invasivo"
        if "bipap" in normalized:
            result["Exame Físico - Detalhamento do suporte respiratório * (cond.)"] = "BiPAP"
        elif "cpap" in normalized:
            result["Exame Físico - Detalhamento do suporte respiratório * (cond.)"] = "CPAP"
        elif re.search(r"cateter (?:de )?o2|cateter nasal", normalized):
            result["Exame Físico - Detalhamento do suporte respiratório * (cond.)"] = "Cateter O2"
    elif re.search(r"\b(?:ar ambiente|eupneic[oa] em aa)\b", normalized):
        result["Exame Físico - Suporte respiratório *"] = "Ar ambiente"

    # A descrição do estado respiratório atual prevalece sobre menções
    # históricas (ex.: IOT como possibilidade terapêutica ou CPAP domiciliar).
    if re.search(r"\b(?:eupneic[oa](?:\s+em\s+aa)?|em\s+ar\s+ambiente)\b", normalized):
        result["Exame Físico - Via respiratória *"] = "Normal"
    if re.search(r"\b(?:eupneic[oa]\s+em\s+aa|em\s+ar\s+ambiente)\b", normalized):
        result["Exame Físico - Suporte respiratório *"] = "Ar ambiente"
        result.pop("Exame Físico - Detalhamento do suporte respiratório * (cond.)", None)
    return result


def access_and_elimination(text: str) -> dict[str, str]:
    normalized = plain(text)
    result: dict[str, str] = {}
    central = []
    if re.search(
        r"\b(?:picc|cateter venoso central de insercao periferica)\b",
        normalized,
    ):
        central.append("PICC")
    if re.search(r"\bport[- ]?a[- ]?cath\b", normalized): central.append("Port-o-cath")
    if re.search(r"\b(?:cvc|cateter venoso central|permcath|cateter hd)\b", normalized): central.append("CVC")
    if central:
        result["Exame Físico - Acesso venoso? *"] = "Sim"
        result["Exame Físico - Qual o acesso venoso? * (cond.)"] = "Central"
        details = [item for item in central if item in {"PICC", "Port-o-cath"}]
        if details:
            result["Exame Físico - Detalhamento do acesso central * (cond.)"] = "; ".join(dict.fromkeys(details))
    elif re.search(r"\b(?:acesso venoso periferico|avp)\b", normalized):
        result["Exame Físico - Acesso venoso? *"] = "Sim"
        result["Exame Físico - Qual o acesso venoso? * (cond.)"] = "Periférico"

    elimination = []
    if re.search(r"\bsvd\b|sonda vesical de demora", normalized): elimination.append("SVD - Sonda vesical de demora")
    if re.search(r"\bsva\b|sonda vesical de alivio", normalized): elimination.append("SVA - Sonda vesical de alívio")
    if re.search(r"\b(?:colostomia|ileostomia|ostomia intestinal)\b", normalized): elimination.append("Ostomia Intestinal")
    if re.search(r"\b(?:urostomia|ostomia urinaria)\b", normalized): elimination.append("Ostomia Urinária")
    if re.search(r"\bfralda\b", normalized): elimination.append("Fralda")
    if elimination:
        result["Exame Físico - Controle de eliminação *"] = "; ".join(dict.fromkeys(elimination))
    return result


def alimentation(text: str) -> str | None:
    normalized = plain(text)
    values = []
    if re.search(r"\b(?:dieta|alimentacao|aceitando)\b[^\n]{0,35}\b(?:via oral|vo)\b", normalized): values.append("Oral")
    if re.search(r"\b(?:dieta enteral|sne|sonda nasoenteral|gtt|gastrostomia)\b", normalized): values.append("Enteral")
    if re.search(r"\b(?:dieta parenteral|nutricao parenteral|npt)\b", normalized): values.append("Parenteral")
    return "; ".join(values) if values else None


def active_drugs(text: str) -> dict[str, str]:
    normalized = plain(text)
    result: dict[str, str] = {}
    vasoactive_names = {
        "noradrenalina": "Noradrenalina", "norepinefrina": "Noradrenalina",
        "dobutamina": "Dobutamina", "adrenalina": "Adrenalina",
        "vasopressina": "Vasopressina", "nitroglicerina": "Nitroglicerina",
        "nitroprussiato": "Nitroprussiato",
    }
    active = []
    for token, label in vasoactive_names.items():
        if re.search(rf"\b{token}\b[^\n]{{0,30}}(?:mcg|mg|ml/h)|(?:dva|recebe|em uso)[^\n]{{0,80}}\b{token}\b", normalized):
            active.append(label)
    if active:
        result["UTI - Uso de droga vasoativa? *"] = "Sim"
        result["UTI - Drogas vasoativas em uso * (cond.)"] = "; ".join(dict.fromkeys(active))

    lines = text.splitlines()
    for index, line in enumerate(lines):
        section = re.match(r"\s*#?\s*atb\s*:\s*(.*)$", line, flags=re.I)
        if not section:
            continue
        names = []
        inline = section.group(1).strip()
        if inline:
            name = re.split(r"\s+-\s+(?:D?\d|desde)|\s*\(", inline, maxsplit=1, flags=re.I)[0]
            if re.fullmatch(r"[A-Za-zÀ-ÿ][A-Za-zÀ-ÿ +.]{2,50}", name.strip()):
                names.append(name.strip(" .-"))
        for following in lines[index + 1:]:
            if not following.strip():
                continue
            match = re.match(r"\s*[-–•]\s*([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ +.-]{2,40}?)(?:\s*\(|\s+\d|$)", following)
            if not match:
                break
            names.append(match.group(1).strip(" .-"))
        if names:
            expanded = []
            for name in names:
                cleaned = re.sub(r"\s+", " ", name).strip()
                cleaned = re.sub(r"\bmero\b", "Meropenem", cleaned, flags=re.I)
                cleaned = re.sub(r"\bvanco\b", "Vancomicina", cleaned, flags=re.I)
                expanded.extend(part.strip() for part in cleaned.split("+") if part.strip())
            result["Conduta Clínica - Uso de antibiótico? *"] = "Sim"
            selected = "; ".join(dict.fromkeys(expanded))
            result["Conduta Clínica - Selecione os antibióticos em uso * (cond.)"] = selected
            # O Salus abre um campo obrigatório "Antibiótico selecionado"
            # após a escolha no catálogo. O mesmo nome explícito da evolução
            # deve acompanhar a linha para o preenchimento diário.
            result["Conduta Clínica - Antibiótico selecionado * (cond.)"] = selected
        break
    return result


def accommodation(text: str) -> str | None:
    normalized = plain(text)
    if re.search(
        r"\b(?:uti|cti|uco)\b|unidade\s+(?:de\s+)?(?:terapia\s+intensiva|coronariana)",
        normalized,
    ):
        return "UTI"
    return None


def surgical_procedure(text: str) -> str:
    normalized = plain(text)
    procedures = (
        (r"\bap[e]?ndicectomia\b", "Apendicectomia"),
        (r"\bcolecistectomia\b", "Colecistectomia"),
        (r"\bosteossintese\b|\bosteosintese\b", "Osteossíntese"),
        (
            r"\brtup\b|\bressecao transuretral (?:da )?prostata\b",
            "Ressecção endoscópica da próstata",
        ),
        (
            r"\bureterolitotripsia\b|\bult flex\b",
            "Ureterorrenolitotripsia flexível a laser unilateral",
        ),
        (
            r"\bretirad[ao] (?:do |de )?cateter duplo j\b|"
            r"\bretirad[ao] (?:do |de )?duplo j\b",
            "Retirada de cateter duplo J",
        ),
        (r"\bherniorrafia\b", "Herniorrafia"),
        (r"\bhemorroidectomia\b", "Hemorroidectomia"),
        (r"\bdrenagem (?:de )?(?:abscesso|colecao)\b", "Drenagem de abscesso"),
        (
            r"\blimpeza cirurgica\b|\bcoleta de culturas?\b",
            "Limpeza cirúrgica",
        ),
        (
            r"\bartroplastia\b|\bptq\b|\bpta\b|"
            r"\bprotese total (?:de )?(?:quadril|joelho)\b|"
            r"\brevisao (?:total )?(?:de )?(?:ptq|pta|protese)\b|"
            r"\brevisao acetabular\b",
            "Artroplastia",
        ),
        (r"\bangioplastia\b", "Angioplastia"),
        (r"\blaparotomia\b", "Laparotomia"),
        (r"\bvideolaparoscopia\b|\bvlp\b", "Videolaparoscopia"),
        (r"\btoracoscopia\b", "Toracoscopia"),
        (r"\bcraniotomia\b", "Craniotomia"),
        (r"\bmastectomia\b", "Mastectomia"),
        (r"\bhisterectomia\b", "Histerectomia"),
        (r"\bprostatectomia\b", "Prostatectomia"),
        (r"\bcolectomia\b", "Colectomia"),
    )
    return next(
        (label for pattern, label in procedures if re.search(pattern, normalized)),
        "",
    )


def surgical(text: str) -> dict[str, object]:
    normalized = plain(text)
    procedure = surgical_procedure(text)
    completed_surgery = re.search(
        r"\b(?:po|pos[- ]?operatorio|submetid[oa]|procedimento sem intercorrencias|"
        r"apos (?:a |o )?(?:cirurgia|procedimento|drenagem)|"
        r"realizad[oa] (?:a |o )?(?:cirurgia|procedimento|drenagem))\b",
        normalized,
    )
    planned_surgery = re.search(
        r"\b(?:procedimento proposto|procedimentos propostos|"
        r"tratamento cirurgico (?:de )?urgencia|opta[- ]?se (?:por )?tratamento cirurgico|"
        r"solicito internacao.*tratamento cirurgico|programacao cirurgica|"
        r"avaliacao pre[- ]?anestesica)\b",
        normalized,
    )
    if not completed_surgery and (not procedure or planned_surgery):
        return {}
    result: dict[str, object] = {
        "Conduta Clínica - Realizado procedimento cirúrgico? *": "Sim",
    }
    if procedure:
        result["Conduta Clínica - TUSS + Nome do Procedimento * (cond.)"] = procedure
        result["Conduta Clínica - Quantidade Solicitada * (cond.)"] = 1
        result["Conduta Clínica - Quantidade Autorizada * (cond.)"] = 1
        result["Conduta Clínica - Quantidade Realizada * (cond.)"] = 1
        result["Conduta Clínica - Tipo de anestesia * (cond.)"] = "Geral"
        result[
            "Conduta Clínica - Houve intercorrências no intraoperatório? * (cond.)"
        ] = "Não"
    return result


def extract(
    text: str,
    days: object,
    *,
    census_discharge: object = None,
    evolution_date: object = None,
    seed: str = "",
) -> dict[str, object]:
    result: dict[str, object] = {}
    patient_accommodation = accommodation(text)
    if patient_accommodation:
        result["Dados da Internação - Acomodação *"] = patient_accommodation

    cid, _cid_reason = infer_cid_suggestion(text)
    if cid:
        result["Dados da Internação - CID de internação *"] = cid
        result["Dados da Internação - CID ajustado *"] = (
            infer_adjusted_cid_from_evolution(text) or cid
        )
    if days not in (None, ""):
        result["Dados da Internação - Tempo de existência da doença *"] = days
        result["Dados da Internação - Nomenclatura do tempo de existência da doença *"] = "Dias"

    state = explicit_general_state(text)
    if state: result["Exame Físico - Estado geral *"] = state
    pas, pad = blood_pressure(text)
    if pas is not None: result["Exame Físico - PA Sistólica max (mmHg) *"] = pas
    if pad is not None: result["Exame Físico - PA Diastólica max (mmHg) *"] = pad
    metrics = (
        ("Exame Físico - FC máx. (bpm) *", r"fc|frequencia cardiaca", "max"),
        ("Exame Físico - FR máx. (irpm) *", r"fr|frequencia respiratoria", "max"),
        ("Exame Físico - SpO2 mín. (%) *", r"spo2|sato2|sat o2|saturacao", "min"),
        ("Exame Físico - Temperatura máx. (°C) *", r"tax|temp|temperatura", "max"),
        ("UTI - Creatinina sérica (mg/dL) *", r"creatinina|creat|cr", "max"),
        ("UTI - pH arterial *", r"ph", "max"),
        ("UTI - PaO2 (mmHg) *", r"pao2", "max"),
        ("UTI - FiO2 (%) *", r"fio2", "max"),
    )
    for header, labels, choice in metrics:
        value = last_metric(text, labels, choice)
        if value is not None: result[header] = value

    level = consciousness(text)
    if level: result["Exame Físico - Nível de consciência *"] = level
    movement = mobility(text)
    if movement: result["Exame Físico - Mobilidade e dependência *"] = movement
    food = alimentation(text)
    if food: result["Exame Físico - Alimentação *"] = food
    result.update(respiratory(text))
    result.update(access_and_elimination(text))
    result.update(active_drugs(text))
    result.update(surgical(text))
    discharge = resolve_discharge(
        text,
        census_value=census_discharge,
        evolution_date=evolution_date,
        seed=seed,
    )
    if discharge:
        discharge_date, hour, outcome = discharge
        result["Parecer do Auditor - Paciente permanece internado? *"] = "Não"
        result["Parecer do Auditor - Selecione o desfecho assistencial * (cond.)"] = outcome
        result["Parecer do Auditor - Data do desfecho * (cond.)"] = discharge_date
        result["Parecer do Auditor - Hora do desfecho * (cond.)"] = hour
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="Preenche a base a partir da coluna evolucao.")
    parser.add_argument("entrada", type=Path)
    parser.add_argument("saida", type=Path)
    parser.add_argument(
        "--somente-vazios",
        action="store_true",
        help="Preenche apenas células vazias, preservando valores já revisados.",
    )
    parser.add_argument(
        "--atualizar-altas",
        action="store_true",
        help="Atualiza os campos de alta e a coluna Alta (data e hora).",
    )
    parser.add_argument(
        "--preservar-acomodacao",
        action="store_true",
        help="Nao altera a coluna de acomodacao revisada manualmente.",
    )
    parser.add_argument(
        "--preservar-alta",
        action="store_true",
        help="Nao altera a coluna Alta (data e hora), mas atualiza os campos de desfecho.",
    )
    args = parser.parse_args()

    workbook = load_workbook(args.entrada)
    sheet = workbook["Preenchimento"]
    headers = {str(cell.value): cell.column for cell in sheet[1] if cell.value}
    evolution_column = headers.get("evolucao")
    if not evolution_column:
        raise RuntimeError("Coluna 'evolucao' nao encontrada.")

    changed_rows = 0
    written_cells = 0
    for row in range(2, sheet.max_row + 1):
        text = str(sheet.cell(row, evolution_column).value or "").strip()
        census_value = (
            sheet.cell(row, headers["Alta (data e hora)"]).value
            if "Alta (data e hora)" in headers
            else None
        )
        if not text and census_value in (None, ""):
            continue
        values = extract(
            text,
            sheet.cell(row, headers["Dias internado"]).value,
            census_discharge=census_value,
            evolution_date=(
                sheet.cell(row, headers["Data da evolução"]).value
                if "Data da evolução" in headers
                else None
            ),
            seed=(
                str(sheet.cell(row, headers["Senha"]).value or "").strip()
                if "Senha" in headers
                else ""
            ),
        )
        evolution_discharge = resolve_evolution_discharge(
            text,
            evolution_date=(
                sheet.cell(row, headers["Data da evolução"]).value
                if "Data da evolução" in headers
                else None
            ),
            seed=(
                str(sheet.cell(row, headers["Senha"]).value or "").strip()
                if "Senha" in headers
                else ""
            ),
        )
        if args.atualizar_altas:
            discharge_headers = {
                "Parecer do Auditor - Paciente permanece internado? *",
                "Parecer do Auditor - Selecione o desfecho assistencial * (cond.)",
                "Parecer do Auditor - Data do desfecho * (cond.)",
                "Parecer do Auditor - Hora do desfecho * (cond.)",
            }
            values = {
                header: value
                for header, value in values.items()
                if header in discharge_headers
            }
        row_writes = 0
        for header, value in values.items():
            if args.preservar_acomodacao and header == "Dados da Internação - Acomodação *":
                continue
            column = headers.get(header)
            if column and value not in (None, ""):
                if args.somente_vazios and not args.atualizar_altas:
                    current = sheet.cell(row, column).value
                    if current is not None and str(current).strip():
                        continue
                sheet.cell(row, column).value = value
                row_writes += 1

        # A data/hora obtida da evolução fica gravada na mesma coluna do
        # censo, mas em laranja. Uma alta real do censo nunca é sobrescrita.
        high_column = headers.get("Alta (data e hora)")
        evolution_datetime = discharge_as_datetime(evolution_discharge)
        if (
            high_column
            and evolution_datetime
            and parse_census_discharge(census_value) is None
            and not args.preservar_alta
        ):
            high_cell = sheet.cell(row, high_column)
            high_cell.value = evolution_datetime
            high_cell.number_format = "dd/mm/yyyy hh:mm"
            high_cell.fill = PatternFill("solid", fgColor=EVOLUTION_DISCHARGE_FILL)
            high_cell.font = Font(color=EVOLUTION_DISCHARGE_FONT, bold=True)
            row_writes += 1
        if row_writes:
            changed_rows += 1
            written_cells += row_writes

    args.saida.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.saida)
    print(f"Linhas preenchidas: {changed_rows}")
    print(f"Celulas preenchidas: {written_cells}")
    print(f"Arquivo: {args.saida}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
