#!/usr/bin/env python3
"""Colore a planilha de preenchimento da evolucao clinica por bloco/secao."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SECTION_COLORS = {
    "Identificacao": ("666666", "E7E6E6"),
    "Dados da Internação": ("1F4E78", "D9EAF7"),
    "Exame Físico": ("548235", "E2F0D9"),
    "Conduta Clínica": ("C65911", "FCE4D6"),
    "Condição Adquirida": ("A61C00", "F4CCCC"),
    "UTI": ("7030A0", "EADCF8"),
    "Parecer do Auditor": ("008C95", "DDEBF7"),
    "Resumo": ("8064A2", "E4DFEC"),
}

DEFAULT_COLORS = ("404040", "F2F2F2")
PATIENT_COLUMNS = {"Nome", "Nome paciente", "Iniciais", "Senha", "Dias internado", "ID internação"}


def section_map_from_campos(wb) -> dict[str, str]:
    if "Campos" not in wb.sheetnames:
        return {}
    ws = wb["Campos"]
    mapping: dict[str, str] = {}
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    try:
        coluna_idx = headers.index("Coluna") + 1
        secao_idx = headers.index("Seção") + 1
    except ValueError:
        return {}
    for row in range(2, ws.max_row + 1):
        column_name = ws.cell(row, coluna_idx).value
        section_name = ws.cell(row, secao_idx).value
        if column_name and section_name:
            mapping[str(column_name)] = str(section_name)
    return mapping


def apply_colors(workbook_path: Path, output_path: Path) -> None:
    wb = load_workbook(workbook_path)
    if "Preenchimento" not in wb.sheetnames:
        raise RuntimeError("Aba 'Preenchimento' nao encontrada.")

    ws = wb["Preenchimento"]
    section_by_column = section_map_from_campos(wb)

    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col).value or "")
        section = "Identificacao" if header in PATIENT_COLUMNS else section_by_column.get(header, "")
        dark, light = SECTION_COLORS.get(section, DEFAULT_COLORS)

        header_cell = ws.cell(1, col)
        header_cell.fill = PatternFill("solid", fgColor=dark)
        header_cell.font = Font(color="FFFFFF", bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_cell.border = border

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=light)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

        ws.column_dimensions[get_column_letter(col)].width = 24 if col > 5 else 18

    for col in range(1, min(ws.max_column, 5) + 1):
        ws.column_dimensions[get_column_letter(col)].width = [32, 12, 14, 16, 14][col - 1]

    ws.row_dimensions[1].height = 48
    ws.freeze_panes = "F2"
    ws.auto_filter.ref = ws.dimensions

    if "Resumo" in wb.sheetnames:
        resumo = wb["Resumo"]
        start_row = 1
        start_col = 4
        resumo.cell(start_row, start_col).value = "Legenda de cores"
        resumo.cell(start_row, start_col).font = Font(bold=True)
        row = start_row + 1
        for section, (dark, light) in SECTION_COLORS.items():
            resumo.cell(row, start_col).value = section
            resumo.cell(row, start_col).fill = PatternFill("solid", fgColor=dark)
            resumo.cell(row, start_col).font = Font(color="FFFFFF", bold=True)
            resumo.cell(row, start_col + 1).value = "Campos deste bloco"
            resumo.cell(row, start_col + 1).fill = PatternFill("solid", fgColor=light)
            row += 1
        resumo.column_dimensions[get_column_letter(start_col)].width = 24
        resumo.column_dimensions[get_column_letter(start_col + 1)].width = 24

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Aplica cores por bloco na planilha de evolucao.")
    parser.add_argument(
        "entrada",
        nargs="?",
        default="exports/preenchimento_evolucao_clinica_LP_K3ZAVM6_2026-07-15.xlsx",
    )
    parser.add_argument(
        "--saida",
        default="exports/preenchimento_evolucao_clinica_LP_K3ZAVM6_2026-07-15_colorido.xlsx",
    )
    args = parser.parse_args()
    apply_colors(Path(args.entrada), Path(args.saida))
    print(f"Arquivo colorido gerado: {args.saida}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
