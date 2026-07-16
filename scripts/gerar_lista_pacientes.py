#!/usr/bin/env python3
"""Exporta a lista de pacientes do hospital selecionado no Salus para Excel."""

from __future__ import annotations

import argparse
import datetime as dt
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from salus_cdp import call_salus_api


def pick(item: dict[str, Any], *names: str) -> Any:
    lowered = {str(k).lower(): v for k, v in item.items()}
    for name in names:
        if name in item:
            return item[name]
        value = lowered.get(name.lower())
        if value is not None:
            return value
    return ""


def extract_rows(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, dict)]
    if not isinstance(payload, dict):
        return []
    for key in ("items", "registros", "internacoes", "data", "dados", "content", "result"):
        value = payload.get(key)
        if isinstance(value, list):
            return [row for row in value if isinstance(row, dict)]
    for value in payload.values():
        if isinstance(value, list) and value and isinstance(value[0], dict):
            return value
    return []


def total_pages(payload: Any, current_count: int, page_size: int) -> int | None:
    if not isinstance(payload, dict):
        return None
    for key in ("totalPaginas", "total_pages", "totalPages", "qtdePaginas"):
        value = payload.get(key)
        if isinstance(value, int):
            return value
    total = None
    for key in ("total", "totalRegistros", "totalElements", "quantidadeTotal"):
        value = payload.get(key)
        if isinstance(value, int):
            total = value
            break
    if total is None:
        return None
    return max(1, (total + page_size - 1) // page_size)


def fetch_patients(user_key: int, prestador: int, empresa_auditoria: int, page_size: int, cdp_url: str) -> list[dict[str, Any]]:
    patients: list[dict[str, Any]] = []
    page = 1
    expected_pages: int | None = None
    while True:
        endpoint = (
            "/api/internacoes?"
            f"user_key={user_key}&IdPrestador={prestador}&IdEmpresaAuditoria={empresa_auditoria}"
            f"&Pagina={page}&TamanhoPagina={page_size}"
        )
        payload = call_salus_api(endpoint, cdp_url=cdp_url)
        rows = extract_rows(payload)
        patients.extend(rows)
        expected_pages = expected_pages or total_pages(payload, len(rows), page_size)
        if expected_pages and page >= expected_pages:
            break
        if len(rows) < page_size:
            break
        page += 1
    return patients


def save_excel(patients: list[dict[str, Any]], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pacientes"
    headers = ["Nome", "Iniciais", "Senha", "Dias internado"]
    ws.append(headers)
    for patient in patients:
        ws.append(
            [
                pick(patient, "nomeCompleto", "Nome", "nomePaciente", "paciente"),
                pick(patient, "nomeIniciais", "Iniciais", "iniciais"),
                pick(patient, "senha", "Senha", "senhaInternacao"),
                pick(patient, "diasInternados", "DiasInternado", "diasInternado", "dias"),
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    widths = [42, 12, 16, 16]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    today = dt.date.today().isoformat()
    parser = argparse.ArgumentParser(description="Gera a planilha de pacientes do Salus.")
    parser.add_argument("--user-key", type=int, default=49)
    parser.add_argument("--prestador", type=int, default=113)
    parser.add_argument("--empresa-auditoria", type=int, default=6)
    parser.add_argument("--tamanho-pagina", type=int, default=500)
    parser.add_argument("--cdp-url", default="http://127.0.0.1:9222")
    parser.add_argument("--saida", default=f"exports/pacientes_sirio_libanes_{today}.xlsx")
    args = parser.parse_args()

    patients = fetch_patients(
        user_key=args.user_key,
        prestador=args.prestador,
        empresa_auditoria=args.empresa_auditoria,
        page_size=args.tamanho_pagina,
        cdp_url=args.cdp_url,
    )
    save_excel(patients, Path(args.saida))
    print(f"Arquivo gerado: {args.saida}")
    print(f"Pacientes: {len(patients)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
