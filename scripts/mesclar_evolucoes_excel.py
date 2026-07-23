#!/usr/bin/env python3
"""Mescla evolucoes de uma base auxiliar na base principal, sem sobrescrever dados.

O pareamento usa primeiro a Senha do atendimento e, na ausencia dela, o nome
normalizado do paciente. Por seguranca, apenas linhas cuja evolucao esta vazia
na base de destino sao preenchidas.
"""

from __future__ import annotations

import argparse
import shutil
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def normalize(value: object) -> str:
    text = str(value or "").strip().upper()
    text = "".join(
        character
        for character in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(character)
    )
    return " ".join(text.split())


def headers(sheet) -> dict[str, int]:
    return {
        normalize(cell.value): cell.column
        for cell in sheet[1]
        if cell.value not in (None, "")
    }


def find_column(columns: dict[str, int], *names: str) -> int | None:
    for name in names:
        column = columns.get(normalize(name))
        if column:
            return column
    return None


def sheet_for(workbook):
    return workbook["Preenchimento"] if "Preenchimento" in workbook.sheetnames else workbook.active


def row_key(sheet, row: int, senha_col: int | None, nome_col: int | None) -> tuple[str, str] | None:
    if senha_col:
        senha = normalize(sheet.cell(row, senha_col).value)
        if senha:
            return "senha", senha
    if nome_col:
        nome = normalize(sheet.cell(row, nome_col).value)
        if nome:
            return "nome", nome
    return None


def main() -> int:
    parser = argparse.ArgumentParser(description="Mescla evolucoes ausentes em uma base principal.")
    parser.add_argument("origem", type=Path)
    parser.add_argument("destino", type=Path)
    parser.add_argument("--aplicar", action="store_true", help="Salva as alteracoes na base de destino.")
    parser.add_argument("--backup", type=Path, help="Caminho da copia de seguranca antes da alteracao.")
    parser.add_argument(
        "--data-evolucao",
        help="Data padrao (DD/MM/AAAA) para evolucoes mescladas cuja data esteja vazia.",
    )
    args = parser.parse_args()

    source_workbook = load_workbook(args.origem, data_only=False)
    target_workbook = load_workbook(args.destino, data_only=False)
    source_sheet = sheet_for(source_workbook)
    target_sheet = sheet_for(target_workbook)
    source_headers = headers(source_sheet)
    target_headers = headers(target_sheet)

    source_evolution_col = find_column(source_headers, "evolucao", "evolução")
    target_evolution_col = find_column(target_headers, "evolucao", "evolução")
    if not source_evolution_col or not target_evolution_col:
        raise RuntimeError("A coluna 'evolucao' nao foi encontrada nas duas bases.")

    source_senha_col = find_column(source_headers, "senha")
    target_senha_col = find_column(target_headers, "senha")
    source_nome_col = find_column(source_headers, "nome", "nome paciente", "paciente")
    target_nome_col = find_column(target_headers, "nome", "nome paciente", "paciente")

    target_rows: dict[tuple[str, str], int] = {}
    duplicate_keys: set[tuple[str, str]] = set()
    for row in range(2, target_sheet.max_row + 1):
        key = row_key(target_sheet, row, target_senha_col, target_nome_col)
        if not key:
            continue
        if key in target_rows:
            duplicate_keys.add(key)
        else:
            target_rows[key] = row

    metadata_names = (
        "Data da evolução",
        "Responsável",
        "Responsavel",
    )
    metadata_pairs: list[tuple[int, int]] = []
    for name in metadata_names:
        source_col = find_column(source_headers, name)
        target_col = find_column(target_headers, name)
        if source_col and target_col and (source_col, target_col) not in metadata_pairs:
            metadata_pairs.append((source_col, target_col))

    source_with_evolution = 0
    additions: list[tuple[int, int, tuple[str, str]]] = []
    conflicts: list[tuple[int, int, tuple[str, str]]] = []
    unmatched: list[tuple[int, tuple[str, str]]] = []
    matched_rows: list[tuple[int, int, tuple[str, str]]] = []
    seen_source_keys: set[tuple[str, str]] = set()

    for source_row in range(2, source_sheet.max_row + 1):
        evolution = source_sheet.cell(source_row, source_evolution_col).value
        if evolution is None or not str(evolution).strip():
            continue
        source_with_evolution += 1
        key = row_key(source_sheet, source_row, source_senha_col, source_nome_col)
        if not key:
            continue
        if key in seen_source_keys or key in duplicate_keys:
            continue
        seen_source_keys.add(key)
        target_row = target_rows.get(key)
        if not target_row:
            unmatched.append((source_row, key))
            continue
        matched_rows.append((source_row, target_row, key))
        current = target_sheet.cell(target_row, target_evolution_col).value
        if current is None or not str(current).strip():
            additions.append((source_row, target_row, key))
        elif str(current).strip() != str(evolution).strip():
            conflicts.append((source_row, target_row, key))

    print(f"Evolucoes preenchidas na origem: {source_with_evolution}")
    print(f"Novas evolucoes a acrescentar: {len(additions)}")
    print(f"Conflitos preservados (destino ja preenchido): {len(conflicts)}")
    print(f"Pacientes da origem nao encontrados no destino: {len(unmatched)}")
    for source_row, target_row, key in additions:
        print(f"ADICIONAR {key[0]}={key[1]} origem_linha={source_row} destino_linha={target_row}")

    if not args.aplicar:
        print("Modo de conferencia: nenhum arquivo foi alterado.")
        return 0

    if args.backup:
        args.backup.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(args.destino, args.backup)
        print(f"Backup: {args.backup}")

    for source_row, target_row, _key in additions:
        target_sheet.cell(target_row, target_evolution_col).value = source_sheet.cell(
            source_row, source_evolution_col
        ).value
        target_sheet.cell(target_row, target_evolution_col).fill = PatternFill(
            "solid", fgColor="C6EFCE"
        )
        for source_col, target_col in metadata_pairs:
            source_value = source_sheet.cell(source_row, source_col).value
            target_value = target_sheet.cell(target_row, target_col).value
            if source_value not in (None, "") and target_value in (None, ""):
                target_sheet.cell(target_row, target_col).value = source_value

    dated_rows = 0
    styled_rows = 0
    target_date_col = find_column(target_headers, "Data da evolução")
    if args.data_evolucao and target_date_col:
        for _source_row, target_row, _key in matched_rows:
            evolution = target_sheet.cell(target_row, target_evolution_col).value
            current_date = target_sheet.cell(target_row, target_date_col).value
            if evolution not in (None, "") and current_date in (None, ""):
                target_sheet.cell(target_row, target_date_col).value = args.data_evolucao
                dated_rows += 1
            evolution_cell = target_sheet.cell(target_row, target_evolution_col)
            if evolution not in (None, "") and evolution_cell.fill.fgColor.rgb in {
                "FFFFC7CE",
                "00FFC7CE",
            }:
                evolution_cell.fill = PatternFill("solid", fgColor="C6EFCE")
                styled_rows += 1

    target_workbook.save(args.destino)
    print(f"Arquivo atualizado: {args.destino}")
    print(f"Evolucoes acrescentadas: {len(additions)}")
    print(f"Datas de evolucao preenchidas: {dated_rows}")
    print(f"Evolucoes sinalizadas em verde: {styled_rows}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
