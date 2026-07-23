#!/usr/bin/env python3
"""Audita como pacientes/evolucoes foram transportados entre duas bases."""

from __future__ import annotations

import argparse
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


def normalize(value: object) -> str:
    text = str(value or "").strip().upper()
    text = "".join(
        character
        for character in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(character)
    )
    return " ".join(text.split())


def read_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Preenchimento"]
    headers = {normalize(cell.value): cell.column for cell in sheet[1] if cell.value}

    def value(row: int, header: str) -> str:
        column = headers.get(normalize(header))
        return str(sheet.cell(row, column).value or "").strip() if column else ""

    rows = []
    for row in range(2, sheet.max_row + 1):
        senha = value(row, "Senha")
        if not senha:
            continue
        rows.append(
            {
                "row": str(row),
                "nome": value(row, "Nome"),
                "senha": senha,
                "id": value(row, "ID internação"),
                "evolucao": value(row, "evolucao"),
            }
        )
    return rows


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("anterior", type=Path)
    parser.add_argument("atual", type=Path)
    args = parser.parse_args()

    previous = read_rows(args.anterior)
    current = read_rows(args.atual)
    previous_by_admission = {(normalize(row["senha"]), normalize(row["id"])): row for row in previous}
    current_by_admission = {(normalize(row["senha"]), normalize(row["id"])): row for row in current}
    current_by_name = {normalize(row["nome"]): row for row in current if normalize(row["nome"])}

    evolved_previous = [row for row in previous if row["evolucao"]]
    carried = []
    changed_admission = []
    absent = []
    for row in evolved_previous:
        key = (normalize(row["senha"]), normalize(row["id"]))
        if key in current_by_admission:
            carried.append((row, current_by_admission[key]))
        elif normalize(row["nome"]) in current_by_name:
            changed_admission.append((row, current_by_name[normalize(row["nome"])]))
        else:
            absent.append(row)

    empty_current = [row for row in current if not row["evolucao"]]
    empty_with_previous_admission = []
    new_admission = []
    for row in empty_current:
        key = (normalize(row["senha"]), normalize(row["id"]))
        if key in previous_by_admission:
            empty_with_previous_admission.append((row, previous_by_admission[key]))
        else:
            new_admission.append(row)

    print(f"Pacientes na base anterior: {len(previous)}")
    print(f"Evolucoes na base anterior: {len(evolved_previous)}")
    print(f"Pacientes na base atual: {len(current)}")
    print(f"Evolucoes transportadas na mesma internacao: {len(carried)}")
    print(f"Evolucoes com mesmo nome, mas senha/internacao diferente: {len(changed_admission)}")
    print(f"Evolucoes de pacientes ausentes da fila atual: {len(absent)}")
    print(f"Linhas vazias atuais que ja estavam sem evolucao antes: {len(empty_with_previous_admission)}")
    print(f"Linhas vazias atuais sem a mesma internacao na base anterior: {len(new_admission)}")

    for previous_row, current_row in changed_admission:
        print(
            "NOVA_INTERNACAO "
            f"nome={current_row['nome']} "
            f"senha_anterior={previous_row['senha']} id_anterior={previous_row['id']} "
            f"senha_atual={current_row['senha']} id_atual={current_row['id']}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
