#!/usr/bin/env python3
"""Aplica listas suspensas e orientacoes na planilha de preenchimento.

Regra:
- Escolha unica, radio, select e booleano recebem lista suspensa.
- Multipla escolha nao recebe validacao restritiva, porque Excel sem macro nao
  permite selecionar varios itens da lista na mesma celula. Nesses campos, usar
  opcoes separadas por ponto e virgula (;).
"""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation


MULTIPLE_TYPES = {"LISTA_MULTIPLA"}
MULTIPLE_CONTROLS = {"CHECKBOX_MULTI", "MULTISELECT"}


def safe_title(value: str, fallback: str) -> str:
    text = re.sub(r"[^A-Za-z0-9_]+", "_", value.strip())
    text = re.sub(r"_+", "_", text).strip("_")
    return (text or fallback)[:28]


def header_indexes(ws) -> dict[str, int]:
    return {
        str(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(1, col).value is not None
    }


def collect_field_meta(wb) -> dict[str, dict[str, str]]:
    if "Campos" not in wb.sheetnames:
        return {}
    ws = wb["Campos"]
    headers = header_indexes(ws)
    result: dict[str, dict[str, str]] = {}
    for row in range(2, ws.max_row + 1):
        column_name = ws.cell(row, headers.get("Coluna", 1)).value
        if not column_name:
            continue
        result[str(column_name)] = {
            "tipo": str(ws.cell(row, headers.get("Tipo", 1)).value or "").upper(),
            "controle": str(ws.cell(row, headers.get("Controle", 1)).value or "").upper(),
            "secao": str(ws.cell(row, headers.get("Seção", 1)).value or ""),
            "campo": str(ws.cell(row, headers.get("Campo", 1)).value or ""),
        }
    return result


def collect_options(wb) -> dict[str, list[str]]:
    if "Opcoes" not in wb.sheetnames:
        return {}
    ws = wb["Opcoes"]
    headers = header_indexes(ws)
    field_col = headers.get("Campo")
    option_col = headers.get("Opção") or headers.get("Opcao")
    if not field_col or not option_col:
        return {}

    options: dict[str, list[str]] = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        field = ws.cell(row, field_col).value
        option = ws.cell(row, option_col).value
        if not field or option in (None, ""):
            continue
        option = str(option)
        if option not in options[str(field)]:
            options[str(field)].append(option)
    return dict(options)


def get_or_create_lists_sheet(wb):
    if "Listas" in wb.sheetnames:
        ws = wb["Listas"]
        ws.delete_rows(1, ws.max_row or 1)
        return ws
    ws = wb.create_sheet("Listas")
    return ws


def apply_validations(workbook_path: Path, output_path: Path, max_rows: int) -> None:
    wb = load_workbook(workbook_path)
    if "Preenchimento" not in wb.sheetnames:
        raise RuntimeError("Aba 'Preenchimento' nao encontrada.")

    preenchimento = wb["Preenchimento"]
    meta_by_column = collect_field_meta(wb)
    options_by_column = collect_options(wb)
    listas = get_or_create_lists_sheet(wb)
    listas.sheet_state = "hidden"

    list_column = 1
    unique_count = 0
    multiple_count = 0

    for col in range(1, preenchimento.max_column + 1):
        header = str(preenchimento.cell(1, col).value or "")
        options = options_by_column.get(header, [])
        if not options:
            continue

        meta = meta_by_column.get(header, {})
        tipo = meta.get("tipo", "")
        controle = meta.get("controle", "")
        is_multiple = tipo in MULTIPLE_TYPES or controle in MULTIPLE_CONTROLS

        if is_multiple:
            multiple_count += 1
            preenchimento.cell(1, col).comment = Comment(
                "Campo de multipla escolha. Preencher uma ou mais opcoes separadas por ponto e virgula (;). "
                "Consultar a aba Opcoes para ver os valores permitidos.",
                "Robo Sallus",
            )
            continue

        unique_count += 1
        list_header = safe_title(header, f"lista_{list_column}")
        listas.cell(1, list_column).value = list_header
        listas.cell(1, list_column).font = Font(bold=True)
        for row_idx, option in enumerate(options, start=2):
            listas.cell(row_idx, list_column).value = option

        letter = get_column_letter(list_column)
        formula = f"{quote_sheetname(listas.title)}!${letter}$2:${letter}${len(options) + 1}"
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.errorTitle = "Opcao invalida"
        validation.error = "Escolha uma opcao da lista suspensa."
        validation.promptTitle = "Escolha unica"
        validation.prompt = "Selecione uma opcao da lista."
        preenchimento.add_data_validation(validation)
        validation.add(f"{get_column_letter(col)}2:{get_column_letter(col)}{max_rows}")

        list_column += 1

    if "Resumo" in wb.sheetnames:
        resumo = wb["Resumo"]
        start_row = resumo.max_row + 2
        resumo.cell(start_row, 1).value = "Listas suspensas aplicadas"
        resumo.cell(start_row, 2).value = unique_count
        resumo.cell(start_row + 1, 1).value = "Campos multipla escolha com orientacao"
        resumo.cell(start_row + 1, 2).value = multiple_count

    preenchimento.sheet_view.showGridLines = True
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Aplica listas suspensas na planilha clinica.")
    parser.add_argument(
        "entrada",
        nargs="?",
        default="exports/preenchimento_evolucao_clinica_LP_K3ZAVM6_2026-07-15.xlsx",
    )
    parser.add_argument("--saida")
    parser.add_argument("--linhas", type=int, default=500)
    args = parser.parse_args()

    entrada = Path(args.entrada)
    saida = Path(args.saida) if args.saida else entrada
    apply_validations(entrada, saida, args.linhas)
    print(f"Listas aplicadas em: {saida}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
